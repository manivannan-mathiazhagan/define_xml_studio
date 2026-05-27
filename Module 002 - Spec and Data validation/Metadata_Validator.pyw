#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ====================================================================================================
# Script Name    : metadata_validator.pyw
#
# Module Name    : Clinical Metadata Studio – Module 2
#
# Author         : Manivannan Mathialagan
# Created Date   : 13-May-2026
#
# Purpose        :
#                  Interactive metadata validation and specification review tool
#                  for SDTM/ADaM study metadata using SharePoint-based specifications
#                  and SAS/XPT dataset metadata comparison.
#
# Description    :
#                  This tool imports specification workbooks from SharePoint using
#                  the harmonized get_spreadsheet.py utility, loads all sheets,
#                  normalizes metadata, scans SAS/XPT datasets, extracts PROC CONTENTS-
#                  style metadata, and performs specification-versus-dataset validation.
#
# Key Features   :
#                  • SharePoint specification import
#                  • Local specification loading support
#                  • Harmonized SDTM/ADaM metadata normalization
#                  • SAS7BDAT and XPT dataset inventory scanning
#                  • PROC CONTENTS-style metadata extraction
#                  • KEEP flag–based variable validation
#                  • Spec vs Dataset validation checks
#                  • Missing variable detection
#                  • Extra variable detection
#                  • Variable type mismatch checks
#                  • Character length validation
#                  • Label mismatch validation
#                  • SUPPQUAL and SUPP_TEMP handling
#                  • Validation summary and XLSX export
#                  • Interactive PyQt5 GUI interface
#
# Existing Utility Dependency:
#                  P:\BSP_LocalDev\_GLIB\harm_bsp_v_1_0\macros\report\api\get_spreadsheet.py
#
# Example Inputs :
#
#                  SharePoint Site:
#                      BSP - Team Mani
#
#                  SharePoint File:
#                      BSP/SDTM_Specs/CSR/SDTM_Specification.xlsx
#
# Run            :
#                  python metadata_validator.pyw
#
# ====================================================================================================

import os
import sys
import subprocess
import importlib
from pathlib import Path
from datetime import datetime

# ====================================================================================================
# AUTO INSTALL
# ====================================================================================================

def install_if_missing(package_name, import_name=None):
    module_name = import_name or package_name
    try:
        return importlib.import_module(module_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
        importlib.invalidate_caches()
        return importlib.import_module(module_name)

pd = install_if_missing("pandas", "pandas")
install_if_missing("openpyxl", "openpyxl")
install_if_missing("PyQt5", "PyQt5")

try:
    pyreadstat = install_if_missing("pyreadstat", "pyreadstat")
except Exception:
    pyreadstat = None

from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook

# ====================================================================================================
# CONFIG
# ====================================================================================================

TEMP_DIR = r"C:\TEMP"
DOWNLOAD_PATH = r"C:\TEMP\metadata_studio_spec.xlsx"
GET_SPREADSHEET_SCRIPT = r"P:\BSP_LocalDev\_GLIB\harm_bsp_v_1_0\macros\report\api\get_spreadsheet.py"

DEFAULT_SITE_NAME = "BSP - Team Mani"
DEFAULT_FILE_PATH = "BSP/SDTM_Specs/CSR/SDTM_Specification.xlsx"

THEME = {
    "app_bg": "#f3f7fd",
    "header": "#cfe7ff",
    "header_text": "#103760",
    "subtitle": "#315f8f",
    "panel": "#ffffff",
    "border": "#b7cde8",
    "table_header": "#d8ebff",
    "status": "#fff8dc",
    "button": "#4d8ef7",
    "button_hover": "#2f75dd",
    "green_bg": "#e7f8ec",
    "green_text": "#116329",
    "red_bg": "#fff0f0",
    "red_text": "#9a1b1b",
    "yellow_bg": "#fff8dc",
}

# ====================================================================================================
# HELPERS
# ====================================================================================================

def safe_text(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()

def safe_upper(value):
    return safe_text(value).upper()

def normalize_dataset_type(value):
    """
    Convert pyreadstat/readstat type values to simple SAS-style type:
        char
        num
    """
    txt = safe_upper(value)

    if not txt:
        return ""

    # Common pyreadstat/readstat values
    if txt in ["STRING", "CHAR", "CHARACTER", "OBJECT", "TEXT"]:
        return "char"

    if txt in ["DOUBLE", "FLOAT", "NUMERIC", "INTEGER", "INT", "NUMBER", "DATE", "DATETIME", "TIME"]:
        return "num"

    # SAS original type strings may sometimes include $ for character formats
    if "$" in txt:
        return "char"

    # If BEST, DATE, DATETIME, TIME etc. are detected, they are numeric storage in SAS
    if txt.startswith(("BEST", "DATE", "DATETIME", "TIME", "YY", "MMDD", "DDMM", "E8601")):
        return "num"

    return txt.lower()

def to_int_or_none(value):
    """
    Convert a value to integer where possible.
    Handles Excel/pandas values like 200, 200.0, '200'.
    """
    txt = safe_text(value)
    if txt == "":
        return None

    try:
        return int(float(txt))
    except Exception:
        return None

def normalize_keep(value):
    txt = safe_upper(value)
    return txt in ["1", "Y", "YES", "TRUE", "X", "KEEP"]

def is_probable_domain_sheet(sheet_name, df):
    name = safe_upper(sheet_name)

    if name in ["README", "NOTES", "INSTRUCTIONS", "CHANGELOG", "CT", "CONTROLLED TERMINOLOGY"]:
        return False

    if df is None or df.empty:
        return False

    # Usually domain sheets have variable-level rows and KEEP column around column E.
    if len(df.columns) >= 5:
        return True

    return False

def clean_df_for_display(df):
    out = df.copy()
    out = out.fillna("")
    return out

def get_keep_column(df):
    # Preferred: E column by position, as per current template logic.
    if df is not None and len(df.columns) >= 5:
        return df.columns[4]

    # Fallback by name.
    for c in df.columns:
        cu = safe_upper(c)
        if cu in ["KEEP", "USE", "USED", "IN_DATASET", "DATASET_KEEP"]:
            return c

    return None

def infer_variable_column(df):
    possible = ["VARIABLE", "VARIABLE NAME", "VAR", "NAME", "VARIABLE_NAME"]
    for c in df.columns:
        if safe_upper(c) in possible:
            return c

    # Common spec templates often have variable name near B/C/D.
    for idx in [1, 2, 3, 0]:
        if len(df.columns) > idx:
            return df.columns[idx]

    return None

def infer_label_column(df):
    possible = ["LABEL", "VARIABLE LABEL", "DESCRIPTION", "VARIABLE DESCRIPTION"]
    for c in df.columns:
        if safe_upper(c) in possible:
            return c
    return None

def infer_type_column(df):
    possible = ["TYPE", "DATA TYPE", "DATATYPE", "DATA_TYPE"]
    for c in df.columns:
        if safe_upper(c) in possible:
            return c
    return None

def infer_ct_columns(df):
    cols = []
    for c in df.columns:
        cu = safe_upper(c)
        if "CT" in cu or "CODELIST" in cu or "CONTROLLED" in cu:
            cols.append(c)
    return cols


EXPECTED_SPEC_COLUMNS = [
    "Variable",
    "Label",
    "ID Var",
    "Keep",
    "Type",
    "Len",
    "Control or Format",
    "Term",
    "Core",
    "Role",
    "Origin",
    "Comments",
    "",
]


def read_domain_sheet_a_to_m_by_position(path, sheet_name):
    """
    Read domain sheet using openpyxl by fixed Excel positions.

    This avoids pandas header issues caused by blank/merged cells.

    Expected mapping:
        A Dataset
        B Variable
        C Label
        D ID Var
        E Keep
        F Type
        G Len
        H Control or Format
        I Term
        J Core
        K Role
        L Origin
        M Comments
    """
    headers = [
        "Dataset",
        "Variable",
        "Label",
        "ID Var",
        "Keep",
        "Type",
        "Len",
        "Control or Format",
        "Term",
        "Core",
        "Role",
        "Origin",
        "Comments",
    ]

    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame(columns=headers)

    ws = wb[sheet_name]

    rows = []

    # Start at row 2 because row 1 is the template header.
    for excel_row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        vals = list(excel_row)

        # Pad if row has fewer cells
        vals = vals + [""] * (13 - len(vals))

        rec = {}
        for i, h in enumerate(headers):
            rec[h] = safe_text(vals[i])

        # If Dataset column is blank, derive from sheet name.
        if rec["Dataset"] == "":
            rec["Dataset"] = sheet_name.upper()

        # Skip fully blank rows.
        if all(rec[h] == "" for h in headers):
            continue

        # Skip rows without variable name.
        if rec["Variable"] == "":
            continue

        # Skip repeated header rows.
        if rec["Variable"].upper() in ["VARIABLE", "VARIABLE NAME"]:
            continue

        rows.append(rec)

    try:
        wb.close()
    except Exception:
        pass

    return pd.DataFrame(rows, columns=headers)



def get_unique_supp_datasets_from_suppqual(path):
    """
    Read SUPPQUAL sheet and return unique SUPP dataset names from column A.

    Example:
        SUPPDM, SUPPAE, SUPPCM
    """
    try:
        df = read_domain_sheet_a_to_m_by_position(path, "SUPPQUAL")
    except Exception:
        return []

    if df is None or df.empty or "Dataset" not in df.columns:
        return []

    out = []
    for val in df["Dataset"].tolist():
        ds = safe_upper(val)
        if ds and ds.startswith("SUPP") and ds not in out:
            out.append(ds)

    return out


def build_supp_rows_from_template(path, supp_dataset):
    """
    Duplicate SUPP_TEMP rows and replace Dataset with the actual SUPP dataset.

    Example:
        SUPP_TEMP Dataset = SUPP--
        Output Dataset = SUPPDM / SUPPAE / SUPPCM
    """
    try:
        template = read_domain_sheet_a_to_m_by_position(path, "SUPP_TEMP")
    except Exception:
        return pd.DataFrame()

    if template is None or template.empty:
        return pd.DataFrame()

    out = template.copy()
    out["Dataset"] = supp_dataset

    # SUPP_TEMP rows are generated only for actual SUPP datasets listed in SUPPQUAL,
    # so they should be treated as expected/kept variables for validation.
    if "Keep" in out.columns:
        out["Keep"] = "1"

    # Remove rows without variable after template read.
    out["Variable"] = out["Variable"].apply(safe_text)
    out = out[out["Variable"] != ""].copy()

    return out

def normalize_spec_template_columns(df):
    """
    Fallback normalizer when a dataframe has already been read.
    Prefer read_domain_sheet_a_to_n() for actual domain sheet import.
    """
    if df is None or df.empty:
        return df

    cols = [
        "Dataset",
        "Variable",
        "Label",
        "ID Var",
        "Keep",
        "Type",
        "Len",
        "Control or Format",
        "Term",
        "Core",
        "Role",
        "Origin",
        "Comments",
    ]

    out = df.copy()
    out = out.dropna(how="all")
    out = out.dropna(axis=1, how="all")
    out = out.iloc[:, :14].copy()
    out.columns = cols[:len(out.columns)]

    for col in cols:
        if col not in out.columns:
            out[col] = ""

    return out[cols]


def is_support_sheet(sheet_name):
    """
    Returns True for template/helper sheets that should not be appended into
    the variable metadata view.

    These sheets are useful in the workbook, but they are not domain variable sheets.
    """
    name = safe_upper(sheet_name).replace(" ", "_")

    exclude_exact = {
        "README",
        "READ_ME",
        "STATUS",
        "DOMAINS",
        "BLANK_SPEC",
        "SOURCE_DATA",
        "SUPPQUAL",
        "SUPP_TEMP",
        "FORMATS",
        "VALUEMETADATA",
        "VALUE_METADATA",
        "DEV_FORMATS",
        "QC_FORMATS",
        "LB_FORMATS",
        "QC_LB_EXT_TESTS",
        "QS_TESTCD",
        "LOOKUPS",
    }

    if name in exclude_exact:
        return True

    # SDTM relationship/support metadata sheets usually start with underscore.
    # Example: _TA, _TE, _TI, _TS, _TV
    if name.startswith("_"):
        return True

    # Format / lookup helper sheets
    if "FORMAT" in name:
        return True

    if "LOOKUP" in name:
        return True

    return False


def is_domain_metadata_sheet(sheet_name, df):
    """
    Domain sheets are appended into one combined metadata table.
    """
    if is_support_sheet(sheet_name):
        return False

    if df is None or df.empty:
        return False

    if len(df.columns) < 5:
        return False

    return True


# ====================================================================================================
# SHAREPOINT DOWNLOAD
# ====================================================================================================

def download_spec_from_sharepoint(site_name, file_path, status_callback=None):
    os.makedirs(TEMP_DIR, exist_ok=True)

    if os.path.exists(DOWNLOAD_PATH):
        try:
            os.remove(DOWNLOAD_PATH)
        except Exception:
            pass

    if not os.path.exists(GET_SPREADSHEET_SCRIPT):
        raise FileNotFoundError(
            "get_spreadsheet.py was not found:\n\n"
            f"{GET_SPREADSHEET_SCRIPT}"
        )

    if status_callback:
        status_callback("Downloading specification from SharePoint...")

    cmd = [
        sys.executable,
        GET_SPREADSHEET_SCRIPT,
        "-s", site_name,
        "-f", file_path,
        "-o", DOWNLOAD_PATH,
        "-d", "Excel",
    ]

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=240, shell=False)

    if result.returncode != 0:
        raise RuntimeError(
            "SharePoint specification download failed.\n\n"
            f"Command:\n{' '.join(cmd)}\n\n"
            f"STDOUT:\n{result.stdout}\n\n"
            f"STDERR:\n{result.stderr}"
        )

    if not os.path.exists(DOWNLOAD_PATH):
        raise FileNotFoundError(
            "Downloaded specification file was not created:\n\n"
            f"{DOWNLOAD_PATH}"
        )

    return DOWNLOAD_PATH

# ====================================================================================================
# TABLE UTILITY
# ====================================================================================================

class DataFrameTable(QtWidgets.QTableWidget):
    def __init__(self):
        super().__init__()
        self.setAlternatingRowColors(False)
        self.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.setWordWrap(True)
        self.verticalHeader().setVisible(False)
        self.setShowGrid(True)

    def load_df(self, df, color_keep=False):
        self.clear()

        if df is None:
            self.setRowCount(0)
            self.setColumnCount(0)
            return

        df = clean_df_for_display(df)

        self.setRowCount(len(df))
        self.setColumnCount(len(df.columns))
        self.setHorizontalHeaderLabels([str(c) for c in df.columns])

        keep_col = get_keep_column(df) if color_keep else None

        for r in range(len(df)):
            keep_flag = False
            if keep_col is not None:
                keep_flag = normalize_keep(df.iloc[r][keep_col])

            for c in range(len(df.columns)):
                val = safe_text(df.iloc[r, c])
                item = QtWidgets.QTableWidgetItem(val)
                item.setToolTip(val)

                if color_keep:
                    if keep_flag:
                        item.setBackground(QtGui.QColor(THEME["green_bg"]))
                    else:
                        item.setBackground(QtGui.QColor("#ffffff"))

                self.setItem(r, c, item)

        self.resizeColumnsToContents()

        # Prevent Comments column from taking excessive width.
        for idx in range(self.columnCount()):
            header_item = self.horizontalHeaderItem(idx)
            if header_item and header_item.text() == "Comments":
                self.setColumnWidth(idx, 300)

        self.horizontalHeader().setStretchLastSection(False)

# ====================================================================================================
# MAIN APPLICATION
# ====================================================================================================

class MetadataStudio(QtWidgets.QWidget):

    def __init__(self):
        super().__init__()

        self.spec_path = ""
        self.dataset_path = ""
        self.excel_data = {}
        self.normalized_spec = pd.DataFrame()
        self.dataset_inventory = pd.DataFrame()
        self.dataset_metadata = pd.DataFrame()
        self.validation_issues = pd.DataFrame()
        self.validation_summary_text = ""
        self.temp_downloaded_spec = ""

        self.setWindowTitle("Clinical Metadata Studio - Module 2")
        self.setMinimumSize(1250, 760)
        self.resize(1700, 950)

        self.build_ui()
        self.apply_style()

    # ----------------------------------------------------------------------------------------------
    # UI
    # ----------------------------------------------------------------------------------------------

    def build_ui(self):
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(10, 8, 10, 8)
        root.setSpacing(8)

        header = QtWidgets.QFrame()
        header.setObjectName("Header")

        h = QtWidgets.QHBoxLayout(header)
        h.setContentsMargins(14, 10, 14, 10)

        title = QtWidgets.QLabel("Clinical Metadata Studio")
        title.setObjectName("Title")

        subtitle = QtWidgets.QLabel("Module 2 • SharePoint Spec Import • SAS/XPT Discovery • Metadata Viewer")
        subtitle.setObjectName("Subtitle")

        h.addWidget(title)
        h.addSpacing(18)
        h.addWidget(subtitle)
        h.addStretch()

        root.addWidget(header)

        controls = QtWidgets.QFrame()
        controls.setObjectName("Controls")
        g = QtWidgets.QGridLayout(controls)
        g.setContentsMargins(10, 8, 10, 8)
        g.setHorizontalSpacing(8)
        g.setVerticalSpacing(6)

        self.use_sharepoint_chk = QtWidgets.QCheckBox("Use SharePoint Spec")
        self.use_sharepoint_chk.setChecked(True)
        self.use_sharepoint_chk.stateChanged.connect(self.update_spec_mode_ui)

        self.site_label = QtWidgets.QLabel("SharePoint Site")
        self.site_edit = QtWidgets.QLineEdit(DEFAULT_SITE_NAME)
        self.site_edit.setPlaceholderText("SharePoint Site Name")

        self.file_label = QtWidgets.QLabel("Spec File Path")
        self.file_edit = QtWidgets.QLineEdit(DEFAULT_FILE_PATH)
        self.file_edit.setPlaceholderText("SharePoint File Path, e.g. BSP/SDTM_Specs/CSR/SDTM_Specification.xlsx")

        self.local_label = QtWidgets.QLabel("Local Spec")
        self.local_spec_edit = QtWidgets.QLineEdit()
        self.local_spec_edit.setPlaceholderText("Local specification file path")

        self.btn_local_spec = QtWidgets.QPushButton("Browse Local Spec")
        self.btn_local_spec.clicked.connect(self.browse_local_spec)

        self.btn_load_spec = QtWidgets.QPushButton("Load Spec")
        self.btn_load_spec.clicked.connect(self.load_spec_by_mode)

        self.dataset_edit = QtWidgets.QLineEdit()
        self.dataset_edit.setPlaceholderText("SAS/XPT Folder")

        self.btn_dataset = QtWidgets.QPushButton("Browse SAS/XPT Folder")
        self.btn_dataset.clicked.connect(self.browse_dataset_folder)

        self.btn_load_data = QtWidgets.QPushButton("Load Data")
        self.btn_load_data.clicked.connect(self.load_data_only)

        self.btn_validate = QtWidgets.QPushButton("Validate")
        self.btn_validate.clicked.connect(self.run_validation_only)

        self.btn_export_validation = QtWidgets.QPushButton("Export Validation XLSX")
        self.btn_export_validation.clicked.connect(self.export_validation_xlsx)

        g.addWidget(self.use_sharepoint_chk, 0, 0)

        g.addWidget(self.site_label, 1, 0)
        g.addWidget(self.site_edit, 1, 1)
        g.addWidget(self.file_label, 1, 2)
        g.addWidget(self.file_edit, 1, 3)
        g.addWidget(self.btn_load_spec, 1, 4)

        g.addWidget(self.local_label, 2, 0)
        g.addWidget(self.local_spec_edit, 2, 1, 1, 3)
        g.addWidget(self.btn_local_spec, 2, 4)

        g.addWidget(QtWidgets.QLabel("Dataset Folder"), 3, 0)
        g.addWidget(self.dataset_edit, 3, 1, 1, 2)
        g.addWidget(self.btn_dataset, 3, 3)
        g.addWidget(self.btn_load_data, 3, 4)

        g.addWidget(self.btn_validate, 4, 3)
        g.addWidget(self.btn_export_validation, 4, 4)

        root.addWidget(controls)

        self.tabs = QtWidgets.QTabWidget()

        self.tab_study = QtWidgets.QWidget()
        self.tab_specs = QtWidgets.QWidget()
        self.tab_normalized = QtWidgets.QWidget()
        self.tab_datasets = QtWidgets.QWidget()
        self.tab_dataset_metadata = QtWidgets.QWidget()
        self.tab_ct = QtWidgets.QWidget()
        self.tab_define = QtWidgets.QWidget()
        self.tab_validation = QtWidgets.QWidget()

        self.tabs.addTab(self.tab_study, "Study Metadata")
        self.tabs.addTab(self.tab_specs, "Specification Sheets")
        self.tabs.addTab(self.tab_normalized, "Normalized Metadata")
        self.tabs.addTab(self.tab_datasets, "Dataset Inventory")
        self.tabs.addTab(self.tab_dataset_metadata, "Dataset Metadata")
        self.tabs.addTab(self.tab_ct, "CT Validation")
        self.tabs.addTab(self.tab_define, "Define Preview")
        self.tabs.addTab(self.tab_validation, "Validation Report")

        root.addWidget(self.tabs, stretch=1)

        self.build_study_tab()
        self.build_specs_tab()
        self.build_normalized_tab()
        self.build_datasets_tab()
        self.build_dataset_metadata_tab()
        self.build_text_tabs()

        self.status = QtWidgets.QLabel("Ready.")
        self.status.setObjectName("Status")
        root.addWidget(self.status)

        # Apply initial SharePoint/Local spec mode after status label is created.
        self.update_spec_mode_ui()

    def build_study_tab(self):
        layout = QtWidgets.QVBoxLayout(self.tab_study)
        self.study_text = QtWidgets.QTextEdit()
        self.study_text.setReadOnly(True)
        layout.addWidget(self.study_text)

    def build_specs_tab(self):
        layout = QtWidgets.QVBoxLayout(self.tab_specs)

        self.spec_summary_label = QtWidgets.QLabel(
            "Combined domain metadata view. Template/support sheets are excluded automatically."
        )
        self.spec_summary_label.setObjectName("InfoLabel")

        self.spec_table = DataFrameTable()

        layout.addWidget(self.spec_summary_label)
        layout.addWidget(self.spec_table)

    def build_normalized_tab(self):
        layout = QtWidgets.QVBoxLayout(self.tab_normalized)
        self.normalized_table = DataFrameTable()
        layout.addWidget(self.normalized_table)

    def build_datasets_tab(self):
        layout = QtWidgets.QVBoxLayout(self.tab_datasets)

        self.dataset_folder_label = QtWidgets.QLabel("Dataset Folder: -")
        self.dataset_folder_label.setObjectName("InfoLabel")

        self.dataset_table = DataFrameTable()

        layout.addWidget(self.dataset_folder_label)
        layout.addWidget(self.dataset_table)

    def build_dataset_metadata_tab(self):
        layout = QtWidgets.QVBoxLayout(self.tab_dataset_metadata)

        info = QtWidgets.QLabel(
            "PROC CONTENTS-style metadata extracted from selected SAS/XPT folder. "
            "Only files directly present in the selected folder are scanned."
        )
        info.setObjectName("InfoLabel")

        self.dataset_metadata_table = DataFrameTable()

        layout.addWidget(info)
        layout.addWidget(self.dataset_metadata_table)

    def build_text_tabs(self):
        # CT and Define are text-preview tabs for now.
        for tab, attr in [
            (self.tab_ct, "ct_text"),
            (self.tab_define, "define_text"),
        ]:
            layout = QtWidgets.QVBoxLayout(tab)
            box = QtWidgets.QTextEdit()
            box.setReadOnly(True)
            setattr(self, attr, box)
            layout.addWidget(box)

        # Validation tab: readable summary + tabular issues.
        layout = QtWidgets.QVBoxLayout(self.tab_validation)

        self.validation_text = QtWidgets.QTextEdit()
        self.validation_text.setReadOnly(True)
        self.validation_text.setMaximumHeight(190)

        self.validation_table = DataFrameTable()

        layout.addWidget(self.validation_text)
        layout.addWidget(self.validation_table)

    # ----------------------------------------------------------------------------------------------
    # STYLE
    # ----------------------------------------------------------------------------------------------

    def apply_style(self):
        self.setStyleSheet(f"""
            QWidget {{
                background-color: {THEME['app_bg']};
                font-family: Segoe UI;
                font-size: 10pt;
                color: #1c2e4a;
            }}

            #Header {{
                background-color: {THEME['header']};
                border-radius: 14px;
                border: 1px solid #9cc8f5;
            }}

            #Title {{
                font-size: 20pt;
                font-weight: bold;
                color: {THEME['header_text']};
            }}

            #Subtitle {{
                color: {THEME['subtitle']};
                font-weight: bold;
            }}

            #Controls {{
                background-color: white;
                border-radius: 12px;
                border: 1px solid {THEME['border']};
            }}

            QPushButton {{
                background-color: {THEME['button']};
                color: white;
                border-radius: 8px;
                padding: 7px 12px;
                font-weight: bold;
            }}

            QPushButton:hover {{
                background-color: {THEME['button_hover']};
            }}

            QLineEdit, QComboBox {{
                background-color: white;
                border: 1px solid #a8bfdc;
                border-radius: 8px;
                padding: 5px;
            }}

            QTableWidget {{
                background-color: white;
                border: 1px solid #8fb2d6;
                gridline-color: #c9d7e6;
            }}

            QHeaderView::section {{
                background-color: {THEME['table_header']};
                padding: 6px;
                border: 1px solid #8fb2d6;
                font-weight: bold;
                color: {THEME['header_text']};
            }}

            QTextEdit {{
                background-color: white;
                border: 1px solid #b7cde8;
                border-radius: 8px;
                padding: 8px;
            }}

            #Status {{
                background-color: {THEME['status']};
                border: 1px solid #dbc46f;
                border-radius: 8px;
                padding: 6px;
                color: #4d3b00;
                font-weight: bold;
            }}
        """)

    # ----------------------------------------------------------------------------------------------
    # STATUS
    # ----------------------------------------------------------------------------------------------

    def set_status(self, text):
        self.status.setText(text)
        QtWidgets.QApplication.processEvents()

    # ----------------------------------------------------------------------------------------------
    # SPEC MODE UI
    # ----------------------------------------------------------------------------------------------

    def update_spec_mode_ui(self):
        use_sp = self.use_sharepoint_chk.isChecked()

        # SharePoint fields visible only in SharePoint mode
        for w in [self.site_label, self.site_edit, self.file_label, self.file_edit]:
            w.setVisible(use_sp)

        # Local spec fields visible only in Local mode
        for w in [self.local_label, self.local_spec_edit, self.btn_local_spec]:
            w.setVisible(not use_sp)

        if use_sp:
            self.btn_load_spec.setText("Load Spec")
            if hasattr(self, "status"):
                self.status.setText("SharePoint spec mode selected.")
        else:
            self.btn_load_spec.setText("Load Spec")
            if hasattr(self, "status"):
                self.status.setText("Local spec mode selected. Browse or enter local spec path.")

    def load_spec_by_mode(self):
        if self.use_sharepoint_chk.isChecked():
            self.load_spec_from_sharepoint()
        else:
            path = self.local_spec_edit.text().strip()
            if not path:
                self.browse_local_spec()
                path = self.local_spec_edit.text().strip()

            if not path:
                QtWidgets.QMessageBox.warning(self, "Missing Local Spec", "Please browse or enter a local specification file.")
                return

            if not os.path.exists(path):
                QtWidgets.QMessageBox.warning(self, "Invalid Local Spec", f"File does not exist:\n\n{path}")
                return

            self.spec_path = path
            self.load_spec_only()

    # ----------------------------------------------------------------------------------------------
    # LOAD SPEC
    # ----------------------------------------------------------------------------------------------

    def load_spec_from_sharepoint(self):
        site_name = self.site_edit.text().strip()
        file_path = self.file_edit.text().strip()

        if not site_name or not file_path:
            QtWidgets.QMessageBox.warning(self, "Missing SharePoint Details", "Please enter SharePoint Site and File Path.")
            return

        try:
            local_path = download_spec_from_sharepoint(site_name, file_path, self.set_status)
            self.temp_downloaded_spec = local_path
            self.spec_path = local_path
            self.local_spec_edit.setText(local_path)
            self.load_spec_only()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "SharePoint Import Failed", str(e))
            self.set_status("SharePoint import failed.")

    def browse_local_spec(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Select Study Specification",
            "",
            "Excel Files (*.xlsx *.xlsm *.xls)"
        )

        if path:
            self.spec_path = path
            self.local_spec_edit.setText(path)

    def browse_dataset_folder(self):
        path = QtWidgets.QFileDialog.getExistingDirectory(self, "Select SAS/XPT Dataset Folder")
        if path:
            self.dataset_path = path
            self.dataset_edit.setText(path)

    def load_spec_only(self):
        """
        Load only the specification workbook and combined spec metadata.
        Does not load datasets and does not run validation.
        """
        if not self.spec_path:
            self.spec_path = self.local_spec_edit.text().strip()

        if not self.spec_path:
            QtWidgets.QMessageBox.warning(self, "Missing Spec", "Please load a SharePoint spec or browse local spec.")
            return

        try:
            self.set_status("Reading specification workbook...")
            self.load_workbook(self.spec_path)

            self.set_status("Building combined specification metadata...")
            self.build_normalized_metadata()

            self.validation_issues = pd.DataFrame()
            self.validation_summary_text = ""

            self.populate_spec_only_tabs()
            self.set_status("Specification loaded successfully. Next, load SAS/XPT datasets.")

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Spec Load Failed", str(e))
            self.set_status("Spec load failed.")

    def load_data_only(self):
        """
        Load only dataset inventory and PROC CONTENTS-style metadata.
        Does not reload spec and does not run validation.
        """
        self.dataset_path = self.dataset_edit.text().strip()

        if not self.dataset_path:
            QtWidgets.QMessageBox.warning(self, "Missing Dataset Folder", "Please browse/select a SAS/XPT dataset folder.")
            return

        if not os.path.exists(self.dataset_path):
            QtWidgets.QMessageBox.warning(self, "Invalid Dataset Folder", f"Folder does not exist:\\n\\n{self.dataset_path}")
            return

        try:
            self.set_status("Reading SAS/XPT dataset files from selected folder...")
            self.load_dataset_inventory()

            self.set_status("Extracting PROC CONTENTS-style metadata from SAS/XPT files...")
            self.extract_dataset_metadata()

            self.validation_issues = pd.DataFrame()
            self.validation_summary_text = ""

            self.populate_dataset_only_tabs()
            self.set_status("Dataset metadata loaded successfully. Click Validate to compare against spec.")

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Dataset Load Failed", str(e))
            self.set_status("Dataset load failed.")

    def run_validation_only(self):
        """
        Run validation only when user clicks Validate.
        Requires both spec metadata and dataset metadata.
        """
        if self.normalized_spec is None or self.normalized_spec.empty:
            QtWidgets.QMessageBox.warning(self, "Spec Not Loaded", "Please load specification first.")
            return

        if self.dataset_metadata is None or self.dataset_metadata.empty:
            QtWidgets.QMessageBox.warning(self, "Dataset Metadata Not Loaded", "Please load SAS/XPT dataset metadata first.")
            return

        self.set_status("Running metadata validation against specification...")
        self.populate_validation_tab()
        self.tabs.setCurrentWidget(self.tab_validation)
        self.set_status("Validation completed.")

    def load_all_metadata(self):
        """
        Backward-compatible wrapper only. Current UI uses separate buttons.
        """
        self.load_spec_only()
        self.load_data_only()
        self.run_validation_only()

    def load_workbook(self, path):
        xl = pd.ExcelFile(path)
        self.excel_data = {}

        for sheet in xl.sheet_names:
            try:
                # Read all sheets as-is. Header row assumed first row.
                df = pd.read_excel(path, sheet_name=sheet)
                self.excel_data[sheet] = df
            except Exception as e:
                print(f"Failed reading {sheet}: {e}")

    # ----------------------------------------------------------------------------------------------
    # NORMALIZE / APPEND SPEC METADATA
    # ----------------------------------------------------------------------------------------------

    def build_normalized_metadata(self):
        """
        Append all valid domain sheets using only columns A:M.

        Also handles SUPPQUAL:
        - SUPPQUAL itself is a helper sheet, not appended as-is.
        - Unique SUPP datasets from SUPPQUAL are identified.
        - SUPP_TEMP is duplicated once per SUPP dataset and appended to the spec metadata.
        """
        frames = []

        for sheet, df_check in self.excel_data.items():
            if not is_domain_metadata_sheet(sheet, df_check):
                continue

            try:
                df_std = read_domain_sheet_a_to_m_by_position(self.spec_path, sheet)
            except Exception:
                df_std = normalize_spec_template_columns(df_check)

            if df_std is None or df_std.empty:
                continue

            df_std["Variable"] = df_std["Variable"].apply(safe_text)
            df_std = df_std[df_std["Variable"] != ""].copy()
            df_std = df_std[~df_std["Variable"].str.upper().isin(["VARIABLE", "NAN"])].copy()

            frames.append(df_std)

        # SUPPQUAL expansion from SUPP_TEMP
        supp_datasets = get_unique_supp_datasets_from_suppqual(self.spec_path)

        for supp_ds in supp_datasets:
            supp_rows = build_supp_rows_from_template(self.spec_path, supp_ds)
            if supp_rows is not None and not supp_rows.empty:
                frames.append(supp_rows)

        final_cols = [
            "Dataset",
            "Variable",
            "Label",
            "ID Var",
            "Keep",
            "Type",
            "Len",
            "Control or Format",
            "Term",
            "Core",
            "Role",
            "Origin",
            "Comments",
        ]

        if frames:
            self.normalized_spec = pd.concat(frames, ignore_index=True)
            self.normalized_spec = self.normalized_spec[final_cols]
        else:
            self.normalized_spec = pd.DataFrame(columns=final_cols)

    # ----------------------------------------------------------------------------------------------
    # DATASET INVENTORY
    # ----------------------------------------------------------------------------------------------

    def load_dataset_inventory(self):
        """
        Build dataset inventory using only the selected folder.

        Inventory shows:
            Dataset
            Dataset Label
            Type
            File
            Variables
            Size MB

        Folder is displayed once as a title label, not repeated per row.
        """
        self.dataset_path = self.dataset_edit.text().strip()

        rows = []

        if self.dataset_path and os.path.exists(self.dataset_path):
            for f in Path(self.dataset_path).glob("*"):
                if f.suffix.lower() not in [".sas7bdat", ".xpt"]:
                    continue

                ds = f.stem.upper()
                ftype = f.suffix.replace(".", "").upper()

                n_vars = ""
                ds_label = ""

                if pyreadstat is not None:
                    try:
                        if ftype == "SAS7BDAT":
                            _, meta = pyreadstat.read_sas7bdat(str(f), metadataonly=True)
                        elif ftype == "XPT":
                            _, meta = pyreadstat.read_xport(str(f), metadataonly=True)
                        else:
                            meta = None

                        if meta is not None:
                            names = list(getattr(meta, "column_names", []) or [])
                            n_vars = len(names)

                            # Dataset label availability differs by file type/source.
                            ds_label = safe_text(
                                getattr(meta, "file_label", "")
                                or getattr(meta, "table_name", "")
                            )

                    except Exception as e:
                        ds_label = f"Metadata read warning: {e}"

                rows.append({
                    "Dataset": ds,
                    "Dataset Label": ds_label,
                    "Type": ftype,
                    "File": f.name,
                    "Variables": n_vars,
                    "Size MB": round(f.stat().st_size / (1024 * 1024), 2),
                })

        self.dataset_inventory = pd.DataFrame(rows)

    # ----------------------------------------------------------------------------------------------
    # DATASET METADATA EXTRACTION
    # ----------------------------------------------------------------------------------------------

    def extract_dataset_metadata(self):
        """
        Extract PROC CONTENTS-style metadata from selected SAS/XPT files.

        Uses metadataonly=True and does not read full data values.
        """
        rows = []

        if self.dataset_inventory is None or self.dataset_inventory.empty:
            self.dataset_metadata = pd.DataFrame(columns=[
                "Dataset", "Variable", "Label", "Type", "Length",
                "Format", "Informat", "Order", "Source File"
            ])
            return

        if pyreadstat is None:
            self.dataset_metadata = pd.DataFrame([{
                "Dataset": "",
                "Variable": "",
                "Label": "",
                "Type": "",
                "Length": "",
                "Format": "",
                "Informat": "",
                "Order": "",
                "Source File": "pyreadstat is not available. Install failed or package unavailable."
            }])
            return

        for _, inv in self.dataset_inventory.iterrows():
            ds = safe_text(inv.get("Dataset"))
            fpath = Path(self.dataset_path) / safe_text(inv.get("File"))
            ftype = safe_upper(inv.get("Type"))

            try:
                if ftype == "SAS7BDAT":
                    _, meta = pyreadstat.read_sas7bdat(str(fpath), metadataonly=True)
                elif ftype == "XPT":
                    _, meta = pyreadstat.read_xport(str(fpath), metadataonly=True)
                else:
                    continue

                names = list(getattr(meta, "column_names", []) or [])
                labels = getattr(meta, "column_labels", []) or []
                original_types = getattr(meta, "original_variable_types", {}) or {}
                readstat_types = getattr(meta, "readstat_variable_types", {}) or {}
                var_formats = getattr(meta, "variable_to_label", {}) or {}
                storage_widths = getattr(meta, "variable_storage_width", {}) or {}
                display_widths = getattr(meta, "variable_display_width", {}) or {}

                for i, var in enumerate(names):
                    label = ""
                    if isinstance(labels, list) and i < len(labels):
                        label = safe_text(labels[i])

                    typ = safe_text(readstat_types.get(var, ""))
                    orig = safe_text(original_types.get(var, ""))

                    # Prefer original type when available; otherwise readstat type.
                    # Display as simple SAS-style type: char / num.
                    out_type = normalize_dataset_type(orig or typ)

                    length = ""
                    if var in storage_widths:
                        length = safe_text(storage_widths.get(var))
                    elif var in display_widths:
                        length = safe_text(display_widths.get(var))

                    rows.append({
                        "Dataset": ds,
                        "Variable": safe_text(var),
                        "Label": label,
                        "Type": out_type,
                        "Length": length,
                        "Format": safe_text(var_formats.get(var, "")),
                        "Informat": "",
                        "Order": i + 1,
                        "Source File": safe_text(fpath.name),
                    })

            except Exception as e:
                rows.append({
                    "Dataset": ds,
                    "Variable": "",
                    "Label": "",
                    "Type": "",
                    "Length": "",
                    "Format": "",
                    "Informat": "",
                    "Order": "",
                    "Source File": f"ERROR reading {fpath.name}: {e}",
                })

        self.dataset_metadata = pd.DataFrame(rows)

    # ----------------------------------------------------------------------------------------------
    # SPEC VS DATASET VALIDATION
    # ----------------------------------------------------------------------------------------------

    def build_spec_dataset_validation(self):
        """
        Build readable validation summary and tabular issue dataframe.

        Logic:
        - Missing Variable:
              Keep=1 in spec, but variable is not present in dataset.
        - Not Marked Keep:
              Variable exists in dataset and exists in spec, but Keep is blank/not active.
        - Extra Variable:
              Variable exists in dataset but does not exist anywhere in spec.
        - Length Greater Than Spec:
              Variable exists in both spec and dataset, and dataset character length > spec length.
        """
        summary = []
        issues = []

        summary.append("Spec vs Dataset Metadata Validation")
        summary.append("=" * 90)
        summary.append("")

        if self.normalized_spec is None or self.normalized_spec.empty:
            summary.append("No normalized specification metadata available.")
            self.validation_issues = pd.DataFrame()
            return "\n".join(summary)

        if self.dataset_metadata is None or self.dataset_metadata.empty:
            summary.append("No dataset metadata available. Select SAS/XPT folder and click Refresh Metadata.")
            self.validation_issues = pd.DataFrame()
            return "\n".join(summary)

        spec = self.normalized_spec.copy()
        data = self.dataset_metadata.copy()

        spec_all_pairs = set(
            (safe_upper(r.get("Dataset")), safe_upper(r.get("Variable")))
            for _, r in spec.iterrows()
            if safe_text(r.get("Dataset")) and safe_text(r.get("Variable"))
        )

        if "Keep" in spec.columns:
            spec_keep = spec[spec["Keep"].apply(normalize_keep)].copy()
        else:
            spec_keep = spec.copy()

        spec_keep_pairs = set(
            (safe_upper(r.get("Dataset")), safe_upper(r.get("Variable")))
            for _, r in spec_keep.iterrows()
            if safe_text(r.get("Dataset")) and safe_text(r.get("Variable"))
        )

        data_pairs = set(
            (safe_upper(r.get("Dataset")), safe_upper(r.get("Variable")))
            for _, r in data.iterrows()
            if safe_text(r.get("Dataset")) and safe_text(r.get("Variable"))
        )

        missing = sorted(spec_keep_pairs - data_pairs)
        not_marked_keep = sorted((data_pairs & spec_all_pairs) - spec_keep_pairs)
        true_extra = sorted(data_pairs - spec_all_pairs)

        for ds, var in missing:
            spec_row = spec_keep[
                (spec_keep["Dataset"].apply(safe_upper) == ds) &
                (spec_keep["Variable"].apply(safe_upper) == var)
            ]

            label = ""
            typ = ""
            length = ""

            if not spec_row.empty:
                label = safe_text(spec_row.iloc[0].get("Label"))
                typ = safe_text(spec_row.iloc[0].get("Type"))
                length = safe_text(spec_row.iloc[0].get("Len"))

            issues.append({
                "Severity": "Error",
                "Check": "Missing Variable",
                "Dataset": ds,
                "Variable": var,
                "Spec Type": typ,
                "Data Type": "",
                "Spec Len": length,
                "Data Len": "",
                "Spec Label": label,
                "Data Label": "",
                "Message": "Variable is marked Keep in spec but not found in dataset."
            })

        for ds, var in not_marked_keep:
            spec_row = spec[
                (spec["Dataset"].apply(safe_upper) == ds) &
                (spec["Variable"].apply(safe_upper) == var)
            ]
            data_row = data[
                (data["Dataset"].apply(safe_upper) == ds) &
                (data["Variable"].apply(safe_upper) == var)
            ]

            s = spec_row.iloc[0] if not spec_row.empty else {}
            d = data_row.iloc[0] if not data_row.empty else {}

            issues.append({
                "Severity": "Warning",
                "Check": "Not Marked Keep",
                "Dataset": ds,
                "Variable": var,
                "Spec Type": safe_text(s.get("Type")) if hasattr(s, "get") else "",
                "Data Type": safe_text(d.get("Type")) if hasattr(d, "get") else "",
                "Spec Len": safe_text(s.get("Len")) if hasattr(s, "get") else "",
                "Data Len": safe_text(d.get("Length")) if hasattr(d, "get") else "",
                "Spec Label": safe_text(s.get("Label")) if hasattr(s, "get") else "",
                "Data Label": safe_text(d.get("Label")) if hasattr(d, "get") else "",
                "Message": "Variable exists in dataset and in spec, but Keep is not marked in spec."
            })

        for ds, var in true_extra:
            data_row = data[
                (data["Dataset"].apply(safe_upper) == ds) &
                (data["Variable"].apply(safe_upper) == var)
            ]

            dlabel = ""
            dtyp = ""
            dlen = ""

            if not data_row.empty:
                dlabel = safe_text(data_row.iloc[0].get("Label"))
                dtyp = safe_text(data_row.iloc[0].get("Type"))
                dlen = safe_text(data_row.iloc[0].get("Length"))

            issues.append({
                "Severity": "Warning",
                "Check": "Extra Variable",
                "Dataset": ds,
                "Variable": var,
                "Spec Type": "",
                "Data Type": dtyp,
                "Spec Len": "",
                "Data Len": dlen,
                "Spec Label": "",
                "Data Label": dlabel,
                "Message": "Variable exists in dataset but is not present anywhere in spec."
            })

        # Metadata checks for all variables that exist in BOTH spec and dataset.
        # This includes Keep=1 and non-Keep rows, so length/type issues are still visible.
        common_all = sorted(spec_all_pairs & data_pairs)

        for ds, var in common_all:
            spec_row = spec[
                (spec["Dataset"].apply(safe_upper) == ds) &
                (spec["Variable"].apply(safe_upper) == var)
            ]
            data_row = data[
                (data["Dataset"].apply(safe_upper) == ds) &
                (data["Variable"].apply(safe_upper) == var)
            ]

            if spec_row.empty or data_row.empty:
                continue

            s = spec_row.iloc[0]
            d = data_row.iloc[0]

            stype = normalize_dataset_type(safe_text(s.get("Type")))
            dtype = normalize_dataset_type(safe_text(d.get("Type")))

            slen = safe_text(s.get("Len"))
            dlen = safe_text(d.get("Length"))

            slabel = safe_text(s.get("Label"))
            dlabel = safe_text(d.get("Label"))

            if stype and dtype and stype != dtype:
                issues.append({
                    "Severity": "Error",
                    "Check": "Type Mismatch",
                    "Dataset": ds,
                    "Variable": var,
                    "Spec Type": stype,
                    "Data Type": dtype,
                    "Spec Len": slen,
                    "Data Len": dlen,
                    "Spec Label": slabel,
                    "Data Label": dlabel,
                    "Message": "Spec Type and Dataset Type do not match."
                })

            spec_len_num = to_int_or_none(slen)
            data_len_num = to_int_or_none(dlen)

            if (
                stype == "char"
                and spec_len_num is not None
                and data_len_num is not None
                and data_len_num > spec_len_num
            ):
                issues.append({
                    "Severity": "Warning",
                    "Check": "Length Greater Than Spec",
                    "Dataset": ds,
                    "Variable": var,
                    "Spec Type": stype,
                    "Data Type": dtype,
                    "Spec Len": slen,
                    "Data Len": dlen,
                    "Spec Label": slabel,
                    "Data Label": dlabel,
                    "Message": "Dataset character length is greater than the length specified in the spec."
                })

            if slabel and dlabel and slabel.strip() != dlabel.strip():
                issues.append({
                    "Severity": "Warning",
                    "Check": "Label Mismatch",
                    "Dataset": ds,
                    "Variable": var,
                    "Spec Type": stype,
                    "Data Type": dtype,
                    "Spec Len": slen,
                    "Data Len": dlen,
                    "Spec Label": slabel,
                    "Data Label": dlabel,
                    "Message": "Spec Label and Dataset Label do not match exactly."
                })

        self.validation_issues = pd.DataFrame(issues)

        summary.append(f"Total spec variables            : {len(spec)}")
        summary.append(f"Spec variables with Keep flag   : {len(spec_keep)}")
        summary.append(f"Dataset metadata variables      : {len(data)}")
        summary.append(f"Validation issues detected      : {len(self.validation_issues)}")
        summary.append("")
        summary.append(f"Missing variables               : {len(missing)}")
        summary.append(f"Variables not marked Keep       : {len(not_marked_keep)}")
        summary.append(f"True extra variables            : {len(true_extra)}")
        summary.append("")
        summary.append("SUPPQUAL Handling:")
        summary.append("  • SUPPQUAL unique datasets are expanded from SUPP_TEMP template rows.")
        summary.append("  • Generated SUPP_TEMP rows are marked Keep=1 for validation.")
        summary.append("")
        summary.append("Length Rule:")
        summary.append("  • Character length warning is raised only when dataset length is greater than spec length.")
        summary.append("")
        summary.append("Next:")
        summary.append("  • CT validation against selected CDISC CT package.")
        summary.append("  • VLM handling.")
        summary.append("  • IG metadata validation using CDISC Library API.")

        return "\n".join(summary)

    # ----------------------------------------------------------------------------------------------
    # POPULATE TABS
    # ----------------------------------------------------------------------------------------------

    def populate_spec_only_tabs(self):
        self.populate_study_tab()
        self.display_combined_spec_sheet()
        self.normalized_table.load_df(self.normalized_spec, color_keep=True)
        self.populate_ct_tab()
        self.populate_define_tab()

        self.validation_text.setPlainText("Validation not run yet. Load datasets and click Validate.")
        self.validation_table.load_df(pd.DataFrame(columns=[
            "Severity", "Check", "Dataset", "Variable", "Spec Type", "Data Type",
            "Spec Len", "Data Len", "Spec Label", "Data Label", "Message"
        ]))

    def populate_dataset_only_tabs(self):
        self.populate_study_tab()

        if hasattr(self, "dataset_folder_label"):
            self.dataset_folder_label.setText(
                f"Dataset Folder: {self.dataset_path if self.dataset_path else '-'}"
            )

        if isinstance(self.dataset_inventory, pd.DataFrame) and not self.dataset_inventory.empty:
            self.dataset_table.load_df(self.dataset_inventory)
        else:
            self.dataset_table.load_df(pd.DataFrame(columns=[
                "Dataset", "Dataset Label", "Type", "File", "Variables", "Size MB"
            ]))

        if isinstance(self.dataset_metadata, pd.DataFrame) and not self.dataset_metadata.empty:
            self.dataset_metadata_table.load_df(self.dataset_metadata)
        else:
            self.dataset_metadata_table.load_df(pd.DataFrame(columns=[
                "Dataset", "Variable", "Label", "Type", "Length",
                "Format", "Informat", "Order", "Source File"
            ]))

        self.validation_text.setPlainText("Validation not run yet. Click Validate to compare spec and dataset metadata.")
        self.validation_table.load_df(pd.DataFrame(columns=[
            "Severity", "Check", "Dataset", "Variable", "Spec Type", "Data Type",
            "Spec Len", "Data Len", "Spec Label", "Data Label", "Message"
        ]))

    def populate_all_tabs(self):
        self.populate_study_tab()
        self.display_combined_spec_sheet()
        self.normalized_table.load_df(self.normalized_spec, color_keep=True)

        if hasattr(self, "dataset_folder_label"):
            self.dataset_folder_label.setText(
                f"Dataset Folder: {self.dataset_path if self.dataset_path else '-'}"
            )

        if isinstance(self.dataset_inventory, pd.DataFrame) and not self.dataset_inventory.empty:
            self.dataset_table.load_df(self.dataset_inventory)
        else:
            self.dataset_table.load_df(pd.DataFrame(columns=[
                "Dataset", "Dataset Label", "Type", "File", "Variables", "Size MB"
            ]))

        if isinstance(self.dataset_metadata, pd.DataFrame) and not self.dataset_metadata.empty:
            self.dataset_metadata_table.load_df(self.dataset_metadata)
        else:
            self.dataset_metadata_table.load_df(pd.DataFrame(columns=[
                "Dataset", "Variable", "Label", "Type", "Length",
                "Format", "Informat", "Order", "Source File"
            ]))

        self.populate_ct_tab()
        self.populate_define_tab()
        self.populate_validation_tab()

    def populate_study_tab(self):
        lines = []
        lines.append("Clinical Metadata Studio - Module 2")
        lines.append("=" * 90)
        lines.append("")
        lines.append(f"Loaded On        : {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
        lines.append(f"Specification    : {self.spec_path}")
        lines.append(f"Dataset Folder   : {self.dataset_path if self.dataset_path else 'Not selected'}")
        lines.append("")
        included_sheets = []
        excluded_sheets = []

        for sheet, df in self.excel_data.items():
            if is_domain_metadata_sheet(sheet, df):
                included_sheets.append(sheet)
            else:
                excluded_sheets.append(sheet)

        lines.append("Included Domain Sheets:")
        for sheet in included_sheets:
            df = self.excel_data.get(sheet)
            lines.append(f"  • {sheet} ({len(df)} rows, {len(df.columns)} columns)")

        lines.append("")
        lines.append("Excluded Template / Helper Sheets:")
        for sheet in excluded_sheets:
            lines.append(f"  • {sheet}")

        lines.append("")
        lines.append(f"Normalized Metadata Rows : {len(self.normalized_spec)}")
        lines.append(f"Dataset Files Detected   : {len(self.dataset_inventory) if isinstance(self.dataset_inventory, pd.DataFrame) else 0}")
        lines.append(f"Dataset Metadata Rows    : {len(self.dataset_metadata) if isinstance(self.dataset_metadata, pd.DataFrame) else 0}")

        lines.append("")
        lines.append("Template Rule Applied:")
        lines.append("  • Column E is treated as KEEP flag where value 1 indicates variable is used/kept.")
        lines.append("  • Dataset folder scan uses ONLY the selected folder (no subfolders).")
        lines.append("  • SUPPQUAL datasets are expanded using SUPP_TEMP template rows.")
        lines.append("  • Fallback KEEP-like column names are also supported if present.")

        self.study_text.setPlainText("\n".join(lines))

    def display_combined_spec_sheet(self):
        """
        Shows one appended view of all relevant domain sheets.
        Template/support/helper sheets are excluded.
        """
        if self.normalized_spec is not None and not self.normalized_spec.empty:
            self.spec_table.load_df(self.normalized_spec, color_keep=True)

            included = sorted(set(self.normalized_spec["Dataset"])) if "Dataset" in self.normalized_spec.columns else []
            self.spec_summary_label.setText(
                f"Combined metadata view | Included domain sheets: {len(included)} | Variables: {len(self.normalized_spec)}"
            )
        else:
            self.spec_table.load_df(pd.DataFrame())
            self.spec_summary_label.setText("No domain metadata sheets detected.")

    def populate_ct_tab(self):
        lines = []
        lines.append("CT Validation Preview")
        lines.append("=" * 90)
        lines.append("")
        lines.append("This tab currently detects CT / Codelist related columns from the uploaded spec.")
        lines.append("Future version will compare these against selected CDISC CT package.")
        lines.append("")

        for sheet, df in self.excel_data.items():
            ct_cols = infer_ct_columns(df)
            if ct_cols:
                lines.append(f"Sheet: {sheet}")
                lines.append("  CT Columns: " + ", ".join([str(c) for c in ct_cols]))
                lines.append("")

        self.ct_text.setPlainText("\n".join(lines))

    def populate_define_tab(self):
        lines = []
        lines.append("Define Preview")
        lines.append("=" * 90)
        lines.append("")
        lines.append("Current Define-ready metadata source:")
        lines.append(f"  • Normalized variable rows: {len(self.normalized_spec)}")
        lines.append("")
        lines.append("Future Define.xml outputs:")
        lines.append("  • ItemGroupDefs")
        lines.append("  • ItemDefs")
        lines.append("  • CodeLists")
        lines.append("  • Value Level Metadata")
        lines.append("  • Methods")
        lines.append("  • Comments")
        lines.append("  • Origins")
        lines.append("")
        lines.append("Note: Only KEEP_FLAG=1 variables should flow to final Define.xml unless otherwise configured.")

        self.define_text.setPlainText("\n".join(lines))

    def populate_validation_tab(self):
        self.validation_summary_text = self.build_spec_dataset_validation()
        self.validation_text.setPlainText(self.validation_summary_text)

        if isinstance(self.validation_issues, pd.DataFrame) and not self.validation_issues.empty:
            self.validation_table.load_df(self.validation_issues)
        else:
            self.validation_table.load_df(pd.DataFrame(columns=[
                "Severity", "Check", "Dataset", "Variable", "Spec Type", "Data Type",
                "Spec Len", "Data Len", "Spec Label", "Data Label", "Message"
            ]))

    # ----------------------------------------------------------------------------------------------
    # EXPORT VALIDATION REPORT
    # ----------------------------------------------------------------------------------------------

    def export_validation_xlsx(self):
        if self.validation_issues is None or self.validation_issues.empty:
            QtWidgets.QMessageBox.warning(
                self,
                "No Validation Results",
                "No validation issues available to export. Please click Validate first."
            )
            return

        default_name = "metadata_validation_report.xlsx"
        out_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Export Validation Report",
            str(Path.home() / default_name),
            "Excel Files (*.xlsx)"
        )

        if not out_path:
            return

        if not out_path.lower().endswith(".xlsx"):
            out_path += ".xlsx"

        try:
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            summary_lines = self.validation_text.toPlainText().splitlines()

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

                # ------------------------------------------------------------------
                # Summary Sheet
                # ------------------------------------------------------------------
                summary_df = pd.DataFrame({"Summary": [""]})
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

                ws_summary = writer.book["Summary"]

                # Remove default dataframe content
                ws_summary.delete_rows(1, ws_summary.max_row)

                row_num = 1

                for line in summary_lines:
                    safe_line = str(line)

                    # Prevent Excel from treating text as formulas.
                    if safe_line.startswith("="):
                        safe_line = "'" + safe_line

                    cell = ws_summary.cell(row=row_num, column=1, value=safe_line)

                    # Bold section headers
                    if (
                        safe_line.startswith("Spec vs")
                        or safe_line.endswith(":")
                    ):
                        cell.font = Font(bold=True)

                    row_num += 1

                ws_summary.column_dimensions["A"].width = 120

                # ------------------------------------------------------------------
                # Issues Sheet
                # ------------------------------------------------------------------
                self.validation_issues.to_excel(writer, sheet_name="Issues", index=False)

                ws_issues = writer.book["Issues"]

                # Auto-width columns
                for col_cells in ws_issues.columns:
                    length = 0
                    col_letter = get_column_letter(col_cells[0].column)

                    for c in col_cells:
                        try:
                            length = max(length, len(str(c.value)))
                        except Exception:
                            pass

                    ws_issues.column_dimensions[col_letter].width = min(max(length + 3, 15), 80)

                # Freeze header row
                ws_issues.freeze_panes = "A2"

            QtWidgets.QMessageBox.information(
                self,
                "Export Complete",
                f"Validation report exported successfully:\n\n{out_path}"
            )

            self.set_status(f"Validation report exported: {out_path}")

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Failed", str(e))
            self.set_status("Validation export failed.")

    # ----------------------------------------------------------------------------------------------
    # CLOSE
    # ----------------------------------------------------------------------------------------------

    def closeEvent(self, event):
        # Delete only temporary SharePoint download; do not delete user's local file.
        if self.temp_downloaded_spec and os.path.exists(self.temp_downloaded_spec):
            try:
                os.remove(self.temp_downloaded_spec)
            except Exception:
                pass
        event.accept()

# ====================================================================================================
# MAIN
# ====================================================================================================

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")

    win = MetadataStudio()

    try:
        win.showMaximized()
    except Exception:
        win.show()

    sys.exit(app.exec_())
