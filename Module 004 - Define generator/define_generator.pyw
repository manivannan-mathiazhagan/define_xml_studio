# ====================================================================================================
# Script Name    : define_xml_generator.pyw
#
# Description    : Desktop GUI utility for generating CDISC Define-XML 2.0 / 2.1 from
#                  SharePoint-based SDTM / ADaM specifications and XPT datasets without
#                  SAS dependency for define.xml generation.
#
# Version        : full updated script with SharePoint metadata loading, Define-XML generation,
#                  CT/format generation, VLM automation, document linking support,
#                  ADaM/SDTM CT integration, and Define reviewer enhancements
#
#                  Key Updates / Enhancements:
#                    [1] Added SharePoint-based specification loading using get_spreadsheet.py
#                    [2] Added automatic Define-XML 2.0 / 2.1 generation
#                    [3] Added SDTM and ADaM metadata support
#                    [4] Added automatic dataset metadata extraction from XPT files
#                    [5] Added metadata editor with editable/non-editable column control
#                    [6] Added automatic KeySequence generation using ID Var
#                    [7] Added automatic CT / format generation from spec and XPT values
#                    [8] Added CDISC Library CT lookup integration
#                    [9] Added SDTM + ADaM CT combined lookup handling
#                   [10] Added automatic nearest-prior ADaM CT fallback selection
#                   [11] Added SDTM VLM generation for --ORRES and SUPP-- QVAL
#                   [12] Added ADaM PARAM/PARAMCD VLM generation logic
#                   [13] Added Define-XML MethodDef generation for Derived variables
#                   [14] Added CommentDef generation for Assigned / Predecessor variables
#                   [15] Added CRF page hyperlink support for Origin=CRF
#                   [16] Added document registry support for ADRG / CSDRG / aCRF
#                   [17] Added optional Documents and Document_Links sheet support
#                   [18] Added dataset / variable / method / comment document linking
#                   [19] Added automatic predecessor handling using spec comments
#                   [20] Added ADaM display/order variable pair handling
#                   [21] Added external dictionary support (MedDRA / WHODrug)
#                   [22] Added CT synonym decode replacement logic
#                   [23] Added Define validation checks and issue reporting
#                   [24] Added export-ready formats and VLM review grids
#                   [25] Added dataset ordering rules for ADaM Define generation
#
#                  Workflow:
#                    - Load specification from SharePoint
#                    - Load XPT datasets
#                    - Review dataset metadata
#                    - Review/edit Type / Format / Origin / Comments
#                    - Generate CT / Formats
#                    - Generate VLM metadata
#                    - Review validation findings
#                    - Generate Define-XML
#
#                  Configuration:
#                    - This tool depends on define_config.json.
#                    - Study-level details and file locations are maintained in the config JSON.
#                    - The config JSON controls items such as:
#                        * Study name / study OID
#                        * Sponsor / protocol details
#                        * SharePoint site and specification path
#                        * XPT input folder
#                        * Define output folder
#                        * CDISC Library API key / CT package date
#                        * Define version and standard selection
#                        * Document settings for aCRF, CSDRG, and ADRG
#                        * Default linked document file names and locations
#
#                  Supported Specification Sheets:
#                    - Domain sheets (SDTM / ADaM metadata)
#                    - Domains
#                    - ValueMetadata
#                    - Documents
#                    - Document_Links
#
#                  Documents Sheet columns:
#                    ID, Title, Href
#
#                  Document_Links Sheet columns:
#                    ID, Document, Pages
#
#                  ValueMetadata columns used internally:
#                    Dataset, Grouping Variable, Group Value,
#                    Where Clause, Result Variable,
#                    Length, Type, Format, Origin, Role, Comment
#
#                  Formats columns used internally:
#                    Format, Code, Decode,
#                    Codelist Code, Term Code,
#                    Source Dataset, Source Variable
#
#                  Output Files:
#                    - define.xml
#                    - CT / Formats review output
#                    - Value Level Metadata review output
#                    - Validation report
#                    - Define reviewer-linked documents
#
#                  Standards Supported:
#                    - SDTM
#                    - ADaM
#
#                  Define Versions Supported:
#                    - Define-XML 2.0
#                    - Define-XML 2.1
# ====================================================================================================

import os
import re
import numbers
import warnings

# Suppress harmless openpyxl warning for Excel data validation extensions.
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl"
)
import sys
import html
import math
import json
import subprocess
import importlib
import urllib.request
import urllib.error
import urllib.parse
import traceback
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime


def install_if_missing(package_name, import_name=None):
    module_name = import_name or package_name
    try:
        return importlib.import_module(module_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
        importlib.invalidate_caches()
        return importlib.import_module(module_name)


pd = install_if_missing("pandas", "pandas")
install_if_missing("openpyxl", "openpyxl")
install_if_missing("PyQt5", "PyQt5")
try:
    pyreadstat = install_if_missing("pyreadstat", "pyreadstat")
except Exception:
    pyreadstat = None

from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook


code = ""  # global fallback to prevent legacy bare-code NameError
# ================================================================================================
# Constants / config
# ================================================================================================

TEMP_DIR = r"C:\TEMP"
DOWNLOAD_PATH = r"C:\TEMP\define_studio_spec.xlsx"
GET_SPREADSHEET_SCRIPT = r"P:\BSP_LocalDev\_GLIB\harm_bsp_v_1_0\macros\report\api\get_spreadsheet.py"
DEFAULT_SITE_NAME = "BSP - Team Mani"
DEFAULT_SPEC_FILE_PATH = "BSP/SDTM_Specs/CSR/SDTM_Specification.xlsx"
DEFAULT_CONFIG_PATH = str(Path(__file__).resolve().parent / "define_config.json")

SPEC_COLUMNS = [
    "Dataset", "Variable", "Label", "ID Var", "Keep", "Type", "Len",
    "Control or Format", "Term", "Core", "Role", "Origin", "Comments"
]

EDITOR_COLUMNS = [
    "Dataset", "Variable", "Label", "Keep", "ID Var", "Length", "Type",
    "Format", "Origin", "Comments", "Data Type", "Data Format", "Order", "Source"
]

EDITABLE_COLUMNS = {"Type", "Format", "Origin", "Comments"}

VLM_COLUMNS = [
    "Dataset", "Grouping Variable", "Group Value", "Group Label",
    "Grouping Variable 1", "Group Value 1",
    "Grouping Variable 2", "Group Value 2",
    "Grouping Variable 3", "Group Value 3",
    "Grouping Variable 4", "Group Value 4",
    "Where Clause", "Result Variable", "Length", "Type", "Format", "Origin", "Role", "Comment"
]

FORMAT_COLUMNS = [
    "Order", "Format", "Code", "Decode",
    "Codelist Code", "Term Code",
    "Source Dataset", "Source Variable", "Decode Variable", "Sort Value"
]

DOMAIN_COLUMNS = ["Dataset", "Description", "Class", "Structure", "Purpose", "Keys", "Documentation", "Location"]
DOCUMENTS_COLUMNS = ["ID", "Title", "Href"]
DOCUMENT_LINKS_COLUMNS = ["ID", "Document", "Pages"]


DOMAIN_CLASS_ORDER = {
    "TRIAL DESIGN": 1,
    "TRIAL DOMAINS": 1,
    "SPECIAL PURPOSE": 2,
    "EVENTS": 3,
    "INTERVENTIONS": 4,
    "FINDINGS": 5,
    "FINDING": 5,
    "FINDINGS ABOUT": 6,
    "RELATIONSHIP": 7,
}

# ADaM Define dataset display order requested for ItemGroupDef: 
#   1) Subject Level, 2) Occurrence, 3) Basic Data Structure.
# Keep SDTM order unchanged using DOMAIN_CLASS_ORDER above.
ADAM_CLASS_ORDER = {
    "SUBJECT LEVEL": 1,
    "SUBJECT LEVEL ANALYSIS DATASET": 1,
    "SUBJECT-LEVEL ANALYSIS DATASET": 1,
    "ADSL": 1,
    "OCCURRENCE": 2,
    "OCCURRENCE DATA STRUCTURE": 2,
    "OCCURRENCE DATASET": 2,
    "BASIC DATA": 3,
    "BASIC DATA STRUCTURE": 3,
    "BDS": 3,
}


def supp_parent_domain(dataset):
    """Return parent domain for SUPP/SQ datasets.

    SUPPDM -> DM, SUPPAE -> AE, SQAPCM -> APCM.
    """
    ds = safe_upper(dataset)
    if ds.startswith("SUPP") and len(ds) > 4:
        return ds[4:]
    if ds.startswith("SQ") and len(ds) > 2:
        return ds[2:]
    return ""


def is_supp_qual_dataset(dataset):
    ds = safe_upper(dataset)
    return ds.startswith("SUPP") or ds.startswith("SQ")


def dataset_class_sort_value(ds, domain_lookup, standard=""):
    ds = safe_upper(ds)
    info = domain_lookup.get(ds, {}) if isinstance(domain_lookup, dict) else {}
    dclass = normalize_domain_class(info.get("Class"))
    structure = safe_upper(info.get("Structure"))
    std = safe_upper(standard)

    if std == "ADAM":
        # ADSL / subject-level datasets should always be first.
        if ds == "ADSL" or "SUBJECT LEVEL" in dclass or "SUBJECT-LEVEL" in dclass or "SUBJECT LEVEL" in structure:
            return 1

        # Occurrence datasets next. Covers common class/structure text variants.
        if "OCCURRENCE" in dclass or "OCCURRENCE" in structure:
            return 2

        # Basic Data Structure after occurrence. Covers BDS or full class text.
        if dclass in {"BDS", "BASIC DATA", "BASIC DATA STRUCTURE"} or "BASIC DATA" in dclass or "BASIC DATA" in structure:
            return 3

        return ADAM_CLASS_ORDER.get(dclass, 99)

    if is_supp_qual_dataset(ds):
        dclass = "RELATIONSHIP"
    return DOMAIN_CLASS_ORDER.get(dclass, 99)


def get_dataset_repeating_reference(dataset, domain_class="", standard=""):
    """Return Define-XML dataset-level Repeating and IsReferenceData values.

    Baseline rules:
      - SDTM Trial Design domains: Repeating=No, IsReferenceData=Yes
      - DM and ADSL: Repeating=No, IsReferenceData=No
      - All other datasets: Repeating=Yes, IsReferenceData=No

    The class check is retained so Domains-sheet metadata such as
    Class='TRIAL DESIGN' also drives the correct setting.
    """
    ds = safe_upper(dataset)
    dclass = normalize_domain_class(domain_class)
    std = safe_upper(standard)

    trial_design_domains = {"TA", "TE", "TI", "TS", "TV", "TD", "TM"}

    if std == "SDTM" and (ds in trial_design_domains or dclass in {"TRIAL DESIGN", "TRIAL DOMAINS"}):
        return "No", "Yes"

    if ds in {"DM", "ADSL"}:
        return "No", "No"

    return "Yes", "No"


def tune_table_widths(view, preferred=None, max_width=260):
    """Keep wide review grids usable by preventing Code/Where Clause columns from taking the whole screen."""
    try:
        view.resizeColumnsToContents()
        model = view.model()
        if hasattr(model, "sourceModel"):
            src = model.sourceModel()
        else:
            src = model
        cols = list(getattr(src, "df", pd.DataFrame()).columns)
        preferred = preferred or {}
        for i, col in enumerate(cols):
            width = preferred.get(col, None)
            if width is not None:
                view.setColumnWidth(i, width)
            else:
                view.setColumnWidth(i, min(view.columnWidth(i), max_width))
        view.horizontalHeader().setStretchLastSection(False)
        view.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
    except Exception:
        try:
            view.resizeColumnsToContents()
        except Exception:
            pass

TYPE_OPTIONS = ["char", "num", "date", "datetime", "float"]
ORIGIN_OPTIONS = ["CRF", "Assigned", "Derived", "Protocol", "Predecessor", "eDT", ""]

THEME = {
    "app_bg": "#f3f7fd",
    "header": "#cfe7ff",
    "header_text": "#103760",
    "subtitle": "#315f8f",
    "panel": "#ffffff",
    "border": "#b7cde8",
    "table_header": "#d8ebff",
    "status": "#fff8dc",
    "status_running": "#d9ecff",
    "status_done": "#e7f7e7",
    "status_error": "#ffe6e6",
    "status_info": "#fff8dc",
    "button": "#4d8ef7",
    "button_hover": "#2f75dd",
    "locked_bg": "#f1f3f6",
    "editable_bg": "#fffde7",
    "warn_bg": "#fff4cc",
    "err_bg": "#ffe9e9",
}


# ================================================================================================
# Utility helpers
# ================================================================================================

def safe_text(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()




def sas_best_text(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, numbers.Number) and not isinstance(value, bool):
        try:
            f = float(value)
            if f.is_integer():
                return str(int(f))
            return format(f, ".15g")
        except Exception:
            return str(value).strip()
    return str(value).strip()


def infer_codelist_datatype_from_codes(code_values, fallback="text"):
    vals = []
    has_decimal = False

    for value in code_values or []:
        txt = safe_text(value)
        if txt == "":
            continue
        try:
            num = float(txt)
        except Exception:
            return "text"

        vals.append(num)
        # Preserve intent from the displayed code as well as numeric value.
        if not num.is_integer() or re.search(r"[.]\d*[1-9]", txt):
            has_decimal = True

    if not vals:
        fb = safe_text(fallback) or "text"
        return fb if fb in {"text", "integer", "float"} else "text"

    return "float" if has_decimal else "integer"

def safe_upper(value):
    return safe_text(value).upper()


def normalize_domain_class(value):
    """Normalize Define-XML dataset class values for P21 compatibility.

    Some specs use RELATIONSHIPS, but Define/P21 expects RELATIONSHIP.
    Keep all other class values as uppercase display values.
    """
    cls = safe_upper(value)
    if cls == "RELATIONSHIPS":
        return "RELATIONSHIP"
    return cls


def normalize_keep(value):
    """Return True for KEEP values including Excel numeric 1.0.

    ADaM/SDTM specs read from Excel can display KEEP as 1.0 when the
    cell is numeric. Treat 1, 1.0, and equivalent numeric values as KEEP=1.
    """
    txt = safe_text(value)
    if not txt:
        return False
    up = txt.upper()
    if up in {"Y", "YES", "TRUE", "X", "KEEP"}:
        return True
    try:
        return float(txt) == 1.0
    except Exception:
        return False


def to_int_or_none(value):
    txt = safe_text(value)
    if txt == "":
        return None
    try:
        return int(float(txt))
    except Exception:
        return None


def normalize_dataset_type(value):
    txt = safe_upper(value)
    if txt in {"STRING", "CHAR", "CHARACTER", "OBJECT", "TEXT"}:
        return "char"
    if txt in {"DOUBLE", "FLOAT", "NUMERIC", "INTEGER", "INT", "NUMBER", "DATE", "DATETIME", "TIME"}:
        return "num"
    if "$" in txt:
        return "char"
    if txt.startswith(("BEST", "DATE", "DATETIME", "TIME", "YY", "MMDD", "DDMM", "E8601")):
        return "num"
    return txt.lower() if txt else ""


def is_support_sheet(sheet_name):
    name = safe_upper(sheet_name).replace(" ", "_")
    exclude_exact = {
        "README", "READ_ME", "STATUS", "DOMAINS", "BLANK_SPEC", "SOURCE_DATA",
        "SUPPQUAL", "SUPP_TEMP", "FORMATS", "VALUEMETADATA", "VALUE_METADATA", "VALUEMETADATA",
        "DEV_FORMATS", "QC_FORMATS", "LB_FORMATS", "QC_LB_EXT_TESTS",
        "QS_TESTCD", "LOOKUPS"
    }
    return name in exclude_exact or name.startswith("_") or "FORMAT" in name or "LOOKUP" in name


def is_domain_metadata_sheet(sheet_name, df):
    return not is_support_sheet(sheet_name) and df is not None and not df.empty and len(df.columns) >= 5


def sort_key_mixed(value):
    txt = safe_text(value)
    if txt == "":
        return (1, "")
    try:
        return (0, float(txt))
    except Exception:
        return (1, txt)


def xml_id(value):
    txt = safe_text(value).replace(" ", "_")
    txt = re.sub(r"[^A-Za-z0-9_.-]", "_", txt)
    if not txt:
        txt = "X"
    if not re.match(r"^[A-Za-z_]", txt):
        txt = "X" + txt
    return txt



def infer_predecessor_value(row, standard=""):
    """Return the text required for Define-XML Origin Type=Predecessor.

    For Origin=Predecessor, use the spec Comments cell directly as the
    predecessor value. Example: if Comments contains DM.STUDYID or AE.AEACN,
    write that exact value under def:Origin Type=Predecessor.

    Do not auto-adapt ADAE.AEACN to AE.AEACN when Comments is populated; the
    specification is the source of truth. Only when Comments is blank, fall
    back to DATASET.VARIABLE to avoid an empty predecessor value and P21 DD0061.
    """
    if not hasattr(row, "get"):
        return ""

    comment = safe_text(row.get("Comments"))
    if comment:
        return comment

    ds = safe_upper(row.get("Dataset"))
    var = safe_upper(row.get("Variable"))
    if ds and var:
        return f"{ds}.{var}"
    return var

def clean_comment_text(comment, origin=""):
    """Remove duplicated origin heading lines from comments shown by the XSL.

    The spec often stores Origin=CRF/Protocol/eDT and Comments starting again with
    CRF, CRF Page, Protocol, eDT, Assigned, etc.  Define XSL already prints the
    origin type from def:Origin, so keeping the same heading in CommentDef makes
    the browser view show duplicates such as:
        CRF
        CRF Page
        Set to ...
    This keeps the useful derivation/source text and removes only leading labels.
    """
    txt = safe_text(comment)
    if not txt:
        return ""
    org = safe_upper(origin)
    drop_exact = {"CRF", "CRF PAGE", "CRF PAGES", "PROTOCOL", "EDT", "E-DT", "ASSIGNED", "PREDECESSOR"}
    if org:
        drop_exact.add(org)
    lines = [ln.strip() for ln in str(txt).replace("\r", "\n").split("\n")]
    cleaned = []
    skipping = True
    label_re = r"^(CRF\s+PAGES?|CRF|PROTOCOL|EDT|E-DT|ASSIGNED|PREDECESSOR|DERIVED)\s*:?\s*$"
    label_with_text_re = r"^(CRF\s+PAGES?|PROTOCOL|EDT|E-DT|ASSIGNED|PREDECESSOR|DERIVED)\s*:\s*(.+)$"
    for ln in lines:
        if skipping:
            u = safe_upper(ln).strip(":")
            if not ln or u in drop_exact or re.fullmatch(label_re, ln, flags=re.I):
                continue
            # Remove leading labels like "CRF Page:" but preserve details after colon.
            m = re.match(label_with_text_re, ln, flags=re.I)
            if m:
                ln = m.group(2).strip()
        skipping = False
        cleaned.append(ln)
    return "\n".join(cleaned).strip()


def parse_crf_pages(origin):
    """Return list of page numbers in strings like 'CRF page 40', 'CRF pages 40, 41', 'CRF 40-41'."""
    txt = safe_text(origin)
    if "CRF" not in txt.upper():
        return []
    pages = []
    for start, end in re.findall(r"(\d+)\s*-\s*(\d+)", txt):
        try:
            s, e = int(start), int(end)
            if e >= s and e - s <= 50:
                pages.extend(range(s, e + 1))
        except Exception:
            pass
    cleaned = re.sub(r"\d+\s*-\s*\d+", " ", txt)
    for n in re.findall(r"\d+", cleaned):
        try:
            pages.append(int(n))
        except Exception:
            pass
    return sorted(set(pages))


def looks_like_date(series):
    vals = [safe_text(v) for v in series.dropna().tolist() if safe_text(v)]
    if not vals:
        return False
    sample = vals[:200]
    patterns = [
        r"^\d{4}-\d{2}-\d{2}$",
        r"^\d{4}-\d{2}$",
        r"^\d{4}$",
        r"^\d{2}[A-Z]{3}\d{4}$",
    ]
    hits = 0
    for v in sample:
        if any(re.match(p, v.upper()) for p in patterns):
            hits += 1
    return hits / max(len(sample), 1) >= 0.8


def looks_like_datetime(series):
    vals = [safe_text(v) for v in series.dropna().tolist() if safe_text(v)]
    if not vals:
        return False
    sample = vals[:200]
    patterns = [
        r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}",
        r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}",
    ]
    hits = 0
    for v in sample:
        if any(re.match(p, v) for p in patterns):
            hits += 1
    return hits / max(len(sample), 1) >= 0.8


def infer_type_from_values(series, storage_type="", storage_length=""):
    vals = [safe_text(v) for v in series.dropna().tolist() if safe_text(v)]
    stype = normalize_dataset_type(storage_type)

    if not vals:
        if stype == "char":
            return "char", to_int_or_none(storage_length) or 1, ""
        return "num", 8, "8"

    ser = pd.Series(vals)
    if looks_like_datetime(ser):
        return "datetime", max(len(v) for v in vals), ""
    if looks_like_date(ser):
        return "date", max(len(v) for v in vals), ""

    numeric_values = []
    numeric_ok = True
    max_decimals = 0
    max_width = 1
    for v in vals[:1000]:
        try:
            float(v)
            numeric_values.append(v)
            max_width = max(max_width, len(v))
            if "." in v:
                max_decimals = max(max_decimals, len(v.split(".")[-1].rstrip("0")))
        except Exception:
            numeric_ok = False
            break

    if numeric_ok and numeric_values:
        if max_decimals > 0:
            width = max(8, max_width)
            return "float", 8, f"{width}.{max_decimals}"
        return "num", 8, "8"

    return "char", max(len(v) for v in vals), ""


# ================================================================================================
# Spec / data loading
# ================================================================================================


def download_spec_from_sharepoint(site_name, file_path, status_callback=None):
    """Download the specification workbook from SharePoint using harmonized get_spreadsheet.py."""
    os.makedirs(TEMP_DIR, exist_ok=True)
    if os.path.exists(DOWNLOAD_PATH):
        try:
            os.remove(DOWNLOAD_PATH)
        except Exception:
            pass

    if not os.path.exists(GET_SPREADSHEET_SCRIPT):
        raise FileNotFoundError(
            "get_spreadsheet.py was not found:\n\n"
            f"{GET_SPREADSHEET_SCRIPT}"
        )

    site_name = safe_text(site_name)
    file_path = safe_text(file_path)
    if not site_name or not file_path:
        raise ValueError("SharePoint Site Name and SharePoint File Path are required.")

    if status_callback:
        status_callback("Downloading specification from SharePoint...")

    cmd = [
        sys.executable,
        GET_SPREADSHEET_SCRIPT,
        "-s", site_name,
        "-f", file_path,
        "-o", DOWNLOAD_PATH,
        "-d", "Excel",
    ]

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300, shell=False)
    if result.returncode != 0:
        raise RuntimeError(
            "SharePoint specification download failed.\n\n"
            f"Command:\n{' '.join(cmd)}\n\n"
            f"STDOUT:\n{result.stdout}\n\n"
            f"STDERR:\n{result.stderr}"
        )

    if not os.path.exists(DOWNLOAD_PATH):
        raise FileNotFoundError(
            "Downloaded specification file was not created:\n\n"
            f"{DOWNLOAD_PATH}"
        )

    return DOWNLOAD_PATH

def read_domain_sheet_a_to_m_by_position(path, sheet_name):
    headers = SPEC_COLUMNS
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame(columns=headers)
    ws = wb[sheet_name]
    rows = []
    for excel_row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        vals = list(excel_row) + [""] * 13
        # Use SAS-like numeric rendering so Excel numeric cells such as
        # KEEP=1.0, ID Var=2.0, Len=200.0 are stored/displayed as 1, 2, 200.
        # Character values are preserved as-is.
        rec = {h: sas_best_text(vals[i]) for i, h in enumerate(headers)}
        if not rec["Dataset"]:
            rec["Dataset"] = sheet_name.upper()
        if all(rec[h] == "" for h in headers):
            continue
        if not rec["Variable"] or rec["Variable"].upper() in {"VARIABLE", "VARIABLE NAME"}:
            continue
        rows.append(rec)
    try:
        wb.close()
    except Exception:
        pass
    return pd.DataFrame(rows, columns=headers)


def normalize_header_name(value):
    txt = safe_text(value).lower()
    txt = re.sub(r"[^a-z0-9]+", "_", txt).strip("_")
    aliases = {
        "domain": "Dataset",
        "dataset": "Dataset",
        "data_set": "Dataset",
        "description": "Description",
        "label": "Description",
        "dataset_label": "Description",
        "class": "Class",
        "structure": "Structure",
        "purpose": "Purpose",
        "keys": "Keys",
        "key": "Keys",
        "documentation": "Documentation",
        "document": "Documentation",
        "comment": "Documentation",
        "comments": "Documentation",
        "location": "Location",
        "xpt": "Location",
    }
    return aliases.get(txt, safe_text(value))


def read_domains_sheet(path):
    """Read the Domains sheet used to populate the Define datasets table."""
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Domains" not in wb.sheetnames and "DOMAINS" not in [s.upper() for s in wb.sheetnames]:
        try:
            wb.close()
        except Exception:
            pass
        return pd.DataFrame(columns=DOMAIN_COLUMNS)
    sheet_name = next(s for s in wb.sheetnames if s.upper() == "DOMAINS")
    ws = wb[sheet_name]

    rows_iter = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows_iter:
        try:
            wb.close()
        except Exception:
            pass
        return pd.DataFrame(columns=DOMAIN_COLUMNS)

    # Find the header row. Usually it contains Dataset/Domain and Description.
    header_idx = 0
    for i, row in enumerate(rows_iter[:20]):
        vals = [normalize_header_name(v) for v in row]
        if "Dataset" in vals and any(v in vals for v in ["Description", "Structure", "Class", "Purpose"]):
            header_idx = i
            break

    headers = [normalize_header_name(v) for v in rows_iter[header_idx]]
    data_rows = []
    for row in rows_iter[header_idx + 1:]:
        rec = {c: "" for c in DOMAIN_COLUMNS}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j] in rec:
                rec[headers[j]] = safe_text(val)
        rec["Dataset"] = safe_upper(rec.get("Dataset"))
        if not rec["Dataset"] or rec["Dataset"] in {"DATASET", "DOMAIN"}:
            continue
        rec["Class"] = normalize_domain_class(rec.get("Class"))
        # Force standard SUPP/SQ dataset metadata for the Define datasets table.
        if is_supp_qual_dataset(rec["Dataset"]):
            parent = supp_parent_domain(rec["Dataset"])
            rec["Description"] = f"Supplemental Qualifiers for {parent}" if parent else "Supplemental Qualifiers"
            rec["Class"] = "RELATIONSHIP"
            rec["Structure"] = "One record per IDVAR, IDVARVAL, and QNAM value per subject."
            rec["Purpose"] = rec.get("Purpose") or "Tabulation"
        if not rec.get("Location"):
            rec["Location"] = f"{rec['Dataset'].lower()}.xpt"
        data_rows.append(rec)

    try:
        wb.close()
    except Exception:
        pass
    return pd.DataFrame(data_rows, columns=DOMAIN_COLUMNS)



def read_value_metadata_sheet(path):
    """Read the SharePoint spec ValueMetadata sheet for user-maintained VLM.

    The sheet name is matched case-insensitively as ValueMetadata / VALUE METADATA.
    Missing expected columns are added as blanks and expected VLM columns are
    ordered first. Extra review columns, if any, are preserved after the standard
    columns.
    """
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sheet_name = ""
        for s in wb.sheetnames:
            norm = re.sub(r"[^A-Z0-9]", "", safe_upper(s))
            if norm == "VALUEMETADATA":
                sheet_name = s
                break
        try:
            wb.close()
        except Exception:
            pass
        if not sheet_name:
            return pd.DataFrame(columns=VLM_COLUMNS)

        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
        df.columns = [safe_text(c) for c in df.columns]
        # Drop fully blank rows.
        if not df.empty:
            df = df.loc[~df.apply(lambda r: all(safe_text(v) == "" for v in r), axis=1)].copy()
        for c in VLM_COLUMNS:
            if c not in df.columns:
                df[c] = ""
        ordered = VLM_COLUMNS + [c for c in df.columns if c not in VLM_COLUMNS]
        return df[ordered].copy()
    except Exception:
        return pd.DataFrame(columns=VLM_COLUMNS)



def _find_sheet_name(path, wanted_names):
    """Find a worksheet by normalized name."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        norm_wanted = {re.sub(r"[^A-Z0-9]", "", safe_upper(x)) for x in wanted_names}
        found = ""
        for s in wb.sheetnames:
            if re.sub(r"[^A-Z0-9]", "", safe_upper(s)) in norm_wanted:
                found = s
                break
        try:
            wb.close()
        except Exception:
            pass
        return found
    except Exception:
        return ""


def read_documents_sheet(path):
    """Read optional Documents sheet.

    Expected columns:
      ID | Title | Href

    The sheet is optional. Missing columns are added as blanks.
    """
    sheet_name = _find_sheet_name(path, ["Documents", "Document"])
    if not sheet_name:
        return pd.DataFrame(columns=DOCUMENTS_COLUMNS)
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
        df.columns = [safe_text(c) for c in df.columns]
        rename = {}
        for c in df.columns:
            cu = safe_upper(c).replace(" ", "_")
            if cu in {"DOCID", "DOCUMENT_ID", "DOCUMENTID", "ID"}:
                rename[c] = "ID"
            elif cu in {"TITLE", "DOCUMENT_TITLE", "DOCUMENTTITLE", "NAME"}:
                rename[c] = "Title"
            elif cu in {"HREF", "FILE", "FILENAME", "FILE_NAME", "PATH", "LOCATION"}:
                rename[c] = "Href"
        df = df.rename(columns=rename)
        for c in DOCUMENTS_COLUMNS:
            if c not in df.columns:
                df[c] = ""
        df = df[DOCUMENTS_COLUMNS].copy()
        df = df.loc[~df.apply(lambda r: all(safe_text(v) == "" for v in r), axis=1)].copy()
        df["ID"] = df["ID"].apply(safe_text)
        df["Title"] = df["Title"].apply(safe_text)
        df["Href"] = df["Href"].apply(safe_text)
        return df[df["ID"] != ""].copy()
    except Exception:
        return pd.DataFrame(columns=DOCUMENTS_COLUMNS)


def read_document_links_sheet(path):
    """Read optional Document_Links sheet.

    Expected columns:
      ID | Document | Pages

    ID examples:
      ADAE        -> dataset-level comment/document reference
      ADSL.TRTSDT -> variable MethodDef if Derived; otherwise CommentDef
    """
    sheet_name = _find_sheet_name(path, ["Document_Links", "Document Links", "DocumentLinks"])
    if not sheet_name:
        return pd.DataFrame(columns=DOCUMENT_LINKS_COLUMNS)
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
        df.columns = [safe_text(c) for c in df.columns]
        rename = {}
        for c in df.columns:
            cu = safe_upper(c).replace(" ", "_")
            if cu in {"ID", "OBJECT", "OBJECT_ID", "OBJECTID", "TARGET", "TARGET_ID"}:
                rename[c] = "ID"
            elif cu in {"DOCUMENT", "DOCID", "DOCUMENT_ID", "DOCUMENTID", "DOC"}:
                rename[c] = "Document"
            elif cu in {"PAGES", "PAGE", "PAGEREFS", "PAGE_REFS", "PAGE_REFERENCE", "PAGE_REFERENCES"}:
                rename[c] = "Pages"
        df = df.rename(columns=rename)
        for c in DOCUMENT_LINKS_COLUMNS:
            if c not in df.columns:
                df[c] = ""
        df = df[DOCUMENT_LINKS_COLUMNS].copy()
        df = df.loc[~df.apply(lambda r: all(safe_text(v) == "" for v in r), axis=1)].copy()
        df["ID"] = df["ID"].apply(lambda x: safe_upper(x).replace(" ", ""))
        df["Document"] = df["Document"].apply(safe_text)
        df["Pages"] = df["Pages"].apply(safe_text)
        return df[(df["ID"] != "") & (df["Document"] != "")].copy()
    except Exception:
        return pd.DataFrame(columns=DOCUMENT_LINKS_COLUMNS)

def read_all_sheets(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    try:
        wb.close()
    except Exception:
        pass
    out = {}
    for sheet in sheets:
        try:
            out[sheet] = read_domain_sheet_a_to_m_by_position(path, sheet)
        except Exception:
            out[sheet] = pd.DataFrame()
    return out


def get_unique_supp_datasets_from_suppqual(path):
    try:
        df = read_domain_sheet_a_to_m_by_position(path, "SUPPQUAL")
    except Exception:
        return []
    if df.empty or "Dataset" not in df.columns:
        return []
    out = []
    for val in df["Dataset"].tolist():
        ds = safe_upper(val)
        if ds.startswith("SUPP") and ds not in out:
            out.append(ds)
    return out


def build_supp_rows_from_template(path, supp_dataset):
    try:
        template = read_domain_sheet_a_to_m_by_position(path, "SUPP_TEMP")
    except Exception:
        return pd.DataFrame()
    if template.empty:
        return pd.DataFrame()
    out = template.copy()
    out["Dataset"] = supp_dataset
    out["Keep"] = "1"
    out["Variable"] = out["Variable"].apply(safe_text)
    return out[out["Variable"] != ""].copy()


def read_dataset_file(path, metadataonly=False):
    """Read SAS XPT with encoding fallback.

    Some XPT files contain Windows-1252 / Latin-1 characters such as smart dashes
    or smart quotes. Pinnacle 21 may read those files without complaint, but
    pyreadstat can fail if it tries UTF-8 first. This function retries common
    SAS/Windows encodings before giving up.
    """
    if pyreadstat is None:
        raise RuntimeError("pyreadstat is required to read XPT files.")

    p = Path(path)
    if p.suffix.lower() != ".xpt":
        raise ValueError(f"Unsupported dataset file. Only .xpt files are accepted: {path}")

    encodings_to_try = [None, "windows-1252", "latin1", "iso-8859-1"]
    errors = []

    for enc in encodings_to_try:
        try:
            kwargs = {"metadataonly": metadataonly}
            if enc is not None:
                kwargs["encoding"] = enc
            return pyreadstat.read_xport(str(p), **kwargs)

        except TypeError as e:
            # Older pyreadstat versions may not support the encoding argument.
            # If so, keep the original clear error instead of masking it.
            errors.append(f"encoding={enc or 'default'}: {e}")
            if enc is not None and "encoding" in str(e).lower():
                break
            raise

        except Exception as e:
            msg = str(e)
            errors.append(f"encoding={enc or 'default'}: {msg}")

            # Retry only for decode/codec style failures. Other read errors should surface.
            retryable = (
                isinstance(e, UnicodeDecodeError)
                or "codec can't decode" in msg.lower()
                or "invalid start byte" in msg.lower()
                or "utf-8" in msg.lower()
            )
            if retryable:
                continue
            raise

    raise RuntimeError(
        "Unable to read XPT file due to character encoding issues.\n\n"
        f"File: {p}\n\n"
        "Tried encodings: default, windows-1252, latin1, iso-8859-1\n\n"
        "Details:\n" + "\n".join(errors)
    )


# ================================================================================================
# CT / Pair rules
# ================================================================================================

def companion_decode_var(variable):
    """Return the decode variable for a code/order variable.

    Examples:
      ARMCD -> ARM
      ACTARMCD -> ACTARM
      TSPARMCD -> TSPARM
      MIPARMCD -> MIPARM
      --TESTCD -> --TEST
    """
    var = safe_upper(variable)
    if var == "ARMCD":
        return "ARM"
    if var == "ACTARMCD":
        return "ACTARM"
    if var == "ETCD":
        return "ELEMENT"
    if var.endswith("TESTCD"):
        return var[:-6] + "TEST"
    if var.endswith("PARMCD"):
        return var[:-6] + "PARM"
    if var in {"PARAMCD", "PARAMN"}:
        return "PARAM"
    if var == "PARAM":
        return "PARAM"
    if var.endswith("PARMCD"):
        return var[:-2]
    if var == "VISITNUM":
        return "VISIT"
    if var == "AVISITN":
        return "AVISIT"
    if re.match(r"^PARCAT\d+N$", var):
        return var[:-1]
    if re.match(r"^PARCAT\d+$", var):
        return var
    if var == "VISITDY":
        return "VISIT"
    if var == "QNAM":
        return "QLABEL"
    adam_dec = adam_decode_display_var(var) if "adam_decode_display_var" in globals() else ""
    if adam_dec:
        return adam_dec
    return ""


def adam_display_order_var(variable):
    """Return ADaM numeric/order companion for a character display variable.

    Examples:
      SEX -> SEXN
      RACE -> RACEN
      ETHNIC -> ETHNICN
      PARAM -> PARAMN
      TRT01P -> TRT01PN
      TRT01A -> TRT01AN
      RELGR1 -> RELGR1N
      ASEV -> ASEVN
      AESEV -> AESEVN
      AEREL -> AERELN
    """
    var = safe_upper(variable)
    if not var:
        return ""

    exact = {
        "SEX": "SEXN",
        "RACE": "RACEN",
        "ETHNIC": "ETHNICN",
        "PARAM": "PARAMN",
        "AVISIT": "AVISITN",
        "VISIT": "VISITNUM",
        "ASEV": "ASEVN",
        "AESEV": "AESEVN",
        "AEREL": "AERELN",
    }
    if var in exact:
        return exact[var]

    # TRT01P/TRT01A, TRT02P/TRT02A, etc.
    if re.match(r"^TRT\d{2}[PA]$", var):
        return var + "N"

    # RELGR1, RELGR2, etc.
    if re.match(r"^RELGR\d+$", var):
        return var + "N"

    # PARCAT1/PARCAT2/PARCAT3, etc.
    if re.match(r"^PARCAT\d+$", var):
        return var + "N"

    return ""


def adam_decode_display_var(variable):
    """Return ADaM character display companion for a numeric/order variable."""
    var = safe_upper(variable)
    if not var:
        return ""

    reverse = {
        "SEXN": "SEX",
        "RACEN": "RACE",
        "ETHNICN": "ETHNIC",
        "PARAMN": "PARAM",
        "AVISITN": "AVISIT",
        "VISITNUM": "VISIT",
        "ASEVN": "ASEV",
        "AESEVN": "AESEV",
        "AERELN": "AEREL",
    }
    if var in reverse:
        return reverse[var]

    # TRT01PN/TRT01AN -> TRT01P/TRT01A
    if re.match(r"^TRT\d{2}[PA]N$", var):
        return var[:-1]

    # RELGR1N -> RELGR1
    if re.match(r"^RELGR\d+N$", var):
        return var[:-1]

    # PARCAT1N -> PARCAT1
    if re.match(r"^PARCAT\d+N$", var):
        return var[:-1]

    return ""


def is_adam_order_display_pair(display_var, order_var):
    """True when display_var/order_var is an ADaM character/numeric value pair."""
    return adam_display_order_var(display_var) == safe_upper(order_var)


def companion_code_var(variable):
    """Return preferred code/order variable when the spec format is placed on a decode variable."""
    var = safe_upper(variable)
    if var == "ARM":
        return "ARMCD"
    if var == "ACTARM":
        return "ACTARMCD"
    if var == "ELEMENT":
        return "ETCD"
    if var.endswith("TEST") and not var.endswith("TESTCD"):
        return var + "CD"
    if var.endswith("PARM") and not var.endswith("PARMCD"):
        return var + "CD"
    if var == "PARAM":
        return "PARAMCD"
    if var.endswith("PARM") and not var.endswith("PARMCD"):
        return var + "CD"
    if var == "VISIT":
        return "VISITNUM"
    if var == "AVISIT":
        return "AVISITN"
    if re.match(r"^PARCAT\d+$", var):
        return var + "N"
    if var == "QLABEL":
        return "QNAM"
    adam_order = adam_display_order_var(var) if "adam_display_order_var" in globals() else ""
    if adam_order:
        return adam_order
    return ""


def is_define_codelist_format(format_name, variable="", var_type=""):
    """Return True only for real CT / CodeList formats.

    ISO/date/time display formats must remain display formats only.  They should
    appear in the XSL format column, but should not create CodeListRef, CodeList,
    or term-count links.  This handles variants such as ISO8601, ISO 8601,
    ISO-8601, ISO_8601, ISO 8601., E8601DA, YYMMDD10., etc.
    """
    raw = safe_text(format_name)
    fmt = safe_upper(raw).strip().rstrip(".")
    fmt_compact = re.sub(r"[^A-Z0-9]", "", fmt)
    vtype = safe_upper(var_type)
    if not fmt:
        return False

    # External dictionaries are CodeLists with ExternalCodeList child, not CodeListItems.
    if is_external_dictionary_format(fmt):
        return True

    # Pure display formats / ISO date-time formats.
    display_formats = {
        "ISO8601", "ISO", "E8601DA", "E8601DT", "E8601TM", "E8601DN",
        "DATE", "DATETIME", "TIME", "YYMMDD10", "DDMMYY10", "MMDDYY10",
        "YYMMDD", "DDMMYY", "MMDDYY", "IS8601DA", "IS8601DT",
    }
    if fmt_compact in display_formats:
        return False
    if fmt_compact.startswith(("ISO8601", "E8601", "IS8601")):
        return False

    # Date/datetime/time variable types should not create CT from display formats.
    if vtype in {"DATE", "DATETIME", "TIME"}:
        return False

    # Numeric SAS display formats: 8, 8.1, 12.3 etc.
    if re.match(r"^\d+(?:\.\d+)?$", fmt):
        return False
    if re.match(r"^(BEST|DATE|DATETIME|TIME|YYMMDD|MMDDYY|DDMMYY|E8601|ISO)", fmt):
        return False

    return True


def _dictionary_search_text(format_name):
    """Normalize a format/control value for external dictionary detection.

    Handles values like MedDRA, meddra, MEDDRA 24.0, WHODRUG, Who Drug,
    WHO-DD, WHO Drug B3MAR2025, DrugDict_F, DRUG DICT, etc.
    """
    txt = safe_upper(format_name)
    txt = re.sub(r"[^A-Z0-9]+", " ", txt)
    return f" {txt.strip()} "


def is_external_dictionary_format(format_name):
    """True for external dictionary formats that should link as dictionaries, not CT items."""
    txt = _dictionary_search_text(format_name)
    compact = txt.replace(" ", "")
    if "MEDDRA" in compact:
        return True
    who_terms = [
        "WHODRUG", "WHODD", "WHODDE", "WHODDEB3",
        "WHODICT", "WHODRUGDICTIONARY", "WHODRUGGLOBAL",
        "DRUGDICT", "DRUGDICTIONARY"
    ]
    if any(term in compact for term in who_terms):
        return True
    if " WHO " in txt and " DRUG " in txt:
        return True
    if " DRUG " in txt and " DICT " in txt:
        return True
    return False


def external_dictionary_key(format_name):
    txt = _dictionary_search_text(format_name)
    compact = txt.replace(" ", "")
    if "MEDDRA" in compact:
        return "MedDRA"
    who_terms = [
        "WHODRUG", "WHODD", "WHODDE", "WHODDEB3",
        "WHODICT", "WHODRUGDICTIONARY", "WHODRUGGLOBAL",
        "DRUGDICT", "DRUGDICTIONARY"
    ]
    if any(term in compact for term in who_terms) or (" WHO " in txt and " DRUG " in txt) or (" DRUG " in txt and " DICT " in txt):
        return "WHODRUG"
    return ""


def define_codelist_name(format_name, ct_map=None):
    """Return the actual CodeList name/OID suffix used in Define.

    IMPORTANT: keep the Define CodeList OID dataset-specific when the format is
    dataset-qualified. Example: LB.UNIT must write/reference CL.LB.UNIT.

    CT lookup is handled separately by ct_lookup_key_from_format(), which strips
    the dataset prefix only for finding CDISC CT/NCI codes.
    """
    fmt = safe_text(format_name)
    if not fmt:
        return ""
    if is_external_dictionary_format(fmt):
        return external_dictionary_key(fmt)
    return fmt


def codelist_oid_for_format(format_name, ct_map=None):
    key = external_dictionary_key(format_name)
    if key == "MedDRA":
        return "CL.MEDDRA"
    if key == "WHODRUG":
        return "CL.WHODRUG"
    fmt = define_codelist_name(format_name, ct_map)
    return f"CL.{xml_id(fmt)}"


def ct_lookup_key_from_format(format_name):
    """Return the CT codelist lookup key from a Define format name.

    Examples:
      APCM.FREQ -> FREQ
      IE.IETESTCD -> IETESTCD
      ACN -> ACN

    External dictionaries and ISO/date display formats intentionally return blank.
    """
    fmt = safe_text(format_name)
    if not fmt:
        return ""
    if is_external_dictionary_format(fmt) or is_iso_8601_format(fmt):
        return ""
    # If format is domain-qualified, use the last token after dot for CDISC CT lookup.
    return safe_upper(fmt.split(".")[-1])


def _json_get_any(obj, keys):
    if not isinstance(obj, dict):
        return ""
    for k in keys:
        if k in obj and obj[k] not in (None, ""):
            return obj[k]
    return ""


def fetch_cdisc_library_ct_map(api_key, ct_version, standard="SDTM", status_callback=None):
    """Fetch CT codelist and term NCI codes from CDISC Library.

    Returns:
      {
        "FREQ": {
            "codelist_code": "C71113",
            "terms": {"QD": "C25473", ...},
            "synonyms": {"QD": "Once Daily", ...}
        },
        ...
      }

    This is deliberately defensive because CDISC Library response shapes can vary by endpoint/version.
    If the API key/version is missing or the service cannot be reached, an empty map is returned and
    define generation continues without NCI aliases.
    """
    api_key = safe_text(api_key)
    ct_version = safe_text(ct_version)
    if not api_key or not ct_version:
        return {}

    std = "adamct" if safe_upper(standard) == "ADAM" else "sdtmct"
    package_id = f"{std}-{ct_version}"
    base = "https://library.cdisc.org/api"
    urls = [
        f"{base}/mdr/ct/packages/{urllib.parse.quote(package_id)}",
        f"{base}/mdr/ct/packages/{urllib.parse.quote(package_id)}/codelists",
    ]

    def request_json(url):
        req = urllib.request.Request(url, headers={
            "api-key": api_key,
            "Accept": "application/json",
        })
        with urllib.request.urlopen(req, timeout=45) as resp:
            return json.loads(resp.read().decode("utf-8", errors="replace"))

    def iter_dicts(x):
        if isinstance(x, dict):
            yield x
            for v in x.values():
                yield from iter_dicts(v)
        elif isinstance(x, list):
            for v in x:
                yield from iter_dicts(v)

    def looks_like_codelist(d):
        return bool(_json_get_any(d, ["submissionValue", "name", "preferredTerm", "label"])) and (
            "terms" in d or "terms" in {safe_text(k).lower(): v for k, v in d.items()}
            or "conceptId" in d or "nciCode" in d or "code" in d
        )

    out = {}
    try:
        payload = None
        last_err = None
        for url in urls:
            try:
                payload = request_json(url)
                break
            except Exception as e:
                last_err = e
                continue
        if payload is None:
            if status_callback:
                status_callback(f"CDISC Library CT lookup skipped/failed: {last_err}")
            return {}

        # Find probable codelist objects.
        for d in iter_dicts(payload):
            sv = safe_upper(_json_get_any(d, ["submissionValue", "submission_value", "name"]))
            if not sv:
                continue
            # Ignore term-level objects without nested terms when possible.
            terms_obj = d.get("terms") or d.get("_links", {}).get("terms") or d.get("concepts") or d.get("items")
            cl_code = safe_text(_json_get_any(d, ["conceptId", "nciCode", "codelistCode", "code"]))
            if sv not in out:
                out[sv] = {"codelist_code": cl_code, "terms": {}, "synonyms": {}}
            elif cl_code and not out[sv].get("codelist_code"):
                out[sv]["codelist_code"] = cl_code
            out[sv].setdefault("synonyms", {})

            # Terms may be embedded in the package response.
            if isinstance(terms_obj, list):
                for t in terms_obj:
                    if not isinstance(t, dict):
                        continue
                    code = safe_text(_json_get_any(t, ["submissionValue", "codedValue", "value", "name"]))
                    nci = safe_text(_json_get_any(t, ["conceptId", "nciCode", "termCode", "code"]))
                    if code:
                        if nci:
                            out[sv]["terms"][code] = nci
                        syn = extract_cdisc_synonyms_from_term(t)
                        if syn:
                            out[sv].setdefault("synonyms", {})[code] = syn

        # Some endpoints return codelist links, not embedded terms. Try to fetch terms for codelists found.
        for d in list(iter_dicts(payload)):
            sv = safe_upper(_json_get_any(d, ["submissionValue", "submission_value", "name"]))
            if not sv or sv not in out:
                continue
            links = d.get("_links", {}) if isinstance(d, dict) else {}
            term_href = ""
            if isinstance(links, dict):
                term_link = links.get("terms") or links.get("self")
                if isinstance(term_link, dict):
                    term_href = term_link.get("href", "")
            if term_href and not out[sv]["terms"]:
                if term_href.startswith("/"):
                    term_href = base + term_href
                try:
                    term_payload = request_json(term_href)
                    for td in iter_dicts(term_payload):
                        code = safe_text(_json_get_any(td, ["submissionValue", "codedValue", "value", "name"]))
                        nci = safe_text(_json_get_any(td, ["conceptId", "nciCode", "termCode", "code"]))
                        if code:
                            if nci:
                                out[sv]["terms"][code] = nci
                            syn = extract_cdisc_synonyms_from_term(td)
                            if syn:
                                out[sv].setdefault("synonyms", {})[code] = syn
                except Exception:
                    pass

        if status_callback:
            status_callback(f"CDISC CT lookup loaded: {len(out)} codelists")
        return out
    except Exception as e:
        if status_callback:
            status_callback(f"CDISC Library CT lookup skipped/failed: {e}")
        return {}



def merge_ct_maps(primary, fallback):
    """Merge two CDISC CT maps.

    The primary map wins when the same codelist/term exists. The fallback map
    fills missing codelists, missing codelist NCI codes, missing terms and
    missing synonyms.

    For ADaM Define-XML this lets ADaM CT be used first while SDTM CT fills
    common/shared codelists that are not present in the ADaM CT package.
    """
    out = {}

    def copy_info(info):
        if not isinstance(info, dict):
            return {"codelist_code": "", "terms": {}, "synonyms": {}}
        return {
            "codelist_code": safe_text(info.get("codelist_code")),
            "terms": dict(info.get("terms", {}) or {}),
            "synonyms": dict(info.get("synonyms", {}) or {}),
        }

    for key, info in (fallback or {}).items():
        out[safe_upper(key)] = copy_info(info)

    for key, info in (primary or {}).items():
        k = safe_upper(key)
        p = copy_info(info)
        if k not in out:
            out[k] = p
            continue

        # ADaM/primary CT should win over SDTM/fallback CT for overlapping terms.
        out[k]["codelist_code"] = p.get("codelist_code") or out[k].get("codelist_code", "")
        out[k]["terms"] = {**out[k].get("terms", {}), **p.get("terms", {})}
        out[k]["synonyms"] = {**out[k].get("synonyms", {}), **p.get("synonyms", {})}

    return out


# Known ADaM CT package effective dates available in CDISC Library.
# Used when the selected SDTM CT date does not have a matching ADaM CT package.
# Example: SDTM CT 2023-09-29 should use the nearest earlier ADaM CT 2023-06-30.
ADAM_CT_EFFECTIVE_DATES = [
    "2026-03-27",
    "2025-09-26",
    "2025-03-28",
    "2024-09-27",
    "2024-03-29",
    "2023-06-30",
    "2023-03-31",
    "2022-06-24",
    "2021-12-17",
    "2020-11-06",
]


def parse_ct_effective_date(value):
    """Return YYYY-MM-DD from CT text such as package labels or plain dates."""
    txt = safe_text(value)
    m = re.search(r"(20\d{2}-\d{2}-\d{2})", txt)
    return m.group(1) if m else ""


def choose_adam_ct_version_for_sdtm_date(sdtm_ct_version, adam_ct_version=""):
    """Choose the ADaM CT package date to use for an ADaM Define.

    If the ADaM CT date requested is unavailable, use the nearest prior ADaM CT
    effective date based on the selected SDTM CT date. This prevents CDISC
    Library 404 errors when SDTM CT has a release but ADaM CT does not.
    """
    requested = parse_ct_effective_date(adam_ct_version) or parse_ct_effective_date(sdtm_ct_version)
    if not requested:
        return ""

    available = sorted({parse_ct_effective_date(x) for x in ADAM_CT_EFFECTIVE_DATES if parse_ct_effective_date(x)})
    if requested in available:
        return requested

    prior = [d for d in available if d <= requested]
    if prior:
        return prior[-1]

    # If the selected date is earlier than the first known ADaM CT package, keep
    # the requested date so the normal CDISC lookup failure message is still meaningful.
    return requested


def fetch_define_ct_map(api_key, standard, sdtm_ct_version="", adam_ct_version="", status_callback=None):
    """Fetch the CT map needed for the selected Define standard.

    SDTM Define:
      - load SDTM CT only.

    ADaM Define:
      - load ADaM CT first;
      - load SDTM CT as fallback;
      - merge so ADaM CT wins for overlapping codelists/terms.

    The current UI has a single CT version field, so callers can pass the same
    date into both sdtm_ct_version and adam_ct_version. If separate UI/config
    fields are added later, this function already supports them.
    """
    std = safe_upper(standard)
    sdtm_ct_version = parse_ct_effective_date(sdtm_ct_version) or safe_text(sdtm_ct_version)
    requested_adam_ct_version = parse_ct_effective_date(adam_ct_version) or safe_text(adam_ct_version) or sdtm_ct_version
    adam_ct_version = choose_adam_ct_version_for_sdtm_date(sdtm_ct_version, requested_adam_ct_version)

    if std == "ADAM":
        if status_callback:
            if requested_adam_ct_version and adam_ct_version and requested_adam_ct_version != adam_ct_version:
                status_callback(
                    f"Selected SDTM CT date {sdtm_ct_version} has no matching ADaM CT package; "
                    f"using nearest prior ADaM CT {adam_ct_version}."
                )
            status_callback(f"Loading ADaM CT package {adam_ct_version} for ADaM Define...")
        adam_map = fetch_cdisc_library_ct_map(
            api_key,
            adam_ct_version,
            standard="ADAM",
            status_callback=status_callback,
        )

        if status_callback:
            status_callback("Loading SDTM CT package as ADaM Define fallback...")
        sdtm_map = fetch_cdisc_library_ct_map(
            api_key,
            sdtm_ct_version,
            standard="SDTM",
            status_callback=status_callback,
        )

        merged = merge_ct_maps(adam_map, sdtm_map)
        if status_callback:
            status_callback(
                f"Combined CT lookup loaded for ADaM Define: "
                f"{len(adam_map or {})} ADaM codelists, "
                f"{len(sdtm_map or {})} SDTM fallback codelists, "
                f"{len(merged or {})} total codelists"
            )
        return merged

    if status_callback:
        status_callback("Loading SDTM CT package for SDTM Define...")
    return fetch_cdisc_library_ct_map(
        api_key,
        sdtm_ct_version,
        standard="SDTM",
        status_callback=status_callback,
    )



def extract_cdisc_synonyms_from_term(term_obj):
    """Extract CDISC Synonyms from a CDISC Library term object.

    CDISC Library payloads can store this under different keys and sometimes as
    list values. Return a semicolon-separated display string.
    """
    if not isinstance(term_obj, dict):
        return ""

    keys = [
        "cdiscSynonyms", "cdisc_synonyms", "CDISCSynonyms", "CDISC Synonyms",
        "synonyms", "Synonyms", "synonym", "Synonym"
    ]

    for key in keys:
        if key in term_obj and term_obj[key] not in (None, ""):
            val = term_obj[key]
            if isinstance(val, list):
                out = []
                for x in val:
                    if isinstance(x, dict):
                        txt = safe_text(
                            x.get("name")
                            or x.get("value")
                            or x.get("synonym")
                            or x.get("text")
                        )
                    else:
                        txt = safe_text(x)
                    if txt:
                        out.append(txt)
                return "; ".join(out)
            return safe_text(val)

    # Fallback: search nested dicts/lists for synonym-like keys.
    for _, val in term_obj.items():
        if isinstance(val, dict):
            got = extract_cdisc_synonyms_from_term(val)
            if got:
                return got
        elif isinstance(val, list):
            for item in val:
                if isinstance(item, dict):
                    got = extract_cdisc_synonyms_from_term(item)
                    if got:
                        return got

    return ""


def companion_order_var(variable, columns):
    cols = {safe_upper(c): c for c in columns}
    var = safe_upper(variable)
    candidates = []
    if var in {"QNAM", "QLABEL"}:
        candidates = ["QNAM"]
    elif var in {"PARAMCD", "PARAM"}:
        candidates = ["PARAMN"]
    elif var in {"ARMCD", "ARM"}:
        candidates = ["ARMCD"]
    elif var in {"ACTARMCD", "ACTARM"}:
        candidates = ["ACTARMCD"]
    elif var in {"ETCD", "ELEMENT"}:
        candidates = ["ETCD"]
    elif re.match(r"^PARCAT\d+$", var):
        candidates = [var + "N"]
    elif re.match(r"^PARCAT\d+N$", var):
        candidates = [var]
    elif var == "VISIT":
        candidates = ["VISITNUM"]
    elif var == "VISITNUM":
        candidates = ["VISITNUM"]
    elif var == "AVISIT":
        candidates = ["AVISITN"]
    elif var == "AVISITN":
        candidates = ["AVISITN"]
    for c in candidates:
        if c in cols:
            return cols[c]
    return ""




def is_iso_8601_format(value):
    """True for ISO 8601 / E8601 display formats; these are not CodeLists."""
    txt = safe_upper(value).strip().rstrip(".")
    compact = re.sub(r"[^A-Z0-9]", "", txt)
    return (
        compact in {"ISO", "ISO8601", "E8601DA", "E8601DT", "E8601TM", "E8601DN", "IS8601DA", "IS8601DT"}
        or compact.startswith(("ISO8601", "E8601", "IS8601"))
    )

def is_dtc_iso_variable(variable, fmt):
    """DTC variables with ISO/E8601 display format should be date/datetime with ISO display format.

    Keep the XPT length as-is (even when blank XPT metadata gives length 1).
    """
    return safe_upper(variable).endswith("DTC") and is_iso_8601_format(fmt)

def default_format_for_variable(variable, current_format=""):
    """Return only the spec/GUI supplied format; do not auto-infer CT.

    Blank Control/Format in the spec must remain blank in Define.
    This prevents automatic codelist links such as VISITNUM/VISIT, ARM/ARMCD,
    TEST/TESTCD, etc., unless the format was explicitly supplied in the
    SharePoint spec or edited by the user in the GUI.
    """
    return safe_text(current_format)


def should_skip_decode_side_format_row(format_name, source_variable):
    """Avoid duplicate CT rows from decode-side variables.

    Example:
      ETCD/ELEMENT pair should create ETCD rows from ETCD codes with ELEMENT decode.
      It should not also create rows where Code=ELEMENT and Decode=ELEMENT.
    """
    fmt = safe_upper(format_name)
    src = safe_upper(source_variable)
    decode_side_vars = {
        "ELEMENT": "ETCD",
        "ARM": "ARMCD",
        "ACTARM": "ACTARMCD",
        "TSPARM": "TSPARMCD",
        "MIPARM": "MIPARMCD",
        "VISIT": "VISITNUM",
        "AVISIT": "AVISITN",
    }
    if src in decode_side_vars and fmt == decode_side_vars[src]:
        return True
    if src.endswith("TEST") and fmt == src + "CD":
        return True
    if src.endswith("PARM") and fmt == src + "CD":
        return True
    return False



def pick_decode_with_ct_synonym(existing_decode, code, ct_row=None):
    """Return Decode using CT Synonyms first, preserving existing Decode otherwise.

    Priority:
      1. CT Synonyms, when present/non-blank
      2. existing Decode already assigned by spec/data/user
      3. code value as final fallback
    """
    synonym = ""
    if ct_row is not None:
        try:
            for col in [
                "CDISC Synonyms", "CDISC Synonym", "Cdisc Synonyms", "cdisc_synonyms",
                "Synonyms", "Synonym", "synonyms", "synonym"
            ]:
                if col in ct_row:
                    synonym = safe_text(ct_row.get(col))
                    if synonym:
                        break
        except Exception:
            pass
    current_decode = safe_text(existing_decode)
    current_code = safe_text(code)
    return synonym or current_decode or current_code


def apply_ct_synonym_decode_rule(formats_df, ct_lookup=None):
    """Apply CT synonym decode rule to the Formats/CT table.

    Non-destructive:
      - CT Synonyms replaces Decode only when non-blank.
      - Blank/missing Synonyms preserves existing Decode.
      - No CT match preserves existing Decode.
    """
    if formats_df is None or getattr(formats_df, "empty", True):
        return formats_df
    df = formats_df.copy()
    if "Decode" not in df.columns:
        df["Decode"] = ""
    if "Code" not in df.columns or "Format" not in df.columns:
        return df

    lookup = {}
    if isinstance(ct_lookup, pd.DataFrame) and not ct_lookup.empty:
        fmt_cols = [c for c in ["Format", "Codelist", "Codelist Code", "CodeList", "CodeList Code"] if c in ct_lookup.columns]
        code_cols = [c for c in ["Code", "Submission Value", "CodedValue", "Coded Value"] if c in ct_lookup.columns]
        if fmt_cols and code_cols:
            fmt_col = fmt_cols[0]
            code_col = code_cols[0]
            for _, r in ct_lookup.iterrows():
                lookup[(safe_upper(r.get(fmt_col)), safe_text(r.get(code_col)))] = r

    def _new_decode(row):
        fmt = safe_text(row.get("Format"))
        code = safe_text(row.get("Code"))
        existing = safe_text(row.get("Decode"))
        ct_row = None
        # If enrichment added Synonyms directly to the same formats row, use it.
        if any(c in row.index for c in ["CDISC Synonyms", "CDISC Synonym", "cdisc_synonyms", "Synonyms", "Synonym"]):
            ct_row = row
        else:
            ct_row = lookup.get((safe_upper(fmt), code))
        return pick_decode_with_ct_synonym(existing, code, ct_row)

    df["Decode"] = df.apply(_new_decode, axis=1)
    return df



def normalize_ct_name_for_match(value):
    """Normalize format/codelist names for CT matching.

    APCM.FREQ -> FREQ
    AE.AESEV -> AESEV
    AESEV -> AESEV
    """
    txt = safe_upper(value)
    if "." in txt:
        txt = txt.split(".")[-1]
    return txt.strip()


def apply_cdisc_synonyms_from_ct_table(formats_df, ct_df):
    """Use CDISC Synonyms from the loaded CT table as Decode in formats_df.

    Match rules:
      - Format can be AESEV or AE.AESEV or APCM.FREQ; use text after dot.
      - Match Format to CT Codelist Name / Codelist Code / Codelist Extensible name where possible.
      - Match Code to CT Submission Value.
      - If CDISC Synonyms exists and is nonblank, it replaces Decode.
      - If not, preserve existing Decode.
    """
    if formats_df is None or getattr(formats_df, "empty", True):
        return formats_df
    if ct_df is None or getattr(ct_df, "empty", True):
        return formats_df

    df = formats_df.copy()
    ct = ct_df.copy()

    # Identify CT columns flexibly.
    ct_cols_upper = {safe_upper(c): c for c in ct.columns}
    fmt_candidates = [
        "FORMAT", "CODELIST NAME", "CODELIST", "CODELIST CODE", "CODELISTCODE",
        "NCI CODE", "CODELIST SUBMISSION VALUE", "SUBMISSION VALUE"
    ]
    code_candidates = ["SUBMISSION VALUE", "SUBMISSIONVALUE", "CODED VALUE", "CODEDVALUE", "CODE"]
    syn_candidates = ["CDISC SYNONYMS", "CDISC SYNONYM", "SYNONYMS", "SYNONYM"]

    codelist_name_col = None
    codelist_code_col = None
    submission_col = None
    synonym_col = None

    for c in ct.columns:
        cu = safe_upper(c).replace("_", " ")
        if cu in {"CODELIST NAME", "CODELIST"} and codelist_name_col is None:
            codelist_name_col = c
        if cu in {"CODELIST CODE", "CODELISTCODE", "NCI CODE"} and codelist_code_col is None:
            codelist_code_col = c
        if cu in {"SUBMISSION VALUE", "SUBMISSIONVALUE", "CODED VALUE", "CODEDVALUE"} and submission_col is None:
            submission_col = c
        if cu in {"CDISC SYNONYMS", "CDISC SYNONYM", "SYNONYMS", "SYNONYM"} and synonym_col is None:
            synonym_col = c

    if submission_col is None or synonym_col is None:
        return df

    # Build lookup mainly by submission value + possible codelist identifiers.
    lookup_by_code = {}
    lookup_by_fmt_code = {}

    for _, r in ct.iterrows():
        subval = safe_text(r.get(submission_col))
        syn = safe_text(r.get(synonym_col))
        if not subval or not syn:
            continue

        # General fallback by code only, useful when code is unique enough in the displayed package.
        lookup_by_code.setdefault(safe_upper(subval), syn)

        possible_fmt_keys = []
        if codelist_name_col:
            possible_fmt_keys.append(normalize_ct_name_for_match(r.get(codelist_name_col)))
        if codelist_code_col:
            possible_fmt_keys.append(normalize_ct_name_for_match(r.get(codelist_code_col)))

        # acronym-like final token from codelist name if possible.
        # Example codelist name "Severity/Intensity Scale for Adverse Events" won't become AESEV,
        # so direct format-specific match may fail; code-only fallback handles such cases.
        for fk in possible_fmt_keys:
            if fk:
                lookup_by_fmt_code[(fk, safe_upper(subval))] = syn

    def choose(row):
        fmt_key = normalize_ct_name_for_match(row.get("Format"))
        code_key = safe_upper(row.get("Code"))
        existing = safe_text(row.get("Decode"))
        syn = lookup_by_fmt_code.get((fmt_key, code_key)) or lookup_by_code.get(code_key)
        return syn or existing or safe_text(row.get("Code"))

    df["Decode"] = df.apply(choose, axis=1)
    return df



def apply_cdisc_synonyms_by_term_code(formats_df, ct_df):
    """Use CDISC Synonyms by Term Code/NCI Code match."""
    if formats_df is None or getattr(formats_df, "empty", True):
        return formats_df
    if ct_df is None or getattr(ct_df, "empty", True):
        return formats_df
    df = formats_df.copy()
    if "Term Code" not in df.columns:
        return df

    term_col = None
    syn_col = None
    for c in ct_df.columns:
        cu = safe_upper(c).replace("_", " ")
        if cu in {"CODE", "TERM CODE", "NCI CODE", "TERM NCI CODE"} and term_col is None:
            term_col = c
        if cu in {"CDISC SYNONYMS", "CDISC SYNONYM", "SYNONYMS", "SYNONYM"} and syn_col is None:
            syn_col = c
    if term_col is None or syn_col is None:
        return df

    mp = {}
    for _, r in ct_df.iterrows():
        k = safe_text(r.get(term_col))
        v = safe_text(r.get(syn_col))
        if k and v:
            mp[k] = v

    df["Decode"] = df.apply(
        lambda r: mp.get(safe_text(r.get("Term Code"))) or safe_text(r.get("Decode")) or safe_text(r.get("Code")),
        axis=1
    )
    return df



def suppress_decode_if_same_for_display_codelist(format_name, code, decode, source_variable=""):
    """Suppress Decode when Code and Decode are identical for display-text variables.

    Uses Source Variable first, then normalized Format name, because formats may be domain-qualified:
      AE.VISIT, CE.VISIT, FA.VISIT -> variable VISIT should suppress same decode.

    Applies to display/text-side variables:
      TEST, PARM, PARAM, VISIT, ARM, ACTARM, --OBJ, --TRT, QLABEL

    Does NOT apply to coded/order variables:
      TESTCD, PARMCD, PARAMCD, VISITNUM, AVISITN, ARMCD, ACTARMCD, QNAM
    """
    fmt_raw = safe_upper(format_name)
    fmt = fmt_raw.split(".")[-1] if "." in fmt_raw else fmt_raw
    src = safe_upper(source_variable)
    c = safe_text(code)
    d = safe_text(decode)

    if not c or not d or c != d:
        return d

    suppress_exact = {"PARAM", "VISIT", "ARM", "ACTARM", "QLABEL"}
    suppress_suffix = ("TEST", "PARM", "OBJ", "TRT")

    coded_suffix = ("TESTCD", "PARMCD", "PARAMCD", "OBJCD", "TRTCD")
    coded_exact = {"VISITNUM", "AVISITN", "ARMCD", "ACTARMCD", "QNAM"}

    check_values = [src, fmt]

    if any(v in coded_exact or v.endswith(coded_suffix) for v in check_values if v):
        return d

    if any(v in suppress_exact or v.endswith(suppress_suffix) for v in check_values if v):
        return ""

    return d


def apply_display_decode_suppression(formats_df):
    """Apply same-code/decode suppression to display-text codelists."""
    if formats_df is None or getattr(formats_df, "empty", True):
        return formats_df
    df = formats_df.copy()
    if {"Format", "Code", "Decode"}.issubset(df.columns):
        df["Decode"] = df.apply(
            lambda r: suppress_decode_if_same_for_display_codelist(
                r.get("Format"),
                r.get("Code"),
                r.get("Decode"),
                r.get("Source Variable", "")
            ),
            axis=1
        )
    return df

def apply_display_decode_suppression(formats_df):
    """Apply same-code/decode suppression to display-text codelists."""
    if formats_df is None or getattr(formats_df, "empty", True):
        return formats_df
    df = formats_df.copy()
    if {"Format", "Code", "Decode"}.issubset(df.columns):
        df["Decode"] = df.apply(
            lambda r: suppress_decode_if_same_for_display_codelist(
                r.get("Format"),
                r.get("Code"),
                r.get("Decode"),
                r.get("Source Variable", "")
            ),
            axis=1
        )
    return df


# Backward-compatible old function name
def apply_test_decode_suppression(formats_df):
    return apply_display_decode_suppression(formats_df)


def is_qnam_qlabel_pair(format_name, source_variable="", decode_variable=""):
    """True when QNAM should use QLABEL as decode text."""
    fmt = safe_upper(format_name)
    src = safe_upper(source_variable)
    dec = safe_upper(decode_variable)
    return fmt == "QNAM" or src == "QNAM" or dec == "QLABEL"



def is_custom_extended_term(row):
    """Return True when term is sponsor/custom and should carry def:ExtendedValue='Yes'.

    P21 DD0029 is raised when CodeListItem/EnumeratedItem has no NCI alias/term code
    and no def:ExtendedValue.  If Term Code is missing, treat it as an extended value.
    """
    return safe_text(row.get("Term Code")) == ""


KNOWN_CODELIST_NCI_CODES = {
    # Frequently used SDTM/ADaM codelists. Used as a safety fallback when CDISC Library
    # lookup is unavailable or the spec format is domain-qualified, e.g. DM.ARMNRS.
    "YNULL": "C66742",
    "ARMNRS": "C142179",
    # ADaM CT fallbacks. These also resolve dataset-qualified formats such as
    # ADEFF.DTYPE and ADLB.DTYPE through normalize_ct_lookup_key().
    "DTYPE": "C81224",
    "SBJTSTAT": "C124296",
}


KNOWN_TERM_NCI_CODES = {
    # C66742 / NY / YNULL - No Yes Response
    "YNULL": {
        "Y": "C49488",
        "YES": "C49488",
        "N": "C49487",
        "NO": "C49487",
        "NA": "C48660",
        "NOT APPLICABLE": "C48660",
        "U": "C17998",
        "UNK": "C17998",
        "UNKNOWN": "C17998",
    },
    "NY": {
        "Y": "C49488",
        "YES": "C49488",
        "N": "C49487",
        "NO": "C49487",
        "NA": "C48660",
        "NOT APPLICABLE": "C48660",
        "U": "C17998",
        "UNK": "C17998",
        "UNKNOWN": "C17998",
    },
    # C142179 / ARMNULRS / ARMNRS - Arm Null Reason
    "ARMNRS": {
        "ASSIGNED, NOT TREATED": "C142238",
        "NOT ASSIGNED": "C142239",
        "SCREEN FAILURE": "C49628",
        "UNPLANNED TREATMENT": "C142240",
        "RANDOMIZED BY MISTAKE": "C139237",
        "REQUIRES PROHIBITED MEDICATION": "C191339",
        "RECOVERY": "C25746",
    },
    "ARMNULRS": {
        "ASSIGNED, NOT TREATED": "C142238",
        "NOT ASSIGNED": "C142239",
        "SCREEN FAILURE": "C49628",
        "UNPLANNED TREATMENT": "C142240",
        "RANDOMIZED BY MISTAKE": "C139237",
        "REQUIRES PROHIBITED MEDICATION": "C191339",
        "RECOVERY": "C25746",
    },
    # ADaM Derivation Type (DTYPE), codelist C81224. Keep submission-value
    # matching case-sensitive to avoid hiding casing problems.
    "DTYPE": {
        "AVERAGE": "C81209",
        "BC": "C92225",
        "BLOCF": "C81201",
        "BOC": "C132340",
        "BOCF": "C92226",
        "COPY": "C184383",
        "EXTRAP": "C139176",
        "HALFLLOQ": "C170546",
        "INTERP": "C81208",
        "LLOD": "C105701",
        "LLOQ": "C170543",
        "LOCF": "C81198",
        "LOV": "C132341",
        "LVPD": "C132342",
        "MAXIMUM": "C82868",
        "MINIMUM": "C82867",
        "ML": "C53331",
        "MOTH": "C81204",
        "MOV": "C81207",
        "NOCB": "C204584",
        "PHANTOM": "C170545",
        "POCF": "C81205",
        "SOCF": "C81200",
        "ULOD": "C174264",
        "ULOQ": "C170544",
        "WC": "C81203",
        "WOC": "C132343",
        "WOCF": "C81199",
        "WOV": "C81206",
    },
    # ADaM Subject Trial Status (SBJTSTAT), codelist C124296.
    "SBJTSTAT": {
        "COMPLETED": "C25250",
        "DISCONTINUED": "C25484",
        "ONGOING": "C53279",
    },
}


def normalize_ct_lookup_key(value):
    """Normalize format/codelist name for CDISC CT lookup.

    Examples:
      DM.ARMNRS -> ARMNRS
      AE.AESEV  -> AESEV
      YNULL     -> YNULL
    """
    txt = safe_upper(value)
    if "." in txt:
        txt = txt.split(".")[-1]
    return txt.strip()


def known_codelist_nci_code(format_name):
    """Return fallback codelist-level NCI code for known CT codelists."""
    return KNOWN_CODELIST_NCI_CODES.get(normalize_ct_lookup_key(format_name), "")


def codelist_alias_code(fdf, format_name=""):
    """Return first nonblank codelist NCI code from a formats dataframe.

    Falls back to known standard CT codelist codes when the workbook/CDISC API did
    not populate Codelist Code. This prevents P21 DD0031 for standard codelists
    such as YNULL and domain-qualified ARMNRS values like DM.ARMNRS.
    """
    if fdf is not None and not getattr(fdf, "empty", True) and "Codelist Code" in fdf.columns:
        for v in fdf["Codelist Code"].tolist():
            txt = safe_text(v)
            if txt:
                return txt
    return known_codelist_nci_code(format_name)


def add_nci_alias(parent, nci_code, q_func):
    """Add Define-XML NCI Alias when code is available."""
    nci_code = safe_text(nci_code)
    if nci_code:
        ET.SubElement(parent, q_func("Alias"), {
            "Context": "nci:ExtCodeID",
            "Name": nci_code
        })


def known_term_nci_code(format_name, code_value):
    """Return fallback term-level NCI code for known CT terms.

    Term submission values are intentionally matched case-sensitively.
    Example: UNIT value ``tsp`` and ``Tsp`` must not be treated as the same term,
    because Pinnacle/CDISC term checks are case-sensitive for submission values.
    Codelist names are still normalized, so DM.ARMNRS can resolve to ARMNRS.
    """
    fmt_key = normalize_ct_lookup_key(format_name)
    code_key = safe_text(code_value)
    if not fmt_key or not code_key:
        return ""

    candidates = [fmt_key]
    if fmt_key == "YNULL":
        candidates.append("NY")
    if fmt_key == "ARMNRS":
        candidates.append("ARMNULRS")

    for key in candidates:
        terms = KNOWN_TERM_NCI_CODES.get(key, {})
        # Exact/case-sensitive match only. Do not use upper/lower fallback here.
        if code_key in terms and safe_text(terms.get(code_key)):
            return safe_text(terms.get(code_key))
    return ""


def get_ct_info_generic(ct_map, format_name):
    """Return CT info from CDISC map using normalized keys.

    This is generic backend logic, not case/value-specific handling.
    """
    if not isinstance(ct_map, dict):
        return {}
    key = normalize_ct_lookup_key(format_name)
    if not key:
        return {}
    return (
        ct_map.get(key)
        or ct_map.get(safe_upper(format_name))
        or ct_map.get(safe_text(format_name))
        or {}
    )


def get_ct_term_code_generic(ct_info, code_value):
    """Return term NCI code from CT info using exact, case-sensitive matching.

    Do not case-fold CT submission values. If data/spec has ``tsp`` but CT has
    ``Tsp`` with a term NCI code, this must remain unmatched and be treated as
    an extended/non-standard value unless the source value is corrected.
    """
    if not isinstance(ct_info, dict):
        return ""
    terms = ct_info.get("terms", {}) or {}
    raw_code = safe_text(code_value)
    if not raw_code:
        return ""

    if raw_code in terms and safe_text(terms.get(raw_code)):
        return safe_text(terms.get(raw_code))

    return ""




def get_ct_standard_code_generic(ct_info, code_value):
    """Return CT submission value only when it is an exact match.

    This function intentionally does not correct casing. Case mismatches should
    remain visible in validation instead of silently assigning the standard NCI code.
    """
    if not isinstance(ct_info, dict):
        return ""
    terms = ct_info.get("terms", {}) or {}
    raw_code = safe_text(code_value)
    if raw_code and raw_code in terms and safe_text(terms.get(raw_code)):
        return raw_code
    return ""

def get_ct_synonym_generic(ct_info, code_value):
    """Return CDISC synonym from CT info using case-insensitive matching."""
    if not isinstance(ct_info, dict):
        return ""
    synonyms = ct_info.get("synonyms", {}) or {}
    raw_code = safe_text(code_value)
    if not raw_code:
        return ""
    if raw_code in synonyms and safe_text(synonyms.get(raw_code)):
        return safe_text(synonyms.get(raw_code))
    code_u = safe_upper(raw_code)
    for k, v in synonyms.items():
        if safe_upper(k) == code_u and safe_text(v):
            return safe_text(v)
    return ""


def standardize_format_for_define(format_name, ct_map=None):
    """Return the Define codelist name to write/reference.

    Keep dataset-qualified codelists exactly as entered in the spec, e.g.
    APCM.FREQ, APCM.UNIT, APCM.ROUTE.  CT/NCI lookup is still normalized
    elsewhere through ct_lookup_key_from_format(), so APCM.FREQ can use FREQ
    controlled terminology without changing the Define CodeList OID/name.

    This prevents OD0048 where ItemDef references CL.APCM.FREQ but the
    generated CodeList was normalized and written only as CL.FREQ.
    """
    fmt = safe_text(format_name)
    if not fmt:
        return ""
    if "." in fmt:
        return fmt
    key = normalize_ct_lookup_key(fmt)
    if isinstance(ct_map, dict) and key and key in ct_map:
        return key
    return fmt



def normalize_format_for_presence_check(format_name, ct_map=None):
    """Normalize format names for comparison between metadata and generated formats.

    Examples:
      DM.ARM    -> ARM if ARM exists in generated or CT context
      CM.UNIT   -> UNIT
      CM.FREQ   -> FREQ
      AE.VISIT  -> VISIT
      ISO 8601  -> ISO 8601
    """
    fmt = safe_text(format_name)
    if not fmt:
        return ""
    if is_external_dictionary_format(fmt) or is_iso_8601_format(fmt):
        return safe_upper(fmt)
    if "." in fmt:
        return safe_upper(fmt.split(".")[-1])
    return safe_upper(fmt)


def format_presence_keys(format_name, ct_map=None):
    """Return all equivalent keys for checking whether a format exists."""
    fmt = safe_text(format_name)
    if not fmt:
        return set()
    keys = {safe_upper(fmt), normalize_format_for_presence_check(fmt, ct_map)}
    std = standardize_format_for_define(fmt, ct_map) if "standardize_format_for_define" in globals() else fmt
    keys.add(safe_upper(std))
    keys.add(normalize_format_for_presence_check(std, ct_map))
    return {k for k in keys if k}


# ================================================================================================
# Table models
# ================================================================================================

class PandasTableModel(QtCore.QAbstractTableModel):
    def __init__(self, df=None, editable_columns=None):
        super().__init__()
        self.df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
        self.editable_columns = set(editable_columns or [])

    def rowCount(self, parent=None):
        return len(self.df)

    def columnCount(self, parent=None):
        return len(self.df.columns)

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if not index.isValid():
            return None
        col = self.df.columns[index.column()]
        val = safe_text(self.df.iat[index.row(), index.column()])
        if role in (QtCore.Qt.DisplayRole, QtCore.Qt.EditRole):
            return val
        if role == QtCore.Qt.BackgroundRole:
            if col in self.editable_columns:
                return QtGui.QColor(THEME["editable_bg"])
            return QtGui.QColor(THEME["locked_bg"])
        if role == QtCore.Qt.ToolTipRole:
            return val
        return None

    def headerData(self, section, orientation, role=QtCore.Qt.DisplayRole):
        if role != QtCore.Qt.DisplayRole:
            return None
        if orientation == QtCore.Qt.Horizontal:
            return str(self.df.columns[section])
        return str(section + 1)

    def flags(self, index):
        if not index.isValid():
            return QtCore.Qt.NoItemFlags
        col = self.df.columns[index.column()]
        flags = QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable
        if col in self.editable_columns:
            flags |= QtCore.Qt.ItemIsEditable
        return flags

    def setData(self, index, value, role=QtCore.Qt.EditRole):
        if role == QtCore.Qt.EditRole and index.isValid():
            col = self.df.columns[index.column()]
            if col in self.editable_columns:
                self.df.iat[index.row(), index.column()] = safe_text(value)
                self.dataChanged.emit(index, index, [QtCore.Qt.DisplayRole, QtCore.Qt.EditRole])
                return True
        return False

    def set_df(self, df):
        self.beginResetModel()
        self.df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
        self.endResetModel()


class ComboDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, options, parent=None):
        super().__init__(parent)
        self.options = options

    def createEditor(self, parent, option, index):
        combo = QtWidgets.QComboBox(parent)
        combo.addItems(self.options)
        combo.setEditable(True)
        return combo

    def setEditorData(self, editor, index):
        value = index.model().data(index, QtCore.Qt.EditRole)
        i = editor.findText(value)
        if i >= 0:
            editor.setCurrentIndex(i)
        else:
            editor.setEditText(value)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), QtCore.Qt.EditRole)


# ================================================================================================
# Define generator
# ================================================================================================

def significant_digits_from_format(format_name, default="8"):
    """Return a safe SignificantDigits value for float ItemDefs.

    If the format has decimals (e.g. 8.2), use the decimal count. Otherwise use
    a conservative default so OD0071 is not triggered.
    """
    txt = safe_text(format_name).rstrip(".")
    m = re.search(r"\b\d+\.(\d+)\b", txt)
    if m:
        try:
            n = int(m.group(1))
            return str(max(n, 1))
        except Exception:
            pass
    return safe_text(default) or "8"


class DefineXmlWriter:
    def __init__(self, standard, define_version, ig_version, ct_version, study_oid, study_name,
                 protocol, metadata_df, formats_df, vlm_df, out_dir, include_acrf=True,
                 include_rg=True, odm_version="1.3.2", study_description="",
                 meddra_version="", whodrug_version="", acrf_file="acrf.pdf",
                 csdrg_file="csdrg.pdf", adrg_file="adrg.pdf", domains_df=None, documents_df=None, document_links_df=None):

        self.standard = safe_upper(standard)
        self.define_version = safe_text(define_version) or "2.0"
        self.ig_version = safe_text(ig_version)
        self.ct_version = safe_text(ct_version)
        self.odm_version = safe_text(odm_version) or "1.3.2"
        self.study_oid = safe_text(study_oid) or "StudyOID"
        self.study_name = safe_text(study_name) or self.study_oid
        self.study_description = safe_text(study_description) or self.study_name
        self.protocol = safe_text(protocol) or self.study_oid
        self.meddra_version = safe_text(meddra_version)
        self.whodrug_version = safe_text(whodrug_version)
        self.include_acrf = include_acrf
        self.include_rg = include_rg
        self.acrf_file = safe_text(acrf_file) or "acrf.pdf"
        self.csdrg_file = safe_text(csdrg_file) or "csdrg.pdf"
        self.adrg_file = safe_text(adrg_file) or "adrg.pdf"
        self.metadata_df = metadata_df.copy()
        self.formats_df = formats_df.copy() if isinstance(formats_df, pd.DataFrame) else pd.DataFrame(columns=FORMAT_COLUMNS)
        self.vlm_df = vlm_df.copy() if isinstance(vlm_df, pd.DataFrame) else pd.DataFrame(columns=VLM_COLUMNS)
        self.domains_df = domains_df.copy() if isinstance(domains_df, pd.DataFrame) else pd.DataFrame(columns=DOMAIN_COLUMNS)
        self.documents_df = documents_df.copy() if isinstance(documents_df, pd.DataFrame) else pd.DataFrame(columns=DOCUMENTS_COLUMNS)
        self.document_links_df = document_links_df.copy() if isinstance(document_links_df, pd.DataFrame) else pd.DataFrame(columns=DOCUMENT_LINKS_COLUMNS)
        self.document_registry = self.build_document_registry()
        self.domain_lookup = {}
        if not self.domains_df.empty:
            for _, drow in self.domains_df.iterrows():
                dsn = safe_upper(drow.get("Dataset"))
                if dsn:
                    rec = {c: safe_text(drow.get(c)) for c in DOMAIN_COLUMNS}
                    rec["Class"] = normalize_domain_class(rec.get("Class"))
                    if is_supp_qual_dataset(dsn):
                        parent = supp_parent_domain(dsn)
                        rec["Description"] = f"Supplemental Qualifiers for {parent}" if parent else "Supplemental Qualifiers"
                        rec["Class"] = "RELATIONSHIP"
                        rec["Structure"] = "One record per IDVAR, IDVARVAL, and QNAM value per subject."
                        rec["Purpose"] = rec.get("Purpose") or "Tabulation"
                    self.domain_lookup[dsn] = rec
        self.out_dir = Path(out_dir)

        self.def_ns = f"http://www.cdisc.org/ns/def/v{self.define_version}"
        self.odm_ns = "http://www.cdisc.org/ns/odm/v1.3"
        self.xlink_ns = "http://www.w3.org/1999/xlink"
        self.xsi_ns = "http://www.w3.org/2001/XMLSchema-instance"
        self.define_xsl = "define2-1-0.xsl" if self.define_version == "2.1" else "define2-0-0.xsl"
        self.define_xsd = "define2-1-0.xsd" if self.define_version == "2.1" else "define2-0-0.xsd"
        ET.register_namespace("", self.odm_ns)
        ET.register_namespace("def", self.def_ns)
        ET.register_namespace("xlink", self.xlink_ns)
        ET.register_namespace("xsi", self.xsi_ns)
        

    def build_document_registry(self):
        """Build document lookup from config defaults plus optional Documents sheet.

        Config defaults are always available:
          - AnnotatedCRF / blankcrf for SDTM aCRF
          - ReviewersGuide for csdrg.pdf / adrg.pdf
        Documents sheet rows can add more documents or override Title/Href by ID.
        """
        docs = {}

        def add_doc(doc_id, title, href, leaf_id=None):
            doc_id = safe_text(doc_id)
            href = safe_text(href)
            if not doc_id or not href:
                return
            docs[safe_upper(doc_id)] = {
                "ID": doc_id,
                "Title": safe_text(title) or doc_id,
                "Href": href,
                "LeafID": safe_text(leaf_id) or f"LF.{xml_id(doc_id)}",
            }

        if self.standard == "SDTM" and self.include_acrf:
            add_doc("AnnotatedCRF", "Annotated Case Report Form", self.acrf_file, "LF.blankcrf")
            add_doc("blankcrf", "Annotated Case Report Form", self.acrf_file, "LF.blankcrf")

        if self.include_rg:
            rg_href = self.csdrg_file if self.standard == "SDTM" else self.adrg_file
            rg_title = "SDTM Reviewer's Guide" if self.standard == "SDTM" else "ADaM Reviewer's Guide"
            add_doc("ReviewersGuide", rg_title, rg_href, "LF.ReviewersGuide")

        if isinstance(self.documents_df, pd.DataFrame) and not self.documents_df.empty:
            for _, r in self.documents_df.iterrows():
                doc_id = safe_text(r.get("ID"))
                href = safe_text(r.get("Href"))
                title = safe_text(r.get("Title")) or doc_id
                if doc_id and href:
                    # Preserve legacy/config leaf IDs for standard IDs when overridden in Documents sheet.
                    existing_leaf = docs.get(safe_upper(doc_id), {}).get("LeafID", "")
                    add_doc(doc_id, title, href, existing_leaf or None)

        return docs

    def document_links_for_object(self, object_id):
        """Return Document_Links rows for a dataset/variable ID."""
        if self.document_links_df is None or getattr(self.document_links_df, "empty", True):
            return []
        oid = safe_upper(object_id).replace(" ", "")
        rows = []
        for _, r in self.document_links_df.iterrows():
            if safe_upper(r.get("ID")).replace(" ", "") == oid:
                rows.append(r)
        return rows

    def has_document_links(self, object_id):
        return len(self.document_links_for_object(object_id)) > 0

    def add_document_refs(self, parent, object_id):
        """Attach DocumentRef/PDFPageRef children from Document_Links.

        Document_Links.Document must match Documents.ID or one of the config IDs
        such as ReviewersGuide / AnnotatedCRF.
        """
        added = 0
        for r in self.document_links_for_object(object_id):
            doc_key = safe_upper(r.get("Document"))
            doc = self.document_registry.get(doc_key)
            if not doc:
                continue
            dr = ET.SubElement(parent, self.dq("DocumentRef"), {"leafID": doc["LeafID"]})
            pages = safe_text(r.get("Pages"))
            if pages:
                ET.SubElement(dr, self.dq("PDFPageRef"), {
                    "Type": "PhysicalRef",
                    "PageRefs": pages.replace(",", " ").strip()
                })
            added += 1
        return added

    def q(self, local, ns=None):
        return f"{{{ns or self.odm_ns}}}{local}"

    def dq(self, local):
        return f"{{{self.def_ns}}}{local}"

    def xq(self, local):
        return f"{{{self.xlink_ns}}}{local}"

    def xiq(self, local):
        return f"{{{self.xsi_ns}}}{local}"

    def add_translated(self, parent, text):
        desc = ET.SubElement(parent, self.q("Description"))
        tt = ET.SubElement(desc, self.q("TranslatedText"), {"{http://www.w3.org/XML/1998/namespace}lang": "en"})
        tt.text = safe_text(text)
        return desc

    def build(self):
        # Match the legacy/SAS define.xml header style used with define2-0-0.xsl.
        odm = ET.Element(self.q("ODM"), {
            self.xiq("schemaLocation"): f"{self.def_ns} {self.define_xsd}",
            "ODMVersion": self.odm_version or "1.3",
            "FileOID": self.study_oid or "study_oid",
            "FileType": "Snapshot",
            "CreationDateTime": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        })
        study = ET.SubElement(odm, self.q("Study"), {"OID": self.study_oid})
        gv = ET.SubElement(study, self.q("GlobalVariables"))
        ET.SubElement(gv, self.q("StudyName")).text = self.study_name
        ET.SubElement(gv, self.q("StudyDescription")).text = self.study_description
        ET.SubElement(gv, self.q("ProtocolName")).text = self.protocol

        std_name = "SDTM-IG" if self.standard == "SDTM" else "ADaM-IG"
        mdv = ET.SubElement(study, self.q("MetaDataVersion"), {
            "OID": f"MDV.{self.study_oid}",
            "Name": f"Study {self.study_name} Data Definitions",
            "Description": f"Study {self.study_name} Data Definitions",
            self.dq("DefineVersion"): "2.1.0" if self.define_version == "2.1" else "2.0.0",
            self.dq("StandardName"): std_name,
            self.dq("StandardVersion"): self.ig_version,
        })

        if self.include_acrf and self.standard == "SDTM" and safe_upper("AnnotatedCRF") in self.document_registry:
            acrf = ET.SubElement(mdv, self.dq("AnnotatedCRF"))
            ET.SubElement(acrf, self.dq("DocumentRef"), {"leafID": self.document_registry[safe_upper("AnnotatedCRF")]["LeafID"]})
        # SupplementalDoc contains the reviewer guide from config plus any Documents-sheet files.
        supplemental_docs = []
        seen_leaf_ids = set()
        for key, doc in self.document_registry.items():
            if doc.get("LeafID") == "LF.blankcrf":
                continue
            leaf_id = doc.get("LeafID")
            if leaf_id and leaf_id not in seen_leaf_ids:
                supplemental_docs.append(doc)
                seen_leaf_ids.add(leaf_id)
        if supplemental_docs:
            suppdoc = ET.SubElement(mdv, self.dq("SupplementalDoc"))
            for doc in supplemental_docs:
                ET.SubElement(suppdoc, self.dq("DocumentRef"), {"leafID": doc["LeafID"]})

        # Build sections, then normalize MetaDataVersion child order.
        # This final reorder prevents Pinnacle DD0007 errors such as MethodDef
        # appearing after an element that closes the ODM metadata sequence.
        self.add_itemgroups(mdv)
        self.add_itemdefs(mdv)
        self.add_codelists(mdv)
        self.add_methods_and_comments(mdv)
        self.add_leaves(mdv)
        self.add_valuelists(mdv)
        self.add_whereclauses(mdv)
        self.reorder_metadata_version_children(mdv)

        return odm


    def reorder_metadata_version_children(self, mdv):
        """Reorder MetaDataVersion children to keep Define-XML schema order valid.

        Define 2.x expects def:ValueListDef/def:WhereClauseDef before dataset/item
        definitions, MethodDef/CommentDef after CodeList, and def:leaf near the end.
        The generator builds some sections in workflow order, so this final pass
        normalizes the XML order before writing the file.
        """
        def local_name(elem):
            tag = elem.tag
            if "}" in tag:
                return tag.rsplit("}", 1)[1]
            return tag

        order = {
            "AnnotatedCRF": 10,
            "SupplementalDoc": 20,
            "ValueListDef": 30,
            "WhereClauseDef": 40,
            "ItemGroupDef": 50,
            "ItemDef": 60,
            "CodeList": 70,
            "MethodDef": 80,
            "CommentDef": 90,
            "leaf": 100,
        }
        children = list(mdv)
        if not children:
            return
        indexed = list(enumerate(children))
        indexed.sort(key=lambda x: (order.get(local_name(x[1]), 999), x[0]))
        mdv[:] = [elem for _, elem in indexed]




    def active_vlm_df(self):
        """Return VLM rows that are actually referenced by generated ItemDefs.

        P21 DD0081 is raised when ValueListDef exists but the parent ItemDef does
        not have ValueListRef.  This can happen when SUPP/VLM rows are generated
        for domains that are not present in the final KEEP=1 metadata.  Filter
        VLM to dataset/result-variable pairs present in metadata before writing
        ValueListDef, WhereClauseDef, VLM ItemDef, and VLM comments.
        """
        if self.vlm_df is None or getattr(self.vlm_df, "empty", True):
            return pd.DataFrame(columns=getattr(self.vlm_df, "columns", []))
        if self.metadata_df is None or getattr(self.metadata_df, "empty", True):
            return self.vlm_df.iloc[0:0].copy()
        meta_keys = set()
        for _, r in self.metadata_df.iterrows():
            ds = safe_upper(r.get("Dataset"))
            var = safe_upper(r.get("Variable"))
            if ds and var:
                meta_keys.add((ds, var))
        mask = self.vlm_df.apply(
            lambda r: (safe_upper(r.get("Dataset")), safe_upper(r.get("Result Variable"))) in meta_keys,
            axis=1
        )
        return self.vlm_df[mask].copy()

    def parent_variable_metadata(self, dataset, variable):
        """Return parent dataset variable metadata used to cap VLM length."""
        if self.metadata_df is None or getattr(self.metadata_df, "empty", True):
            return {}
        ds = safe_upper(dataset)
        var = safe_upper(variable)
        df = self.metadata_df.copy()
        if not {"Dataset", "Variable"}.issubset(df.columns):
            return {}
        m = df[
            (df["Dataset"].astype(str).str.upper() == ds)
            & (df["Variable"].astype(str).str.upper() == var)
        ]
        if m.empty:
            return {}
        r = m.iloc[0]
        return {
            "Type": safe_text(r.get("Type")),
            "Length": safe_text(r.get("Length")) or safe_text(r.get("Len")),
            "Format": safe_text(r.get("Format")) or safe_text(r.get("Control or Format")),
        }

    def capped_vlm_length(self, dataset, variable, value_level_length):
        """Cap VLM ItemDef Length so it never exceeds the parent variable length.

        P21 DD0123 is triggered when Value Level Length is greater than the
        corresponding dataset variable length. This is especially common when
        value-level ORRES values look numeric and are inferred as length 8 while
        the XPT variable length is smaller, for example DAORRES length 3.
        """
        vlm_len = to_int_or_none(value_level_length)
        parent = self.parent_variable_metadata(dataset, variable)
        parent_len = to_int_or_none(parent.get("Length"))

        if parent_len is None:
            return safe_text(value_level_length)
        if vlm_len is None:
            return str(parent_len)
        return str(min(vlm_len, parent_len))

    def referenced_codelist_oids(self):
        """Return CodeList OIDs referenced by ItemDef/ValueList ItemDef.

        This prevents DD0082 by writing only terminology that is actually linked
        from ItemDefs.  Dataset-qualified formats such as LB.UNIT remain
        CL.LB.UNIT; CT lookup still normalizes internally to UNIT.
        """
        refs = set()
        ct_map = getattr(self, "ct_alias_map", {}) if hasattr(self, "ct_alias_map") else {}

        def add_ref(fmt, variable="", dtype=""):
            fmt = safe_text(fmt)
            if not fmt:
                return
            if is_define_codelist_format(fmt, variable, dtype):
                refs.add(codelist_oid_for_format(fmt, ct_map))

        if self.metadata_df is not None and not getattr(self.metadata_df, "empty", True):
            for _, row in self.metadata_df.iterrows():
                var = safe_upper(row.get("Variable"))
                dtype = safe_text(row.get("Type")) or "char"
                fmt = default_format_for_variable(var, safe_text(row.get("Format")))
                add_ref(fmt, var, dtype)

        active_vlm = self.active_vlm_df()
        if active_vlm is not None and not getattr(active_vlm, "empty", True):
            for _, row in active_vlm.iterrows():
                add_ref(row.get("Format"), row.get("Result Variable"), row.get("Type"))

        return refs

    def add_whereclauses(self, mdv):
        """Write WhereClauseDef elements in schema-valid position.

        This method is intentionally generic and works from self.vlm_df.
        It should be called after ValueListDef and before ItemGroupDef.
        """
        active_vlm = self.active_vlm_df()
        if active_vlm.empty:
            return

        for _, row in active_vlm.iterrows():
            ds = safe_upper(row.get("Dataset"))
            res = safe_upper(row.get("Result Variable"))
            key = self.vlm_key(row)
            if not ds or not res or not key:
                continue

            wc = ET.SubElement(mdv, self.dq("WhereClauseDef"), {
                "OID": f"WC.{ds}.{res}.{key}"
            })

            try:
                pairs = self.vlm_group_pairs(row)
            except Exception:
                pairs = []

            for gvar, gval in pairs:
                gvar = safe_upper(gvar)
                gval = safe_text(gval)
                if not gvar or not gval:
                    continue
                rc = ET.SubElement(wc, self.q("RangeCheck"), {
                    "Comparator": "EQ",
                    "SoftHard": "Soft",
                    self.dq("ItemOID"): f"IT.{ds}.{gvar}"
                })
                ET.SubElement(rc, self.q("CheckValue")).text = gval

    def add_itemgroups(self, mdv):
        meta = self.metadata_df.copy()
        meta["Order"] = pd.to_numeric(meta.get("Order", ""), errors="coerce")
        dataset_order = sorted(
            [safe_upper(x) for x in meta.get("Dataset", pd.Series(dtype=str)).dropna().unique()],
            key=lambda x: (dataset_class_sort_value(x, self.domain_lookup, self.standard), x)
        )
        for ds in dataset_order:
            ddf = meta[meta["Dataset"].apply(safe_upper) == ds].copy()
            ds = safe_upper(ds)
            dinfo = dict(self.domain_lookup.get(ds, {}))
            if is_supp_qual_dataset(ds):
                parent = supp_parent_domain(ds)
                dinfo["Description"] = f"Supplemental Qualifiers for {parent}" if parent else "Supplemental Qualifiers"
                dinfo["Class"] = "RELATIONSHIP"
                dinfo["Structure"] = "One record per IDVAR, IDVARVAL, and QNAM value per subject."
                dinfo["Purpose"] = dinfo.get("Purpose") or "Tabulation"
            desc = safe_text(dinfo.get("Description")) or ds
            dclass = normalize_domain_class(dinfo.get("Class"))
            structure = safe_text(dinfo.get("Structure")) or "One record per subject per event/assessment as applicable"
            purpose = safe_text(dinfo.get("Purpose")) or ("Tabulation" if self.standard == "SDTM" else "Analysis")
            location = safe_text(dinfo.get("Location")) or f"{ds.lower()}.xpt"
            documentation = safe_text(dinfo.get("Documentation"))

            # Dataset location hyperlink for XSL display.
            # Some Define 2.0 stylesheets read the child def:leaf directly,
            # while others expect def:ArchiveLocationID on ItemGroupDef.
            # Write both using the same leaf ID so the Location column renders as a clickable xpt link.
            leaf_id = f"LF.{ds}"
            if not location:
                location = f"{ds.lower()}.xpt"
            else:
                location = safe_text(location)
                # If Domains sheet has only dataset name (e.g. AE), write ae.xpt.
                # If it already has AE.xpt or a folder path, keep it as provided.
                if Path(location).suffix == "":
                    location = f"{location.lower()}.xpt"

            repeating, is_reference_data = get_dataset_repeating_reference(ds, dclass, self.standard)

            attrs = {
                "OID": f"IG.{ds}",
                "Name": ds,
                "SASDatasetName": ds,
                "Repeating": repeating,
                "IsReferenceData": is_reference_data,
                "Purpose": purpose,
                self.dq("Structure"): structure,
                self.dq("ArchiveLocationID"): leaf_id,
            }
            # Define-XML/P21 DD0046: Domain is valid for SDTM/SEND-style domains,
            # but ADaM datasets such as ADAE/ADLB must not carry Domain=.
            if self.standard == "SDTM":
                attrs["Domain"] = ds
            if dclass:
                attrs[self.dq("Class")] = dclass
            if (documentation and not documentation.lower().endswith((".pdf", ".html", ".htm", ".docx", ".rtf", ".txt"))) or self.has_document_links(ds):
                attrs[self.dq("CommentOID")] = f"COM.{ds}"
            ig = ET.SubElement(mdv, self.q("ItemGroupDef"), attrs)
            self.add_translated(ig, desc)
            ddf = ddf.sort_values("Order", na_position="last")
            for idx, row in ddf.iterrows():
                var = safe_upper(row.get("Variable"))
                attrs = {
                    "ItemOID": f"IT.{ds}.{var}",
                    "OrderNumber": str(int(row["Order"])) if not pd.isna(row.get("Order")) else str(idx + 1),
                    "Mandatory": "Yes" if safe_upper(row.get("Core")) in {"REQ", "REQUIRED"} else "No",
                }
                keyseq = to_int_or_none(row.get("ID Var"))
                if keyseq is not None:
                    attrs["KeySequence"] = str(keyseq)
                fmt = default_format_for_variable(var, safe_text(row.get("Format")))
                # CodeListRef must be written under ItemDef (not ItemRef) for define2-0-0.xsl.
                # Kept off ItemRef so codelists render correctly in the browser view.
                origin = safe_upper(row.get("Origin"))
                comments = clean_comment_text(row.get("Comments"), origin)
                if origin == "DERIVED" and comments:
                    attrs["MethodOID"] = f"MT.{ds}.{var}"
                ET.SubElement(ig, self.q("ItemRef"), attrs)
            leaf = ET.SubElement(ig, self.dq("leaf"), {"ID": leaf_id, self.xq("href"): location})
            ET.SubElement(leaf, self.dq("title")).text = Path(str(location).replace("\\", "/")).name

    def add_itemdefs(self, mdv):
        for _, row in self.metadata_df.iterrows():
            ds = safe_upper(row.get("Dataset"))
            var = safe_upper(row.get("Variable"))
            dtype = safe_text(row.get("Type")) or "char"
            length = safe_text(row.get("Length"))
            label = safe_text(row.get("Label")) or var
            fmt = default_format_for_variable(var, safe_text(row.get("Format")))

            # P21 DD0149: SEND/SDTM ISO 8601 --DTC/--DUR variables must not be text.
            # DTC length > 10 is treated as datetime; length <= 10 is date.
            # DUR variables use durationDatetime.
            if var.endswith("DTC") or var.endswith("DUR"):
                fmt = "ISO 8601"

            define_dtype = self.item_define_datatype(var, dtype, length, fmt)

            attrs = {
                "OID": f"IT.{ds}.{var}",
                "Name": var,
                "SASFieldName": var,
                "DataType": define_dtype,
            }
            # Define-XML/P21 DD0068/DD0149: ISO date/time/duration ItemDefs
            # should not carry a normal character Length.
            no_length_datatypes = {
                "date", "datetime", "time", "partialDate", "partialTime",
                "partialDatetime", "incompleteDatetime", "durationDatetime",
                "intervalDatetime"
            }
            if define_dtype not in no_length_datatypes:
                attrs["Length"] = str(length) if length else "200" if dtype == "char" else "8"
            # Define-XML/P21 OD0071: float ItemDef requires SignificantDigits.
            if define_dtype == "float":
                attrs["SignificantDigits"] = significant_digits_from_format(fmt, default="8")
            code_list_oid = ""
            if fmt:
                # Define-XML 2.0: do not write nonstandard FormatName on ItemDef.
                # ISO/date formats use def:DisplayFormat; CT/dictionaries use CodeListRef.
                if is_iso_8601_format(fmt):
                    attrs[self.dq("DisplayFormat")] = "ISO 8601"
                if is_define_codelist_format(fmt, var, dtype):
                    code_list_oid = codelist_oid_for_format(fmt, getattr(self, "ct_alias_map", {}))
            origin = safe_upper(row.get("Origin"))
            comments = clean_comment_text(row.get("Comments"), origin)
            # Define-XML XSL displays Predecessor value directly from def:Origin.
            # Do not also create/comment-reference the same text, otherwise the XSL shows it twice.
            # Derived comments continue to become MethodDef via ItemRef/@MethodOID.
            var_object_id = f"{ds}.{var}"
            if (comments and origin not in {"DERIVED", "PREDECESSOR"}) or (origin != "DERIVED" and self.has_document_links(var_object_id)):
                attrs[self.dq("CommentOID")] = f"COM.{ds}.{var}"
            item = ET.SubElement(mdv, self.q("ItemDef"), attrs)
            self.add_translated(item, label)
            if code_list_oid:
                ET.SubElement(item, self.q("CodeListRef"), {"CodeListOID": code_list_oid})
            # If this variable has value-level metadata, add ValueListRef like SAS define output.
            active_vlm = self.active_vlm_df()
            if not active_vlm.empty:
                vlm_match = active_vlm[
                    (active_vlm["Dataset"].astype(str).str.upper() == ds)
                    & (active_vlm["Result Variable"].astype(str).str.upper() == var)
                ]
                if not vlm_match.empty:
                    ET.SubElement(item, self.dq("ValueListRef"), {"ValueListOID": f"VL.{ds}.{var}"})
            self.add_origin(item, row)

    def define_datatype(self, dtype):
        t = safe_upper(dtype)
        if t == "CHAR":
            return "text"
        if t == "DATE":
            return "date"
        if t == "DATETIME":
            return "datetime"
        if t == "FLOAT":
            return "float"
        return "integer" if t == "NUM" else "text"

    def codelist_datatype_for_format(self, fmt, fallback=""):
        """Infer the Define CodeList datatype for a referenced format from its Code values.

        This is used before writing ItemDef so numeric variables that reference a
        codelist with decimal codes, for example VISITNUM values 3.01/6.01, are
        written as DataType="float" instead of integer. P21 OD0080 requires the
        variable ItemDef DataType and CodeList DataType to align.
        """
        fmt_txt = safe_text(fmt)
        if not fmt_txt or getattr(self, "formats_df", pd.DataFrame()).empty:
            return safe_text(fallback)
        try:
            ct_map = getattr(self, "ct_alias_map", {}) if hasattr(self, "ct_alias_map") else {}
            target_name = define_codelist_name(fmt_txt, ct_map)
            work = self.formats_df.copy()
            if "Format" not in work.columns or "Code" not in work.columns:
                return safe_text(fallback)
            work["__DefineFormat"] = work["Format"].apply(lambda x: define_codelist_name(x, ct_map))
            fdf = work[work["__DefineFormat"].astype(str) == target_name].copy()
            if fdf.empty:
                # fallback to exact source format match, useful before CT alias map is populated
                fdf = work[work["Format"].astype(str).str.upper() == safe_upper(fmt_txt)].copy()
            if fdf.empty:
                return safe_text(fallback)
            return infer_codelist_datatype_from_codes(fdf["Code"].tolist(), fallback=safe_text(fallback) or "text")
        except Exception:
            return safe_text(fallback)

    def item_define_datatype(self, variable, dtype, length="", fmt=""):
        """Return Define-XML DataType for an ItemDef.

        Fixes P21 DD0149 for ISO 8601 --DTC/--DUR variables:
        - --DUR -> durationDatetime
        - --DTC length > 10 -> datetime
        - --DTC length <= 10 -> date

        Also keeps numeric variables as integer/float based on the existing type.
        """
        var = safe_upper(variable)
        if var.endswith("DUR"):
            return "durationDatetime"
        if var.endswith("DTC"):
            ln = to_int_or_none(length)
            return "datetime" if (ln is not None and ln > 10) else "date"

        base_dtype = self.define_datatype(dtype)

        # When a numeric variable references a codelist, align the ItemDef datatype
        # to the actual codelist Code values. Example: VISITNUM may be numeric in
        # XPT, but the codelist can include 3.01/6.01, so both ItemDef and
        # CodeList must be float, not integer.
        if fmt and base_dtype in {"integer", "float"} and is_define_codelist_format(fmt, variable, dtype):
            cl_dtype = self.codelist_datatype_for_format(fmt, fallback=base_dtype)
            if cl_dtype in {"integer", "float"}:
                return cl_dtype

        return base_dtype

    def referenced_codelist_datatypes(self):
        """Map CodeList OID -> DataType based on the variables that reference it.

        P21 OD0080 requires CodeList/@DataType to match the referencing
        ItemDef/@DataType. This is important for numeric codelists such as
        VISITNUM, where the codelist must be integer/float rather than text.
        """
        out = {}
        ct_map = getattr(self, "ct_alias_map", {}) if hasattr(self, "ct_alias_map") else {}

        def add(fmt, variable="", dtype="", length=""):
            fmt = safe_text(fmt)
            if not fmt or not is_define_codelist_format(fmt, variable, dtype):
                return
            oid = codelist_oid_for_format(fmt, ct_map)
            dt = self.item_define_datatype(variable, dtype, length, fmt)
            if dt in {"date", "datetime", "time", "partialDate", "partialTime",
                      "partialDatetime", "incompleteDatetime", "durationDatetime",
                      "intervalDatetime"}:
                dt = "text"
            if oid not in out:
                out[oid] = dt
            elif out[oid] != dt:
                # Prefer float if mixed integer/float; otherwise keep text only when no numeric use exists.
                if "float" in {out[oid], dt}:
                    out[oid] = "float"
                elif "integer" in {out[oid], dt} and "text" not in {out[oid], dt}:
                    out[oid] = "integer"

        if self.metadata_df is not None and not getattr(self.metadata_df, "empty", True):
            for _, row in self.metadata_df.iterrows():
                var = safe_upper(row.get("Variable"))
                dtype = safe_text(row.get("Type")) or "char"
                length = safe_text(row.get("Length")) or safe_text(row.get("Len"))
                fmt = default_format_for_variable(var, safe_text(row.get("Format")))
                add(fmt, var, dtype, length)

        active_vlm = self.active_vlm_df()
        if active_vlm is not None and not getattr(active_vlm, "empty", True):
            for _, row in active_vlm.iterrows():
                add(row.get("Format"), row.get("Result Variable"), row.get("Type"), row.get("Length"))

        return out

    def add_origin(self, item, row):
        origin = safe_text(row.get("Origin"))
        upper = origin.upper()
        if not origin:
            return

        # Important: do not add TranslatedText that repeats the origin heading.
        # The Define 2.0 XSL already prints the Origin Type.  If we also write
        # <TranslatedText>Protocol</TranslatedText> or <TranslatedText>CRF</TranslatedText>,
        # the browser view shows duplicate lines such as Protocol / Protocol.
        if "CRF" in upper:
            org = ET.SubElement(item, self.dq("Origin"), {"Type": "CRF"})
            pages = parse_crf_pages(origin)
            if pages and self.standard == "SDTM":
                dr = ET.SubElement(org, self.dq("DocumentRef"), {"leafID": "LF.blankcrf"})
                ET.SubElement(dr, self.dq("PDFPageRef"), {
                    "Type": "PhysicalRef",
                    "PageRefs": " ".join(str(p) for p in pages)
                })
            return

        known = {
            "DERIVED": "Derived",
            "ASSIGNED": "Assigned",
            "PREDECESSOR": "Predecessor",
            "PROTOCOL": "Protocol",
            "EDT": "eDT",
            "E-DT": "eDT",
        }
        if upper in known:
            org = ET.SubElement(item, self.dq("Origin"), {"Type": known[upper]})
            if upper == "PREDECESSOR":
                pred_value = infer_predecessor_value(row, getattr(self, "standard", ""))
                if pred_value:
                    self.add_translated(org, pred_value)
            return

        # For any custom origin, write only Type.  The detailed source/method
        # belongs in CommentDef/MethodDef, not as a duplicate origin line.
        ET.SubElement(item, self.dq("Origin"), {"Type": origin})

    def add_valuelists(self, mdv):
        active_vlm = self.active_vlm_df()
        if active_vlm.empty:
            return
        for (ds, res), rdf in active_vlm.groupby(["Dataset", "Result Variable"], sort=False):
            ds = safe_upper(ds); res = safe_upper(res)
            vl = ET.SubElement(mdv, self.dq("ValueListDef"), {"OID": f"VL.{ds}.{res}"})
            for i, (_, row) in enumerate(rdf.iterrows(), start=1):
                key = self.vlm_key(row)
                attrs = {
                    "ItemOID": f"IT.{ds}.{res}.{key}",
                    "OrderNumber": str(i),
                    "Mandatory": "Yes" if ds.startswith("SUPP") and res in {"QNAM", "QVAL"} else "No",
                }
                row_fmt = safe_text(row.get("Format"))
                # CodeListRef is written under the VLM ItemDef below, not under ItemRef.
                if safe_upper(row.get("Origin")) == "DERIVED" and clean_comment_text(row.get("Comment"), row.get("Origin")):
                    attrs["MethodOID"] = f"MT.{ds}.{res}.{key}"
                ir = ET.SubElement(vl, self.q("ItemRef"), attrs)
                ET.SubElement(ir, self.dq("WhereClauseRef"), {"WhereClauseOID": f"WC.{ds}.{res}.{key}"})

        for _, row in active_vlm.iterrows():
            ds = safe_upper(row.get("Dataset")); res = safe_upper(row.get("Result Variable"))
            key = self.vlm_key(row)
            define_dtype = self.item_define_datatype(res, row.get("Type"), row.get("Length"), row.get("Format"))
            attrs = {
                "OID": f"IT.{ds}.{res}.{key}",
                "Name": res,
                "SASFieldName": res,
                "DataType": define_dtype,
            }
            # Define-XML/P21 DD0068/DD0149: date/time/duration VLM ItemDefs must not carry Length.
            no_length_datatypes = {
                "date", "datetime", "time", "partialDate", "partialTime",
                "partialDatetime", "incompleteDatetime", "durationDatetime",
                "intervalDatetime"
            }
            if define_dtype not in no_length_datatypes:
                attrs["Length"] = self.capped_vlm_length(ds, res, row.get("Length")) or "200"
            # Define-XML/P21 OD0071: float VLM ItemDef requires SignificantDigits.
            if define_dtype == "float":
                attrs["SignificantDigits"] = significant_digits_from_format(row.get("Format"), default="8")
            vlm_origin = safe_upper(row.get("Origin"))
            vlm_comment = clean_comment_text(row.get("Comment"), vlm_origin)
            if vlm_comment and vlm_origin != "DERIVED":
                attrs[self.dq("CommentOID")] = f"COM.{ds}.{res}.{key}"
            fmt = safe_text(row.get("Format"))
            code_list_oid = ""
            if fmt:
                # Define-XML 2.0: no FormatName on VLM ItemDef either.
                if is_iso_8601_format(fmt):
                    attrs[self.dq("DisplayFormat")] = "ISO 8601"
                code_list_oid = codelist_oid_for_format(fmt, getattr(self, "ct_alias_map", {})) if is_define_codelist_format(fmt, res, row.get("Type")) else ""
            item = ET.SubElement(mdv, self.q("ItemDef"), attrs)
            self.add_translated(item, f"{res} where {safe_text(row.get('Where Clause'))}")
            if code_list_oid:
                ET.SubElement(item, self.q("CodeListRef"), {"CodeListOID": code_list_oid})
            # VLM origin/comment should behave like variable-level metadata.
            try:
                self.add_origin(item, {"Origin": row.get("Origin"), "Comments": row.get("Comment")})
            except Exception:
                pass

    def vlm_key(self, row):
        # Legacy/SAS-style key: grouping variable/value pairs only.
        # Example: DATESTCD.DISP or LBTESTCD.ALB.LBCAT.CHEMISTRY.LBSPEC.SERUM
        parts = []
        gv = safe_text(row.get("Grouping Variable")); val = safe_text(row.get("Group Value"))
        if gv and val:
            parts.extend([safe_upper(gv), safe_text(val)])
        for i in range(1, 5):
            gv = safe_text(row.get(f"Grouping Variable {i}")); val = safe_text(row.get(f"Group Value {i}"))
            if gv and val:
                parts.extend([safe_upper(gv), safe_text(val)])
        key = ".".join(parts)
        return key if key else xml_id(safe_text(row.get("Where Clause")) or safe_text(row.get("Result Variable")))

    def vlm_group_pairs(self, row):
        pairs = []
        gv = safe_text(row.get("Grouping Variable")); val = safe_text(row.get("Group Value"))
        if gv and val:
            pairs.append((gv, val))
        for i in range(1, 5):
            gv = safe_text(row.get(f"Grouping Variable {i}")); val = safe_text(row.get(f"Group Value {i}"))
            if gv and val:
                pairs.append((gv, val))
        return pairs

    def add_external_dictionary_codelists(self, mdv):
        """Add external dictionary CodeLists once.

        Important: the Define 2.0 XSL shows the External Dictionaries table only when
        a CodeList contains unprefixed ExternalCodeList in the ODM namespace.  Therefore MedDRA/WHO Drug must be
        written as CodeList + ExternalCodeList, not as regular CodeListItem rows.

        Detection is deliberately broad and case-insensitive. It scans Format,
        Control or Format, Data Format and all row values because some specs keep
        WHODRUG/MedDRA in different columns or with mixed casing.
        """
        used = []

        def remember_from_value(value):
            key = external_dictionary_key(value)
            if key and key not in used:
                used.append(key)

        # Scan key columns first.
        preferred_cols = [
            "Format", "Control or Format", "Control", "Data Format",
            "Codelist", "CodeList", "Dictionary", "External Dictionary"
        ]
        for df in [self.metadata_df, self.vlm_df, self.formats_df]:
            if df is None or getattr(df, "empty", True):
                continue
            for col in preferred_cols:
                if col in df.columns:
                    for val in df[col].tolist():
                        remember_from_value(val)
            # Fallback: scan every text cell in case the dictionary marker is stored elsewhere.
            for _, r in df.iterrows():
                for val in r.tolist():
                    remember_from_value(val)

        # If versions are filled in the GUI/config, include the dictionary rows as well.
        # This prevents the table from being missed when the format text was normalized earlier.
        if safe_text(self.meddra_version) and "MedDRA" not in used:
            used.append("MedDRA")
        if safe_text(self.whodrug_version) and "WHODRUG" not in used:
            used.append("WHODRUG")

        already_written = set()
        for key in used:
            if key == "MedDRA":
                oid = "CL.MEDDRA"
                name = "MedDRA Dictionary"
                dictionary = "MedDRA"
                version = safe_text(self.meddra_version)
            elif key == "WHODRUG":
                oid = "CL.WHODRUG"
                name = "WHO Drug Dictionary"
                dictionary = "WHO Drug"
                version = safe_text(self.whodrug_version)
            else:
                continue

            if oid in already_written:
                continue
            already_written.add(oid)

            cl = ET.SubElement(mdv, self.q("CodeList"), {
                "OID": oid,
                "Name": name,
                "DataType": "text",
            })
            ext_attrs = {"Dictionary": dictionary}
            if version:
                ext_attrs["Version"] = version
            ET.SubElement(cl, self.q("ExternalCodeList"), ext_attrs)

    def add_codelists(self, mdv):
        # First add MedDRA/WHO Drug as external dictionaries once, based on metadata use.
        self.add_external_dictionary_codelists(mdv)

        if self.formats_df.empty:
            return

        try:
            write_formats = apply_display_decode_suppression(self.formats_df)
        except Exception:
            write_formats = self.formats_df.copy()

        ct_map = getattr(self, "ct_alias_map", {}) if hasattr(self, "ct_alias_map") else {}
        referenced_oids = self.referenced_codelist_oids()
        referenced_datatypes = self.referenced_codelist_datatypes()

        write_formats = write_formats.copy()
        # Group by the actual Define codelist name, not by the CT lookup key.
        # LB.UNIT must remain a distinct Define codelist OID (CL.LB.UNIT), while
        # CT/NCI enrichment still looks up UNIT through ct_lookup_key_from_format().
        write_formats["__DefineFormat"] = write_formats["Format"].apply(lambda x: define_codelist_name(x, ct_map))

        for fmt, fdf in write_formats.groupby("__DefineFormat", sort=False):
            raw_fmt = safe_text(fdf["Format"].iloc[0]) if "Format" in fdf.columns and not fdf.empty else fmt
            fmt = safe_text(fmt)
            if not fmt or is_external_dictionary_format(fmt):
                continue

            cl_oid = codelist_oid_for_format(raw_fmt, ct_map)
            if referenced_oids and cl_oid not in referenced_oids:
                continue

            fdf = fdf.copy()
            if "Code" in fdf.columns:
                fdf = fdf[fdf["Code"].apply(safe_text) != ""].copy()
            if fdf.empty:
                continue

            # Use UNIT/FREQ/ROUTE etc. only for CT lookup; keep LB.UNIT/APCM.FREQ as the Define OID.
            ct_key = ct_lookup_key_from_format(raw_fmt) or ct_lookup_key_from_format(fmt) or raw_fmt
            ct_info = get_ct_info_generic(ct_map, ct_key) or get_ct_info_generic(ct_map, raw_fmt)

            codelist_nci = codelist_alias_code(fdf, raw_fmt) or safe_text(ct_info.get("codelist_code")) or known_codelist_nci_code(raw_fmt)

            prepared_rows = []
            for _, row_obj in fdf.iterrows():
                code_value = safe_text(row_obj.get("Code"))
                if not code_value:
                    continue

                ct_syn = get_ct_synonym_generic(ct_info, code_value)
                existing_decode = safe_text(row_obj.get("Decode"))
                dec0 = ct_syn or existing_decode
                dec_value = suppress_decode_if_same_for_display_codelist(
                    fmt,
                    code_value,
                    dec0,
                    row_obj.get("Source Variable", "")
                )

                term_nci = (
                    safe_text(row_obj.get("Term Code"))
                    or get_ct_term_code_generic(ct_info, code_value)
                    or known_term_nci_code(fmt, code_value)
                    or known_term_nci_code(raw_fmt, code_value)
                )

                # Preserve source value exactly. Do not fix casing during Define generation.
                # CT term/NCI matching is case-sensitive.
                prepared_rows.append({
                    "code_value": code_value,
                    "decode_value": safe_text(dec_value),
                    "term_nci": safe_text(term_nci),
                })

            if not prepared_rows:
                continue

            # Decide CodeList/@DataType from the actual codelist Code values.
            # This prevents numeric codelists such as VISITNUM from being written
            # as text, and also keeps mixed/non-numeric codelists as text.
            code_list_datatype = infer_codelist_datatype_from_codes(
                [item_data.get("code_value") for item_data in prepared_rows],
                fallback=referenced_datatypes.get(cl_oid, "text")
            )

            cl = ET.SubElement(mdv, self.q("CodeList"), {
                "OID": cl_oid,
                "Name": define_codelist_name(raw_fmt, ct_map),
                "DataType": code_list_datatype,
            })

            # OD0079: keep only one term per CodedValue within a CodeList.
            # When UNIT/FREQ/ROUTE are collected from multiple domains, the same
            # term can appear more than once. Merge them before writing XML.
            deduped_rows = []
            seen_codes = set()
            for item_data in prepared_rows:
                code_key = safe_text(item_data.get("code_value"))
                if code_key in seen_codes:
                    continue
                seen_codes.add(code_key)
                deduped_rows.append(item_data)
            prepared_rows = deduped_rows

            all_decode_blank = all(item_data["decode_value"] == "" for item_data in prepared_rows)

            for order_number, item_data in enumerate(prepared_rows, start=1):
                code_value = item_data["code_value"]
                decode_value = item_data["decode_value"]
                term_nci = item_data["term_nci"]

                attrs = {
                    "CodedValue": code_value,
                    "OrderNumber": str(order_number)
                }
                if not term_nci:
                    attrs[self.dq("ExtendedValue")] = "Yes"

                if all_decode_blank:
                    item = ET.SubElement(cl, self.q("EnumeratedItem"), attrs)
                    add_nci_alias(item, term_nci, self.q)
                else:
                    item = ET.SubElement(cl, self.q("CodeListItem"), attrs)
                    final_decode = safe_text(decode_value) or code_value
                    decode = ET.SubElement(item, self.q("Decode"))
                    tt = ET.SubElement(
                        decode,
                        self.q("TranslatedText"),
                        {"{http://www.w3.org/XML/1998/namespace}lang": "en"}
                    )
                    tt.text = final_decode
                    add_nci_alias(item, term_nci, self.q)

            add_nci_alias(cl, codelist_nci, self.q)


    def add_methods_and_comments(self, mdv):

        # Terminology version comment is written late to keep Define-XML element order schema-valid.
        term_notes = []
        if self.ct_version:
            term_notes.append(f"CDISC CT Version: {self.ct_version}")
        if self.meddra_version:
            term_notes.append(f"MedDRA Version: {self.meddra_version}")
        if self.whodrug_version:
            term_notes.append(f"WHO Drug Version: {self.whodrug_version}")
        # Do not write COM.TerminologyVersions unless it is referenced.
        # Unreferenced CommentDef causes P21 DD0079.
        if False and term_notes:
            c = ET.SubElement(mdv, self.dq("CommentDef"), {"OID": "COM.TerminologyVersions"})
            self.add_translated(c, "; ".join(term_notes))

        # Variable-level methods/comments. Document_Links are inferred by ID:
        #   DS.VAR + Origin=Derived -> MethodDef
        #   DS.VAR + other Origin    -> CommentDef
        for _, row in self.metadata_df.iterrows():
            ds = safe_upper(row.get("Dataset")); var = safe_upper(row.get("Variable"))
            origin = safe_upper(row.get("Origin")); comment = clean_comment_text(row.get("Comments"), origin)
            object_id = f"{ds}.{var}"
            has_docs = self.has_document_links(object_id)
            if origin == "DERIVED" and (comment or has_docs):
                m = ET.SubElement(mdv, self.q("MethodDef"), {
                    "OID": f"MT.{ds}.{var}",
                    "Name": f"Algorithm to derive {ds}.{var}",
                    "Type": "Computation",
                })
                self.add_translated(m, comment or "See referenced document.")
                self.add_document_refs(m, object_id)
            elif origin != "DERIVED" and ((comment and origin != "PREDECESSOR") or has_docs):
                c = ET.SubElement(mdv, self.dq("CommentDef"), {"OID": f"COM.{ds}.{var}"})
                # For Predecessor, do not repeat the predecessor text as comment because the XSL
                # already displays it from def:Origin. Use a neutral line only when document links exist.
                ctext = comment if origin != "PREDECESSOR" else "See referenced document."
                self.add_translated(c, ctext or "See referenced document.")
                self.add_document_refs(c, object_id)

        # Dataset-level documentation comments from Domains sheet plus Document_Links rows with ID=Dataset.
        if hasattr(self, "domain_lookup"):
            for ds, dinfo in self.domain_lookup.items():
                documentation = safe_text(dinfo.get("Documentation"))
                has_docs = self.has_document_links(ds)
                if (documentation and not documentation.lower().endswith((".pdf", ".html", ".htm", ".docx", ".rtf", ".txt"))) or has_docs:
                    c = ET.SubElement(mdv, self.dq("CommentDef"), {"OID": f"COM.{ds}"})
                    self.add_translated(c, documentation or "See referenced document.")
                    self.add_document_refs(c, ds)

        # VLM methods/comments
        active_vlm = self.active_vlm_df()
        if not active_vlm.empty:
            for _, row in active_vlm.iterrows():
                origin = safe_upper(row.get("Origin"))
                comment = clean_comment_text(row.get("Comment"), origin)
                if comment:
                    ds = safe_upper(row.get("Dataset")); res = safe_upper(row.get("Result Variable")); key = self.vlm_key(row)
                    if origin == "DERIVED":
                        m = ET.SubElement(mdv, self.q("MethodDef"), {
                            "OID": f"MT.{ds}.{res}.{key}",
                            "Name": f"Algorithm to derive {ds}.{res}.{key}",
                            "Type": "Computation",
                        })
                        self.add_translated(m, comment)
                    else:
                        c = ET.SubElement(mdv, self.dq("CommentDef"), {"OID": f"COM.{ds}.{res}.{key}"})
                        self.add_translated(c, comment)

    def add_leaves(self, mdv):
        seen = set()
        for _, doc in self.document_registry.items():
            leaf_id = safe_text(doc.get("LeafID"))
            href = safe_text(doc.get("Href"))
            title = safe_text(doc.get("Title")) or Path(href).name
            if not leaf_id or not href or leaf_id in seen:
                continue
            seen.add(leaf_id)
            leaf = ET.SubElement(mdv, self.dq("leaf"), {"ID": leaf_id, self.xq("href"): href})
            ET.SubElement(leaf, self.dq("title")).text = title

    def remove_blank_decode_nodes(self, root):
        """Remove CodeList Decode nodes where TranslatedText is blank.

        This prevents XSL display like:
            "Visit 2" = ""
        when Decode was intentionally suppressed.
        """
        try:
            for decode in list(root.findall(".//" + self.q("Decode"))):
                texts = decode.findall(".//" + self.q("TranslatedText"))
                all_blank = True
                for tt in texts:
                    if safe_text(tt.text):
                        all_blank = False
                        break
                if all_blank:
                    parent = None
                    for elem in root.iter():
                        if decode in list(elem):
                            parent = elem
                            break
                    if parent is not None:
                        parent.remove(decode)
        except Exception:
            pass


    def write(self):
        self.out_dir.mkdir(parents=True, exist_ok=True)
        root = self.build()
        self.remove_blank_decode_nodes(root)
        tree = ET.ElementTree(root)
        ET.indent(tree, space="  ", level=0)
        out_xml = self.out_dir / "define.xml"

        # ElementTree cannot write the stylesheet PI between XML declaration and root
        # in the exact legacy form, so write XML body first and then prepend headers.
        xml_body = ET.tostring(root, encoding="unicode", short_empty_elements=True)
        header = (
            '<?xml version="1.0" encoding="ISO-8859-1" ?>\n'
            f'<?xml-stylesheet type="text/xsl" href="{self.define_xsl}"?>\n'
        )
        # Use XML character references for anything outside ISO-8859-1.
        out_xml.write_bytes((header + xml_body).encode("ISO-8859-1", errors="xmlcharrefreplace"))
        return out_xml


# ================================================================================================
# Main GUI
# ================================================================================================

class DefineStudio(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.spec_path = ""
        self.dataset_path = ""
        self.excel_data = {}
        self.normalized_spec = pd.DataFrame(columns=SPEC_COLUMNS)
        self.domains_df = pd.DataFrame(columns=DOMAIN_COLUMNS)
        self.documents_df = pd.DataFrame(columns=DOCUMENTS_COLUMNS)
        self.document_links_df = pd.DataFrame(columns=DOCUMENT_LINKS_COLUMNS)
        self.dataset_inventory = pd.DataFrame()
        self.dataset_metadata = pd.DataFrame()
        self.datasets = {}  # dataset name -> dataframe
        self.editor_df = pd.DataFrame(columns=EDITOR_COLUMNS)
        self.formats_df = pd.DataFrame(columns=FORMAT_COLUMNS)
        self.vlm_df = pd.DataFrame(columns=VLM_COLUMNS)
        self.spec_vlm_df = pd.DataFrame(columns=VLM_COLUMNS)
        self.validation_df = pd.DataFrame()
        self.cdisc_api_key = ""
        self.ct_alias_map = {}
        self.spec_model = None
        self.dataset_metadata_model = None
        self.review_export_available = False
        self.define_generated = False

        self.setWindowTitle("Define XML Generator")
        self.setMinimumSize(1300, 800)
        self.resize(1750, 980)
        self.build_ui()
        self.set_workflow_state("initial")

        # JSON-only mode: automatically load define_config.json from script folder.
        try:
            self.load_config()
        except Exception:
            pass
        self.apply_style()

    def build_ui(self):
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(10, 8, 10, 8)
        root.setSpacing(8)

        header = QtWidgets.QLabel("Define XML Generator")
        header.setAlignment(QtCore.Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                color: black; background: #bfe9f7; padding: 10px 0 8px 0; border-radius: 16px;
                font-family: 'Times New Roman'; font-size: 20pt; font-weight: bold;
            }
        """)
        root.addWidget(header)

        contact_note = QtWidgets.QLabel(
            "For queries / suggestions / issues: Manivannan.Mathialagan@veristat.com"
        )
        contact_note.setAlignment(QtCore.Qt.AlignCenter)
        contact_note.setWordWrap(True)
        contact_note.setStyleSheet("""
            QLabel {
                background: #fff3c9;
                color: #184a78;
                border: 1px solid #e6d27d;
                border-radius: 10px;
                padding: 6px 10px;
                font-family: 'Times New Roman';
                font-size: 11pt;
                font-style: italic;
            }
        """)
        root.addWidget(contact_note)

        controls = QtWidgets.QFrame(); controls.setObjectName("Controls")
        self.controls_frame = controls
        grid = QtWidgets.QGridLayout(controls); grid.setContentsMargins(12, 10, 12, 10)
        self.controls_grid = grid
        self.config_edit = QtWidgets.QLineEdit(DEFAULT_CONFIG_PATH)
        self.config_edit.setVisible(False)

        self.site_edit = QtWidgets.QLineEdit(DEFAULT_SITE_NAME)
        self.sp_file_edit = QtWidgets.QLineEdit(DEFAULT_SPEC_FILE_PATH)
        self.dataset_edit = QtWidgets.QLineEdit()
        self.btn_data = QtWidgets.QPushButton("Browse XPT Folder")
        self.btn_load_spec = QtWidgets.QPushButton("Load Spec and Data")
        self.btn_refresh_formats = QtWidgets.QPushButton("Generate Format and VLM")
        self.load_spec_vlm_chk = QtWidgets.QCheckBox("Load VLM from SharePoint spec ValueMetadata")
        self.load_spec_vlm_chk.setChecked(True)
        self.btn_validate = QtWidgets.QPushButton("Validate Define Inputs")
        self.btn_define = QtWidgets.QPushButton("Generate Define.xml")
        self.btn_export = QtWidgets.QPushButton("Export Review XLSX")

        self.standard_combo = QtWidgets.QComboBox(); self.standard_combo.addItems(["SDTM", "ADaM"])
        self.define_combo = QtWidgets.QComboBox(); self.define_combo.addItems(["2.0", "2.1"])
        self.ig_edit = QtWidgets.QLineEdit("3.4")
        self.ct_edit = QtWidgets.QLineEdit("2024-03-29")
        self.study_oid_edit = QtWidgets.QLineEdit("STUDY")
        self.study_name_edit = QtWidgets.QLineEdit("Study")
        self.protocol_edit = QtWidgets.QLineEdit("Protocol")
        self.study_desc_edit = QtWidgets.QLineEdit("Study")
        self.odm_edit = QtWidgets.QLineEdit("1.3.2")
        self.meddra_edit = QtWidgets.QLineEdit("")
        self.whodrug_edit = QtWidgets.QLineEdit("")
        self.acrf_file_edit = QtWidgets.QLineEdit("acrf.pdf")
        self.csdrg_file_edit = QtWidgets.QLineEdit("csdrg.pdf")
        self.adrg_file_edit = QtWidgets.QLineEdit("adrg.pdf")
        self.include_acrf = QtWidgets.QCheckBox("Include aCRF")
        self.include_acrf.setChecked(True)
        self.include_rg = QtWidgets.QCheckBox("Include RG")
        self.include_rg.setChecked(True)
        # Compact JSON-driven action panel: primary workflow buttons plus ValueMetadata checkbox.
        # SharePoint/spec/XPT/study metadata fields still exist and are populated from
        # define_config.json, but they are intentionally not displayed in the UI.
        self.hidden_config_widgets = [
            self.config_edit, self.site_edit, self.sp_file_edit, self.dataset_edit,
            self.standard_combo, self.define_combo, self.ig_edit, self.ct_edit,
            self.study_oid_edit, self.study_name_edit, self.protocol_edit,
            self.study_desc_edit, self.odm_edit, self.meddra_edit, self.whodrug_edit,
            self.acrf_file_edit, self.csdrg_file_edit, self.adrg_file_edit,
            self.include_acrf, self.include_rg, self.btn_data,
        ]
        for w in self.hidden_config_widgets:
            w.setVisible(False)

        self.action_buttons = [
            self.btn_load_spec,
            self.btn_refresh_formats,
            self.btn_validate,
            self.btn_define,
            self.btn_export,
        ]
        button_colors = [
            ("#2f80ed", "#1f64c8"),  # Load Spec and Data
            ("#7b61ff", "#5f45d8"),  # Generate Format and VLM
            ("#f59f00", "#d98200"),  # Validate Define Inputs
            ("#d6336c", "#ad2453"),  # Generate Define.xml
            ("#495057", "#343a40"),  # Export Review XLSX
        ]
        for i, btn in enumerate(self.action_buttons):
            base, hover = button_colors[i]
            btn.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
            btn.setMinimumWidth(0)
            btn.setStyleSheet(
                f"QPushButton {{ background-color: {base}; color: white; border-radius: 10px; padding: 6px 10px; min-height: 34px; font-family: 'Times New Roman'; font-size: 11pt; font-weight: bold; }}"
                f"QPushButton:hover {{ background-color: {hover}; }}"
                "QPushButton:disabled { background-color: #d9d9d9; color: #7a7a7a; }"
            )
            grid.addWidget(btn, 0, i)
            grid.setColumnStretch(i, 1)
        grid.addWidget(self.load_spec_vlm_chk, 1, 0, 1, len(self.action_buttons))
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(6)
        controls.setMaximumHeight(90)
        root.addWidget(controls)

        # Status panel is intentionally placed near the top so workflow progress is visible
        # even when the user is working inside any tab. New messages are appended at the bottom.
        self.status_list = QtWidgets.QListWidget()
        self.status_list.setObjectName("StatusList")
        self.status_list.setMaximumHeight(64)
        self.status_list.setAlternatingRowColors(False)
        self.status_list.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        root.addWidget(self.status_list)

        self.status = QtWidgets.QLabel("")
        self.status.setVisible(False)

        self.tabs = QtWidgets.QTabWidget()
        self.tabs.tabBar().setExpanding(True)
        self.tabs.tabBar().setUsesScrollButtons(False)
        root.addWidget(self.tabs, 1)

        self.tab_summary = QtWidgets.QWidget(); self.summary_text = QtWidgets.QTextEdit(); self.summary_text.setReadOnly(True)
        lay = QtWidgets.QVBoxLayout(self.tab_summary); lay.addWidget(self.summary_text); self.tabs.addTab(self.tab_summary, "Summary")

        self.tab_spec = QtWidgets.QWidget(); spec_layout = QtWidgets.QVBoxLayout(self.tab_spec)
        self.spec_view = QtWidgets.QTableView(); self.spec_view.setSortingEnabled(True)
        self.spec_model = PandasTableModel(pd.DataFrame(columns=SPEC_COLUMNS)); self.spec_view.setModel(self.spec_model)
        spec_layout.addWidget(QtWidgets.QLabel("Loaded spec metadata is shown immediately after Load Spec. This is read-only."))
        self.spec_proxy = QtCore.QSortFilterProxyModel(self); self.spec_proxy.setSourceModel(self.spec_model); self.spec_proxy.setFilterKeyColumn(-1); self.spec_proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.spec_view.setModel(self.spec_proxy)
        self.spec_filter = QtWidgets.QLineEdit(); self.spec_filter.setPlaceholderText("Filter loaded spec..."); self.spec_filter.textChanged.connect(lambda t: self.apply_tab_filter(self.spec_proxy, t))
        spec_layout.addWidget(self.spec_filter)
        spec_layout.addWidget(self.spec_view); self.tabs.addTab(self.tab_spec, "Loaded Spec")

        self.tab_xpt = QtWidgets.QWidget(); xpt_layout = QtWidgets.QVBoxLayout(self.tab_xpt)
        self.dataset_metadata_view = QtWidgets.QTableView(); self.dataset_metadata_view.setSortingEnabled(True)
        self.dataset_metadata_model = PandasTableModel(pd.DataFrame()); self.dataset_metadata_view.setModel(self.dataset_metadata_model)
        self.dataset_metadata_proxy = QtCore.QSortFilterProxyModel(self); self.dataset_metadata_proxy.setSourceModel(self.dataset_metadata_model); self.dataset_metadata_proxy.setFilterKeyColumn(-1); self.dataset_metadata_proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive); self.dataset_metadata_view.setModel(self.dataset_metadata_proxy)
        self.dataset_metadata_filter = QtWidgets.QLineEdit(); self.dataset_metadata_filter.setPlaceholderText("Filter XPT metadata..."); self.dataset_metadata_filter.textChanged.connect(lambda t: self.apply_tab_filter(self.dataset_metadata_proxy, t)); xpt_layout.addWidget(self.dataset_metadata_filter)
        xpt_layout.addWidget(self.dataset_metadata_view)
        self.tabs.addTab(self.tab_xpt, "XPT Metadata")

        self.tab_editor = QtWidgets.QWidget(); editor_layout = QtWidgets.QVBoxLayout(self.tab_editor)
        self.editor_view = QtWidgets.QTableView(); self.editor_view.setSortingEnabled(True); self.editor_view.setAlternatingRowColors(False)
        self.editor_model = PandasTableModel(self.editor_df, EDITABLE_COLUMNS); self.editor_view.setModel(self.editor_model)
        self.editor_proxy = QtCore.QSortFilterProxyModel(self); self.editor_proxy.setSourceModel(self.editor_model); self.editor_proxy.setFilterKeyColumn(-1); self.editor_proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive); self.editor_view.setModel(self.editor_proxy)
        editor_layout.addWidget(QtWidgets.QLabel("Editable columns: Type, Format, Origin, Comments. Length comes from dataset metadata. KEEP=1 only flows to Define."))
        self.editor_filter = QtWidgets.QLineEdit(); self.editor_filter.setPlaceholderText("Filter metadata editor..."); self.editor_filter.textChanged.connect(lambda t: self.apply_tab_filter(self.editor_proxy, t)); editor_layout.addWidget(self.editor_filter)
        editor_layout.addWidget(self.editor_view); self.tabs.addTab(self.tab_editor, "Metadata Editor")

        self.tab_formats = QtWidgets.QWidget(); fmt_layout = QtWidgets.QVBoxLayout(self.tab_formats)
        self.formats_view = QtWidgets.QTableView(); self.formats_model = PandasTableModel(self.formats_df, set(FORMAT_COLUMNS)); self.formats_view.setModel(self.formats_model)
        self.formats_proxy = QtCore.QSortFilterProxyModel(self); self.formats_proxy.setSourceModel(self.formats_model); self.formats_proxy.setFilterKeyColumn(-1); self.formats_proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive); self.formats_view.setModel(self.formats_proxy)
        fmt_layout.addWidget(QtWidgets.QLabel("Formats/CT generated only from KEEP=1 metadata rows where Format is non-missing. Blank Decode is automatically set to Code during refresh. All columns are editable for review."))
        self.formats_filter = QtWidgets.QLineEdit(); self.formats_filter.setPlaceholderText("Filter formats / CT..."); self.formats_filter.textChanged.connect(lambda t: self.apply_tab_filter(self.formats_proxy, t)); fmt_layout.addWidget(self.formats_filter)
        fmt_layout.addWidget(self.formats_view); self.tabs.addTab(self.tab_formats, "Formats / CT")

        self.tab_vlm = QtWidgets.QWidget(); vlm_layout = QtWidgets.QVBoxLayout(self.tab_vlm)
        self.vlm_view = QtWidgets.QTableView(); self.vlm_model = PandasTableModel(self.vlm_df, set(VLM_COLUMNS)); self.vlm_view.setModel(self.vlm_model)
        self.vlm_proxy = QtCore.QSortFilterProxyModel(self); self.vlm_proxy.setSourceModel(self.vlm_model); self.vlm_proxy.setFilterKeyColumn(-1); self.vlm_proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive); self.vlm_view.setModel(self.vlm_proxy)
        vlm_layout.addWidget(QtWidgets.QLabel("VLM is auto-generated: SDTM --ORRES and SUPP-- QVAL only; ADaM chooses AVAL or AVALC per PARAMCD/PARAM where-clause."))
        self.vlm_filter = QtWidgets.QLineEdit(); self.vlm_filter.setPlaceholderText("Filter value metadata..."); self.vlm_filter.textChanged.connect(lambda t: self.apply_tab_filter(self.vlm_proxy, t)); vlm_layout.addWidget(self.vlm_filter)
        vlm_layout.addWidget(self.vlm_view); self.tabs.addTab(self.tab_vlm, "Value Metadata")

        self.tab_validation = QtWidgets.QWidget(); val_layout = QtWidgets.QVBoxLayout(self.tab_validation)
        self.validation_summary_label = QtWidgets.QLabel("Errors - 0, Warnings - 0")
        self.validation_summary_label.setObjectName("ValidationSummary")
        val_layout.addWidget(self.validation_summary_label)
        self.validation_view = QtWidgets.QTableView(); self.validation_model = PandasTableModel(pd.DataFrame(columns=["Severity", "Check", "Dataset", "Variable", "Message"])); self.validation_view.setModel(self.validation_model)
        self.validation_proxy = QtCore.QSortFilterProxyModel(self); self.validation_proxy.setSourceModel(self.validation_model); self.validation_proxy.setFilterKeyColumn(-1); self.validation_proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive); self.validation_view.setModel(self.validation_proxy)
        self.validation_filter = QtWidgets.QLineEdit(); self.validation_filter.setPlaceholderText("Filter validation issues..."); self.validation_filter.textChanged.connect(lambda t: self.apply_tab_filter(self.validation_proxy, t)); val_layout.addWidget(self.validation_filter)
        val_layout.addWidget(self.validation_view); self.tabs.addTab(self.tab_validation, "Validation")

        self.set_status("Ready - Define XML Generator loaded", state="info")
        self.btn_data.clicked.connect(self.browse_data)
        self.btn_load_spec.clicked.connect(self.load_spec_data_and_build_metadata)
        self.btn_refresh_formats.clicked.connect(self.generate_formats_and_vlm)
        self.btn_validate.clicked.connect(self.validate_define_inputs)
        self.btn_define.clicked.connect(self.generate_define_xml)
        self.btn_export.clicked.connect(self.export_review_xlsx)
        self.load_json_config()

    def apply_style(self):
        self.setStyleSheet(f"""
            QWidget {{ background-color: {THEME['app_bg']}; font-family: Segoe UI; font-size: 10pt; color: #1c2e4a; }}
            #Header {{ background-color: {THEME['header']}; border-radius: 14px; border: 1px solid #9cc8f5; }}
            #Title {{ font-size: 20pt; font-weight: bold; color: {THEME['header_text']}; }}
            #Subtitle {{ color: {THEME['subtitle']}; font-weight: bold; }}
            #Controls {{ background-color: white; border-radius: 12px; border: 1px solid {THEME['border']}; }}
            QPushButton {{ background-color: {THEME['button']}; color: white; border-radius: 8px; padding: 7px 12px; font-weight: bold; }}
            QPushButton:hover {{ background-color: {THEME['button_hover']}; }}
            QLineEdit, QComboBox {{ background-color: white; border: 1px solid #a8bfdc; border-radius: 8px; padding: 5px; }}
            QTableView {{ background-color: white; border: 1px solid #8fb2d6; gridline-color: #c9d7e6; }}
            QHeaderView::section {{ background-color: {THEME['table_header']}; padding: 6px; border: 1px solid #8fb2d6; font-weight: bold; color: {THEME['header_text']}; }}
            QTextEdit {{ background-color: white; border: 1px solid #b7cde8; border-radius: 8px; padding: 8px; }}
            #Status {{ background-color: {THEME['status']}; border: 1px solid #dbc46f; border-radius: 8px; padding: 6px; color: #4d3b00; font-weight: bold; }}
            #StatusList {{ background-color: white; border: 1px solid #b7cde8; border-radius: 8px; padding: 4px; font-weight: bold; }}
        """)

    def set_status(self, msg, state="info"):
        """Append one status message to the bottom of the status list.

        state values:
          running = blue line for the active step
          done    = green line for completed step
          error   = red line for failed step
          info    = yellow/neutral line for general notes
        """
        msg = safe_text(msg)
        if not msg:
            return
        stamp = datetime.now().strftime("%H:%M:%S")
        label = f"{stamp}  {msg}"
        self.status.setText(label)
        if hasattr(self, "status_list"):
            item = QtWidgets.QListWidgetItem(label)
            color_map = {
                "running": THEME.get("status_running", "#d9ecff"),
                "done": THEME.get("status_done", "#e7f7e7"),
                "error": THEME.get("status_error", "#ffe6e6"),
                "info": THEME.get("status_info", "#fff8dc"),
            }
            item.setBackground(QtGui.QColor(color_map.get(state, color_map["info"])))
            if state == "error":
                item.setForeground(QtGui.QColor("#8a0000"))
            elif state == "running":
                item.setForeground(QtGui.QColor("#003c78"))
            elif state == "done":
                item.setForeground(QtGui.QColor("#155724"))
            self.status_list.addItem(item)
            self.status_list.scrollToBottom()
        QtWidgets.QApplication.processEvents()

    def set_buttons_busy(self, busy=True):
        for btn in getattr(self, "action_buttons", []):
            btn.setEnabled(not busy)
        QtWidgets.QApplication.processEvents()

    def apply_tab_filter(self, proxy, text):
        # Case-insensitive filter across all columns in the current table.
        proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive)
        proxy.setFilterKeyColumn(-1)
        proxy.setFilterFixedString(text)

    def _hide_config_fields_when_json_locked(self):
        """Keep the top panel button-only.

        Values are read from define_config.json and kept in hidden widgets for downstream
        code. The visible UI shows the main workflow buttons and the ValueMetadata checkbox.
        """
        for w in getattr(self, "hidden_config_widgets", []):
            w.setVisible(False)
        for btn in getattr(self, "action_buttons", []):
            btn.setVisible(True)
        try:
            self.controls_frame.setMaximumHeight(96)
        except Exception:
            pass

    def browse_config(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select JSON config", str(Path.home()), "JSON Files (*.json);;All Files (*)")
        if path:
            self.config_edit.setText(path)
            if True:
                self.load_json_config()

    def load_json_config(self):
        """Load define_config.json from the same folder as this .pyw.

        Supported JSON styles:
        1) Nested sections: sharepoint/spec, study, versions, paths, documents, gui
        2) Flat keys retained for backward compatibility.
        """
        # Config file name/location is fixed: define_config.json beside this script.
        path = DEFAULT_CONFIG_PATH
        if not os.path.exists(path):
            self.set_status(f"define_config.json not found in script folder; using current/default values: {path}")
            self.toggle_config_lock()
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Invalid JSON config", str(e))
            return

        def nested_get(*paths, default=""):
            for path_keys in paths:
                cur = cfg
                ok = True
                for k in path_keys.split("."):
                    if isinstance(cur, dict) and k in cur:
                        cur = cur[k]
                    else:
                        ok = False
                        break
                if ok and cur is not None:
                    return cur
            return default

        def set_text(widget, *paths, default=""):
            val = nested_get(*paths, default=default)
            if val is not None and str(val) != "":
                widget.setText(str(val))

        set_text(self.site_edit, "sharepoint.site_name", "sharepoint_site", "site", "site_name")
        set_text(self.sp_file_edit, "sharepoint.spec_file", "sharepoint.file_path", "sharepoint_file", "spec_file", "file_path", "spec_path")
        set_text(self.dataset_edit, "paths.xpt_folder", "paths.xpt_path", "xpt_folder", "xpt_path", "data_path", "dataset_path")

        set_text(self.ig_edit, "versions.ig_version", "ig_version", "sdtm_ig_version", "adam_ig_version")
        set_text(self.ct_edit, "versions.ct_version", "ct_version")
        set_text(self.odm_edit, "versions.odm_version", "odm_version", default="1.3.2")
        set_text(self.meddra_edit, "versions.meddra_version", "meddra_version")
        set_text(self.whodrug_edit, "versions.whodrug_version", "whodrug_version", "who_drug_version")
        self.cdisc_api_key = safe_text(nested_get("cdisc_library.api_key", "cdisc.api_key", "cdisc_api_key", "api_key"))

        set_text(self.study_oid_edit, "study.study_oid", "study_oid")
        set_text(self.study_name_edit, "study.study_name", "study_name")
        set_text(self.study_desc_edit, "study.study_description", "study.description", "study_description")
        set_text(self.protocol_edit, "study.protocol", "protocol")

        set_text(self.acrf_file_edit, "documents.acrf_file", "acrf_file", default="acrf.pdf")
        set_text(self.csdrg_file_edit, "documents.csdrg_file", "csdrg_file", default="csdrg.pdf")
        set_text(self.adrg_file_edit, "documents.adrg_file", "adrg_file", default="adrg.pdf")

        std = safe_upper(nested_get("study.standard", "standard"))
        if std in {"SDTM", "ADAM"}:
            self.standard_combo.setCurrentText("ADaM" if std == "ADAM" else "SDTM")
        dv = safe_text(nested_get("versions.define_version", "define_version"))
        if dv in {"2.0", "2.1"}:
            self.define_combo.setCurrentText(dv)

        inc_acrf = nested_get("documents.include_acrf", "include_acrf", default=None)
        if inc_acrf is not None:
            self.include_acrf.setChecked(bool(inc_acrf))
        inc_rg = nested_get("documents.include_rg", "include_rg", default=None)
        if inc_rg is not None:
            self.include_rg.setChecked(bool(inc_rg))

        use_spec_vlm = nested_get("gui.load_vlm_from_sharepoint_spec", "load_vlm_from_sharepoint_spec", default=None)
        if use_spec_vlm is not None and hasattr(self, "load_spec_vlm_chk"):
            self.load_spec_vlm_chk.setChecked(bool(use_spec_vlm))

        # Allow JSON to control initial checkbox state, default checked.
        use_json = nested_get("gui.use_json_defaults", "use_json_defaults", default=True)

        self.toggle_config_lock()
        self.set_status("JSON config loaded: define_config.json")

    def toggle_config_lock(self):
        locked = True
        # When checked, fields are fed from JSON and locked. Uncheck to edit individual boxes.
        widgets = [
            self.site_edit, self.sp_file_edit, self.dataset_edit,
            self.standard_combo, self.define_combo, self.ig_edit, self.ct_edit, self.odm_edit,
            self.meddra_edit, self.whodrug_edit, self.study_oid_edit, self.study_name_edit,
            self.study_desc_edit, self.protocol_edit, self.acrf_file_edit, self.csdrg_file_edit,
            self.adrg_file_edit, self.include_acrf, self.include_rg,
        ]
        for w in widgets:
            w.setEnabled(not locked)
        self.config_edit.setEnabled(False)
        self.config_edit.setVisible(False)
        self._hide_config_fields_when_json_locked()

    def browse_data(self):
        path = QtWidgets.QFileDialog.getExistingDirectory(self, "Select XPT dataset folder", str(Path.home()))
        if path:
            self.dataset_edit.setText(path)

    def load_spec(self):
        site_name = self.site_edit.text().strip()
        file_path = self.sp_file_edit.text().strip()
        if not site_name or not file_path:
            QtWidgets.QMessageBox.warning(self, "SharePoint details missing", "Please enter SharePoint Site and Spec File path.")
            return
        try:
            path = download_spec_from_sharepoint(site_name, file_path, self.set_status)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "SharePoint spec download failed", str(e))
            self.set_status("SharePoint spec download failed")
            return

        self.set_status("Loading SharePoint specification...")
        self.spec_path = path
        self.excel_data = read_all_sheets(path)
        self.domains_df = read_domains_sheet(path)
        self.spec_vlm_df = read_value_metadata_sheet(path)
        self.documents_df = read_documents_sheet(path)
        self.document_links_df = read_document_links_sheet(path)
        frames = []
        included = []
        excluded = []
        for sheet, df in self.excel_data.items():
            if is_domain_metadata_sheet(sheet, df):
                frames.append(df.copy())
                included.append(sheet)
            else:
                excluded.append(sheet)
        for supp_ds in get_unique_supp_datasets_from_suppqual(path):
            supp_rows = build_supp_rows_from_template(path, supp_ds)
            if not supp_rows.empty:
                frames.append(supp_rows)
        self.normalized_spec = pd.concat(frames, ignore_index=True)[SPEC_COLUMNS] if frames else pd.DataFrame(columns=SPEC_COLUMNS)
        self.summary_text.setPlainText(
            "Specification loaded from SharePoint\n" + "="*80 + "\n\n" +
            f"SharePoint Site: {site_name}\n" +
            f"SharePoint File: {file_path}\n" +
            f"Included domain sheets: {len(included)}\n" + "\n".join(f"  - {s}" for s in included) + "\n\n" +
            f"Excluded helper sheets: {len(excluded)}\n" + "\n".join(f"  - {s}" for s in excluded) + "\n\n" +
            f"Domains sheet rows: {len(self.domains_df)}\n" +
            f"Normalized metadata rows: {len(self.normalized_spec)}\n" +
            f"ValueMetadata rows: {len(self.spec_vlm_df)}\n"
        )
        self.spec_model.set_df(self.normalized_spec)
        self.spec_view.resizeColumnsToContents()
        self.tabs.setCurrentWidget(self.tab_spec)
        self.set_status(f"Spec loaded from SharePoint: {len(self.normalized_spec)} normalized rows")

    def load_data(self):
        folder = self.dataset_edit.text().strip()
        if not folder or not os.path.isdir(folder):
            QtWidgets.QMessageBox.warning(self, "XPT folder missing", "Please select a valid XPT data folder.")
            return
        if pyreadstat is None:
            QtWidgets.QMessageBox.critical(self, "pyreadstat missing", "pyreadstat could not be installed/imported.")
            return
        self.set_status("Reading dataset metadata and values...")
        self.dataset_path = folder
        inv_rows = []
        meta_rows = []
        self.datasets = {}
        files = [p for p in Path(folder).glob("*.xpt")]
        if not files:
            QtWidgets.QMessageBox.warning(self, "No XPT files", "The selected folder does not contain any .xpt files. Only XPT data is accepted.")
            self.set_status("No XPT files found")
            return
        for p in files:
            ds = p.stem.upper()
            try:
                df, meta = read_dataset_file(p, metadataonly=False)
                df.columns = [safe_upper(c) for c in df.columns]
                self.datasets[ds] = df
                names = list(getattr(meta, "column_names", []) or list(df.columns))
                labels = dict(zip([safe_upper(x) for x in names], list(getattr(meta, "column_labels", []) or [""]*len(names))))
                storage_types = getattr(meta, "readstat_variable_types", {}) or {}
                formats = getattr(meta, "variable_display_width", {}) or {}
                original_types = getattr(meta, "original_variable_types", {}) or {}
                labels2 = getattr(meta, "column_names_to_labels", {}) or {}
                inv_rows.append({
                    "Dataset": ds,
                    "File": p.name,
                    "Variables": len(df.columns),
                    "Records": len(df),
                    "Size MB": round(p.stat().st_size/(1024*1024), 2),
                })
                for i, col in enumerate(df.columns, start=1):
                    stype = storage_types.get(col) or storage_types.get(col.lower()) or storage_types.get(col.upper()) or str(df[col].dtype)
                    label = labels.get(col) or labels2.get(col) or labels2.get(col.lower()) or ""
                    inferred_type, inferred_len, inferred_fmt = infer_type_from_values(df[col], stype, "")
                    data_len = inferred_len if inferred_type == "char" else 8
                    meta_rows.append({
                        "Dataset": ds, "Variable": col, "Label": safe_text(label),
                        "Data Type": inferred_type, "Data Length": data_len,
                        "Data Format": inferred_fmt, "Storage Type": normalize_dataset_type(stype),
                        "Order": i, "Source": p.name,
                    })
            except Exception as e:
                inv_rows.append({"Dataset": ds, "File": p.name, "Variables": "", "Records": "", "Size MB": "", "Warning": str(e)})
        self.dataset_inventory = pd.DataFrame(inv_rows)
        self.dataset_metadata = pd.DataFrame(meta_rows)
        self.dataset_metadata_model.set_df(self.dataset_metadata)
        self.dataset_metadata_view.resizeColumnsToContents()
        self.tabs.setCurrentWidget(self.tab_xpt)
        self.summary_text.append("\nData loaded\n" + "="*80 + f"\nFolder: {folder}\nDatasets: {len(self.datasets)}\nVariables: {len(self.dataset_metadata)}\n")
        self.set_status(f"XPT data loaded: {len(self.datasets)} datasets, {len(self.dataset_metadata)} variables")

    def build_metadata_editor(self):
        if self.normalized_spec.empty:
            QtWidgets.QMessageBox.warning(self, "Spec not loaded", "Please load spec first.")
            return
        if self.dataset_metadata.empty:
            QtWidgets.QMessageBox.warning(self, "XPT data not loaded", "Please load XPT data first. Metadata Editor will include only KEEP=1 variables that are present in the selected XPT files.")
            return
        self.set_status("Building editable metadata table...")
        spec = self.normalized_spec.copy()
        spec["Dataset"] = spec["Dataset"].apply(safe_upper)
        spec["Variable"] = spec["Variable"].apply(safe_upper)
        spec = spec[spec["Keep"].apply(normalize_keep)].copy()
        dmeta = self.dataset_metadata.copy()
        if not dmeta.empty:
            dmeta["Dataset"] = dmeta["Dataset"].apply(safe_upper)
            dmeta["Variable"] = dmeta["Variable"].apply(safe_upper)
        # Metadata Editor should include only variables that are both:
        #   1) KEEP=1 in the specification
        #   2) physically present in the selected XPT datasets
        # Use inner join so spec-only variables are excluded from the editable define metadata.
        merged = spec.merge(dmeta, on=["Dataset", "Variable"], how="inner", suffixes=("", "_DATA"))
        rows = []
        for _, r in merged.iterrows():
            data_type = safe_text(r.get("Data Type"))
            spec_type = safe_text(r.get("Type"))
            final_type = data_type or spec_type or "char"
            data_len = safe_text(r.get("Data Length"))
            spec_len = safe_text(r.get("Len"))
            final_len = data_len or spec_len or ("200" if final_type == "char" else "8")
            fmt = default_format_for_variable(r.get("Variable"), safe_text(r.get("Control or Format")))
            if is_dtc_iso_variable(r.get("Variable"), fmt):
                # Keep final_len from XPT metadata, but show correct Define metadata.
                final_type = "date"
                fmt = "ISO 8601"
            rows.append({
                "Dataset": safe_upper(r.get("Dataset")),
                "Variable": safe_upper(r.get("Variable")),
                "Label": safe_text(r.get("Label")) or safe_text(r.get("Label_DATA")),
                "Keep": sas_best_text(r.get("Keep")),
                "ID Var": sas_best_text(r.get("ID Var")),
                "Length": final_len,
                "Type": final_type,
                "Format": fmt,
                "Origin": safe_text(r.get("Origin")),
                "Comments": safe_text(r.get("Comments")),
                "Data Type": data_type,
                "Data Format": safe_text(r.get("Data Format")),
                "Order": safe_text(r.get("Order")),
                "Source": safe_text(r.get("Source")),
            })
        self.editor_df = pd.DataFrame(rows, columns=EDITOR_COLUMNS)
        self.editor_model.set_df(self.editor_df)
        self.install_editor_delegates()
        self.editor_view.resizeColumnsToContents()
        self.editor_view.setColumnWidth(2, 280)
        self.editor_view.setColumnWidth(9, 320)
        self.set_status(f"Metadata editor built with KEEP=1 and XPT-present rows: {len(self.editor_df)}")

    def install_editor_delegates(self):
        for col_name, options in [("Type", TYPE_OPTIONS), ("Origin", ORIGIN_OPTIONS)]:
            if col_name in self.editor_model.df.columns:
                idx = list(self.editor_model.df.columns).index(col_name)
                self.editor_view.setItemDelegateForColumn(idx, ComboDelegate(options, self.editor_view))

    def current_editor_df(self):
        self.editor_df = self.editor_model.df.copy()
        return self.editor_df.copy()

    def generate_formats(self):
        meta = self.current_editor_df()
        if meta.empty:
            QtWidgets.QMessageBox.warning(self, "Metadata missing", "Build Metadata Editor first.")
            return
        self.set_status("Generating formats from spec-defined codelists and XPT values...")
        ct_version = self.ct_edit.text().strip()
        self.ct_alias_map = fetch_define_ct_map(
            self.cdisc_api_key,
            self.standard_combo.currentText(),
            sdtm_ct_version=ct_version,
            adam_ct_version=ct_version,
            status_callback=self.set_status
        )
        rows = []
        seen_sources = set()

        for _, m in meta.iterrows():
            ds = safe_upper(m.get("Dataset"))
            spec_var = safe_upper(m.get("Variable"))
            fmt = default_format_for_variable(spec_var, safe_text(m.get("Format")))
            vtype = safe_text(m.get("Type"))

            # Formats/CT are created only when the spec/GUI Format is non-missing
            # and it is a real codelist format. Display formats like ISO8601 are skipped.
            if not is_define_codelist_format(fmt, spec_var, vtype):
                continue
            # External dictionaries are linked in define.xml as def:ExternalCodeList,
            # not expanded into CodeListItem rows in the Formats/CT sheet.
            if is_external_dictionary_format(fmt):
                continue
            if ds not in self.datasets:
                continue

            df = self.datasets[ds]
            cols_upper = {safe_upper(c): c for c in df.columns}

            # Codelist behavior depends on where the format is specified:
            #   * Format on code variable (ARMCD/PARAMCD/--TESTCD): code = code variable, decode = companion text.
            #   * Format on decode/text variable (ARM/ELEMENT/VISIT/PARAM/ACTARM): code = that text variable,
            #     decode is blank, and the companion code/order variable is used only for sorting.
            possible_code = companion_code_var(spec_var)
            std_is_adam = safe_upper(self.standard_combo.currentText()) == "ADAM"

            if std_is_adam and possible_code and possible_code in cols_upper and is_adam_order_display_pair(spec_var, possible_code):
                # ADaM value-pair rule when the format is on the character/display variable:
                #   character variable = permitted/display value written to the codelist
                #   numeric companion  = sort/order value only
                # Examples:
                #   RACE  / RACEN   -> ADSL.RACE  Code=RACE values,  sorted by RACEN
                #   PARAM / PARAMN  -> PARAM      Code=PARAM values, sorted by PARAMN
                #   TRT01P/TRT01PN -> TRT01P values, sorted by TRT01PN
                # Do NOT use the numeric companion as Code for the character codelist;
                # otherwise ADSL.RACE incorrectly shows Code=1,2,3 with Decode=WHITE/ASIAN.
                code_var = spec_var
                decode_var = ""
                order_seed = possible_code
            elif possible_code and possible_code in cols_upper:
                code_var = spec_var
                decode_var = ""
                order_seed = possible_code
            else:
                code_var = spec_var
                decode_var = companion_decode_var(spec_var)
                order_seed = code_var

            if code_var not in cols_upper:
                continue
            code_col = cols_upper[code_var]

            dec_col = ""
            if decode_var and decode_var in cols_upper:
                dec_col = cols_upper[decode_var]

            order_var = order_seed if order_seed in cols_upper else (companion_order_var(code_var, df.columns) or companion_order_var(spec_var, df.columns))
            order_col = ""
            if order_var and safe_upper(order_var) in cols_upper:
                order_col = cols_upper[safe_upper(order_var)]

            # Avoid generating the exact same format from both code and decode rows.
            source_key = (ds, fmt, code_col, dec_col, order_col)
            if source_key in seen_sources:
                continue
            seen_sources.add(source_key)

            use_cols = [code_col]
            if dec_col and dec_col not in use_cols:
                use_cols.append(dec_col)
            if order_col and order_col not in use_cols:
                use_cols.append(order_col)

            tmp = df[use_cols].drop_duplicates().copy()
            tmp = tmp[tmp[code_col].apply(lambda x: safe_text(x) != "")]

            sort_cols = [c for c in [order_col, code_col, dec_col] if c and c in tmp.columns]
            if sort_cols:
                tmp = tmp.sort_values(sort_cols)

            for _, tr in tmp.iterrows():
                code = sas_best_text(tr.get(code_col))
                decode = sas_best_text(tr.get(dec_col)) if dec_col else ""
                # For DOMAIN codelist, Decode should be the dataset label/description from Domains sheet.
                if safe_upper(fmt) == "DOMAIN" or safe_upper(code_col) in {"DOMAIN", "RDOMAIN"}:
                    if hasattr(self, "domains_df") and not self.domains_df.empty:
                        dmap = {safe_upper(r.get("Dataset")): safe_text(r.get("Description")) for _, r in self.domains_df.iterrows()}
                        decode = dmap.get(safe_upper(code), decode or code)
                # For self-decode codelists such as EPOCH/PARCAT1, use the value as decode.
                if not decode and safe_upper(code_col) in {"PARAM", "PARCAT1", "PARCAT2", "PARCAT3", "VISIT", "AVISIT", "EPOCH", "DOMAIN", "RDOMAIN", "ELEMENT"}:
                    decode = code
                sortv = sas_best_text(tr.get(order_col)) if order_col else ""
                ct_info = get_ct_info_generic(getattr(self, "ct_alias_map", {}), fmt)
                # Use CDISC Synonyms as Decode when CT match has a nonblank synonym.
                # If synonym is missing, preserve the already assigned Decode.
                ct_synonym = get_ct_synonym_generic(ct_info, code)
                final_decode = ct_synonym or decode or code
                term_code = get_ct_term_code_generic(ct_info, code) or known_term_nci_code(fmt, code)
                rows.append({
                    "Order": "",
                    "Format": fmt,
                    "Code": code,
                    "Decode": final_decode,
                    "Codelist Code": safe_text(ct_info.get("codelist_code", "")) or known_codelist_nci_code(fmt),
                    "Term Code": term_code,
                    "Source Dataset": ds,
                    "Source Variable": safe_upper(code_col),
                    "Decode Variable": safe_upper(dec_col),
                    "Sort Value": sortv,
                })

        out = pd.DataFrame(rows, columns=FORMAT_COLUMNS)
        if not out.empty:
            # Final safety rule: if Decode is blank, assign Decode as Code.
            out["Decode"] = out.apply(lambda r: safe_text(r.get("Decode")) or safe_text(r.get("Code")), axis=1)
            out = out.drop_duplicates(subset=["Format", "Code", "Decode"]).copy()
            out["_sort"] = out["Sort Value"].apply(sort_key_mixed)
            out = out.sort_values(["Format", "_sort", "Code", "Decode"]).drop(columns=["_sort"])
            out["Order"] = out.groupby("Format").cumcount() + 1
        self.formats_df = out

        # Re-apply CT synonym map after dataframe creation, using Format/Code.
        try:
            if hasattr(self, "ct_alias_map") and not self.formats_df.empty:
                def _decode_from_ct_map(r):
                    info = get_ct_info_generic(self.ct_alias_map, r.get("Format"))
                    syn = get_ct_synonym_generic(info, r.get("Code"))
                    return syn or safe_text(r.get("Decode")) or safe_text(r.get("Code"))
                self.formats_df["Decode"] = self.formats_df.apply(_decode_from_ct_map, axis=1)

                def _fill_ct_codes(r):
                    info = get_ct_info_generic(self.ct_alias_map, r.get("Format"))
                    return safe_text(r.get("Codelist Code")) or safe_text(info.get("codelist_code", "")) or known_codelist_nci_code(r.get("Format"))
                self.formats_df["Codelist Code"] = self.formats_df.apply(_fill_ct_codes, axis=1)

                def _fill_term_codes(r):
                    info = get_ct_info_generic(self.ct_alias_map, r.get("Format"))
                    return safe_text(r.get("Term Code")) or get_ct_term_code_generic(info, r.get("Code"))
                self.formats_df["Term Code"] = self.formats_df.apply(_fill_term_codes, axis=1)
        except Exception:
            pass


        # Remove duplicate decode-side CT rows such as ETCD rows created again from ELEMENT.
        try:
            if not self.formats_df.empty and {"Format", "Source Variable"}.issubset(self.formats_df.columns):
                mask_skip = self.formats_df.apply(
                    lambda r: should_skip_decode_side_format_row(r.get("Format"), r.get("Source Variable")),
                    axis=1
                )
                self.formats_df = self.formats_df.loc[~mask_skip].copy()
                # Also remove exact duplicate Format/Code/Decode rows.
                keep_cols = [c for c in ["Format", "Code", "Decode"] if c in self.formats_df.columns]
                if keep_cols:
                    self.formats_df = self.formats_df.drop_duplicates(subset=keep_cols, keep="first").copy()
        except Exception:
            pass

        # CT decode rule: use CT Synonyms as Decode only when Synonyms is available.
        # Otherwise preserve the already assigned Decode; never overwrite with blank.
        try:
            self.formats_df = apply_ct_synonym_decode_rule(self.formats_df)
        except Exception:
            pass

        # CDISC Library uses "CDISC Synonyms"; use it as Decode when present.
        try:
            synonym_cols = [
                "CDISC Synonyms", "CDISC Synonym", "cdisc_synonyms",
                "Synonyms", "Synonym"
            ]
            existing_syn_cols = [c for c in synonym_cols if c in self.formats_df.columns]
            if existing_syn_cols:
                syn_col = existing_syn_cols[0]
                self.formats_df["Decode"] = self.formats_df.apply(
                    lambda r: safe_text(r.get(syn_col)) or safe_text(r.get("Decode")) or safe_text(r.get("Code")),
                    axis=1
                )
        except Exception:
            pass

        # Normalize CDISC Library CT column names after CT enrichment.
        try:
            if not self.formats_df.empty:
                # If CT merge created CDISC synonym columns with suffixes, pick the first available.
                syn_candidates = [c for c in self.formats_df.columns if safe_upper(c).replace("_", " ") in {
                    "CDISC SYNONYMS", "CDISC SYNONYM", "SYNONYMS", "SYNONYM"
                }]
                if syn_candidates:
                    syn_col = syn_candidates[0]
                    self.formats_df["Decode"] = self.formats_df.apply(
                        lambda r: safe_text(r.get(syn_col)) or safe_text(r.get("Decode")) or safe_text(r.get("Code")),
                        axis=1
                    )
        except Exception:
            pass

        # Apply CDISC Synonyms from the loaded CT/browser table to Formats/CT.
        # This fixes cases like AESEV where CDISC Synonyms = "1; Grade 1".
        try:
            ct_df_for_syn = self.get_loaded_ct_dataframe() if hasattr(self, "get_loaded_ct_dataframe") else pd.DataFrame()
            self.formats_df = apply_cdisc_synonyms_from_ct_table(self.formats_df, ct_df_for_syn)
        except Exception:
            pass

        # Apply CDISC Synonyms by Term Code when Term Code is already populated.
        try:
            ct_df_for_syn = self.get_loaded_ct_dataframe() if hasattr(self, "get_loaded_ct_dataframe") else pd.DataFrame()
            self.formats_df = apply_cdisc_synonyms_by_term_code(self.formats_df, ct_df_for_syn)
        except Exception:
            pass

        # For display-text codelists, if Code and Decode are identical, suppress Decode.
        # This avoids duplicate display in Define.xml while keeping coded variables intact.
        try:
            self.formats_df = apply_display_decode_suppression(self.formats_df)
        except Exception:
            pass

        # SUPPQUAL QNAM/QLABEL pairing:
        # QNAM is the code/order, QLABEL is the decode. Do not create separate QLABEL codelist rows.
        try:
            if not self.formats_df.empty and {"Format", "Code", "Decode"}.issubset(self.formats_df.columns):
                # Drop decode-side QLABEL rows if any were created against QNAM.
                if "Source Variable" in self.formats_df.columns:
                    self.formats_df = self.formats_df[
                        ~(
                            (self.formats_df["Format"].astype(str).str.upper() == "QNAM")
                            & (self.formats_df["Source Variable"].astype(str).str.upper() == "QLABEL")
                        )
                    ].copy()
                # Ensure QNAM rows are sorted by code/QNAM.
                if "Sort Value" in self.formats_df.columns:
                    qmask = self.formats_df["Format"].astype(str).str.upper() == "QNAM"
                    self.formats_df.loc[qmask, "Sort Value"] = self.formats_df.loc[qmask, "Code"]
        except Exception:
            pass

        # Match macro-style display behavior: display-text lists with Code=Decode have blank Decode.
        try:
            self.formats_df = apply_display_decode_suppression(self.formats_df)
        except Exception:
            pass

        # Generic CDISC CT backend enrichment:
        # - resolves domain-qualified formats through normalized CT key (e.g. DM.ARMNRS -> ARMNRS)
        # - fills Codelist Code and Term Code for all CT-managed formats
        # - uses CDISC synonym as Decode when available
        try:
            ct_map = getattr(self, "ct_alias_map", {}) if hasattr(self, "ct_alias_map") else {}
            if isinstance(ct_map, dict) and not self.formats_df.empty and {"Format", "Code"}.issubset(self.formats_df.columns):
                for idx, r in self.formats_df.iterrows():
                    raw_fmt = r.get("Format")
                    fmt_std = standardize_format_for_define(raw_fmt, ct_map)
                    ct_info = get_ct_info_generic(ct_map, fmt_std) or get_ct_info_generic(ct_map, raw_fmt)
                    if ct_info:
                        self.formats_df.at[idx, "Format"] = fmt_std
                        cl_code = safe_text(ct_info.get("codelist_code"))
                        term_code = get_ct_term_code_generic(ct_info, r.get("Code"))
                        ct_syn = get_ct_synonym_generic(ct_info, r.get("Code"))
                        if cl_code and not safe_text(self.formats_df.at[idx, "Codelist Code"]):
                            self.formats_df.at[idx, "Codelist Code"] = cl_code
                        if term_code and not safe_text(self.formats_df.at[idx, "Term Code"]):
                            self.formats_df.at[idx, "Term Code"] = term_code
                        if ct_syn:
                            self.formats_df.at[idx, "Decode"] = ct_syn
        except Exception:
            pass

        # Normalize generated Format names for presence checks and CT-backed define output.
        try:
            ct_map = getattr(self, "ct_alias_map", {}) if hasattr(self, "ct_alias_map") else {}
            if isinstance(ct_map, dict) and not self.formats_df.empty and "Format" in self.formats_df.columns:
                self.formats_df["Format"] = self.formats_df["Format"].apply(
                    lambda x: standardize_format_for_define(x, ct_map)
                )
        except Exception:
            pass












        self.formats_model.set_df(self.formats_df)
        tune_table_widths(self.formats_view, {
            "Order": 55,
            "Format": 150,
            "Code": 150,
            "Decode": 260,
            "Codelist Code": 120,
            "Term Code": 110,
            "Source Dataset": 110,
            "Source Variable": 130,
            "Decode Variable": 130,
            "Sort Value": 100,
        }, max_width=220)
        self.set_status(f"Formats generated: {len(self.formats_df)} rows")
        self.tabs.setCurrentWidget(self.tab_formats)

    def generate_vlm(self):
        meta = self.current_editor_df()
        if meta.empty:
            QtWidgets.QMessageBox.warning(self, "Metadata missing", "Build Metadata Editor first.")
            return
        self.set_status("Generating VLM...")
        standard = safe_upper(self.standard_combo.currentText())
        rows = []
        use_spec_vlm = bool(getattr(self, "load_spec_vlm_chk", None) and self.load_spec_vlm_chk.isChecked())
        if use_spec_vlm:
            df = getattr(self, "spec_vlm_df", pd.DataFrame(columns=VLM_COLUMNS)).copy()
            for c in VLM_COLUMNS:
                if c not in df.columns:
                    df[c] = ""
            ordered = VLM_COLUMNS + [c for c in df.columns if c not in VLM_COLUMNS]
            self.vlm_df = df[ordered].copy()
            if self.vlm_df.empty:
                self.set_status("ValueMetadata sheet is checked but no rows were found in the SharePoint spec", "error")
                QtWidgets.QMessageBox.warning(
                    self,
                    "ValueMetadata missing/empty",
                    "Load VLM from SharePoint spec is checked, but the ValueMetadata sheet was not found or has no rows."
                )
        else:
            if standard == "SDTM":
                rows.extend(self.generate_sdtm_vlm_rows())
            else:
                rows.extend(self.generate_adam_vlm_rows())
            self.vlm_df = pd.DataFrame(rows, columns=VLM_COLUMNS)
        self.vlm_df = self.normalize_vlm_lengths_against_parent(self.vlm_df)
        self.vlm_model.set_df(self.vlm_df)
        tune_table_widths(self.vlm_view, {
            "Dataset": 90,
            "Grouping Variable": 125,
            "Group Value": 150,
            "Group Label": 180,
            "Where Clause": 260,
            "Result Variable": 140,
            "Length": 80,
            "Type": 90,
            "Format": 140,
            "Origin": 120,
            "Role": 120,
            "Comment": 300,
        }, max_width=220)
        self.set_status(f"VLM loaded from SharePoint ValueMetadata: {len(self.vlm_df)} rows" if bool(getattr(self, "load_spec_vlm_chk", None) and self.load_spec_vlm_chk.isChecked()) else f"VLM generated: {len(self.vlm_df)} rows")
        self.tabs.setCurrentWidget(self.tab_vlm)

    def result_type_len_format(self, series):
        dtype, length, fmt = infer_type_from_values(series)
        return dtype, length, fmt

    def parent_variable_metadata(self, dataset, variable):
        """Return parent variable metadata from the editor grid for a VLM row."""
        try:
            meta = self.current_editor_df()
        except Exception:
            meta = getattr(self, "editor_df", pd.DataFrame())
        if meta is None or getattr(meta, "empty", True):
            return {}
        ds = safe_upper(dataset)
        var = safe_upper(variable)
        m = meta[
            (meta["Dataset"].astype(str).str.upper() == ds)
            & (meta["Variable"].astype(str).str.upper() == var)
        ]
        if m.empty:
            return {}
        r = m.iloc[0]
        return {
            "Type": safe_text(r.get("Type")),
            "Length": safe_text(r.get("Length")),
            "Format": safe_text(r.get("Format")),
        }

    def capped_vlm_length(self, dataset, variable, value_level_length):
        """Cap VLM ItemDef Length so it never exceeds the parent variable length.

        Pinnacle 21 DD0123 is triggered when Value Level Length is greater than
        the corresponding dataset variable length.  This can happen when ORRES
        values look numeric and are inferred as numeric length 8 while the XPT
        variable (for example DAORRES) has a shorter physical length such as 3.
        """
        vlm_len = to_int_or_none(value_level_length)
        parent = self.parent_variable_metadata(dataset, variable)
        parent_len = to_int_or_none(parent.get("Length"))

        if parent_len is None:
            return safe_text(value_level_length)
        if vlm_len is None:
            return str(parent_len)
        return str(min(vlm_len, parent_len))

    def normalize_vlm_lengths_against_parent(self, vlm_df):
        """Normalize generated/imported VLM lengths against parent variable length."""
        if vlm_df is None or getattr(vlm_df, "empty", True):
            return vlm_df
        df = vlm_df.copy()
        if not {"Dataset", "Result Variable", "Length"}.issubset(df.columns):
            return df
        df["Length"] = df.apply(
            lambda r: self.capped_vlm_length(r.get("Dataset"), r.get("Result Variable"), r.get("Length")),
            axis=1
        )
        return df

    def generate_sdtm_vlm_rows(self):
        rows = []
        for ds, df in self.datasets.items():
            cols = list(df.columns)
            if ds.startswith("SUPP") and "QVAL" in cols and "QNAM" in cols:
                label_var = "QLABEL" if "QLABEL" in cols else "QNAM"
                for qnam, qdf in df.groupby("QNAM", dropna=True):
                    if safe_text(qnam) == "":
                        continue
                    qlabel = safe_text(qdf[label_var].dropna().iloc[0]) if label_var in qdf and not qdf[label_var].dropna().empty else safe_text(qnam)
                    dtype, length, fmt = self.result_type_len_format(qdf["QVAL"])
                    rows.append(self.vlm_row(ds, "QNAM", qnam, qlabel, [], f"QNAM EQ {qnam}({qlabel})", "QVAL", dtype, length, fmt, "CRF", "Topic", ""))
                continue

            testcd_vars = [c for c in cols if c.endswith("TESTCD")]
            result_vars = [c for c in cols if c.endswith("ORRES")]
            for testcd_var in testcd_vars:
                test_var = testcd_var[:-6] + "TEST"
                if test_var not in cols:
                    continue
                for res_var in result_vars:
                    if not res_var.startswith(ds) and len(ds) <= 4:
                        # keep permissive for non-standard custom names
                        pass
                    group_cols = []
                    for suffix in ["CAT", "SCAT", "SPEC", "METHOD", "POS", "LAT"]:
                        gv = ds + suffix if len(ds) <= 4 else suffix
                        if gv in cols:
                            group_cols.append(gv)
                    base_cols = [testcd_var, test_var] + group_cols + [res_var]
                    tdf = df[[c for c in base_cols if c in cols]].copy()
                    tdf = tdf[tdf[res_var].apply(lambda x: safe_text(x) != "")]
                    if tdf.empty:
                        continue
                    groupby_cols = [testcd_var, test_var] + group_cols
                    for key, gdf in tdf.groupby(groupby_cols, dropna=False):
                        if not isinstance(key, tuple):
                            key = (key,)
                        values = dict(zip(groupby_cols, key))
                        testcd = safe_text(values.get(testcd_var)); test = safe_text(values.get(test_var))
                        if not testcd:
                            continue
                        grp_pairs = [(g, values.get(g)) for g in group_cols if safe_text(values.get(g))]
                        dtype, length, fmt = self.result_type_len_format(gdf[res_var])
                        wc_parts = [f"{testcd_var} EQ {testcd}({test})"] + [f"{g} EQ {safe_text(v)}" for g, v in grp_pairs]
                        rows.append(self.vlm_row(ds, testcd_var, testcd, test, grp_pairs, " and ".join(wc_parts), res_var, dtype, length, fmt, "CRF", "Topic", ""))
        return rows

    def adam_value_is_numeric(self, value):
        """Return True when a value can be treated as numeric for ADaM VLM routing."""
        txt = safe_text(value)
        if txt == "":
            return False
        try:
            float(txt)
            return True
        except Exception:
            return False

    def adam_series_all_numeric(self, series):
        """True when every nonblank value in a where-clause is numeric."""
        vals = [v for v in series.tolist() if safe_text(v) != ""]
        if not vals:
            return False
        return all(self.adam_value_is_numeric(v) for v in vals)

    def adam_series_has_nonnumeric(self, series):
        """True when any nonblank value in a where-clause is character/non-numeric."""
        vals = [v for v in series.tolist() if safe_text(v) != ""]
        if not vals:
            return False
        return any(not self.adam_value_is_numeric(v) for v in vals)

    def generate_adam_vlm_rows(self):
        """Generate ADaM VLM rows using only one result variable per where-clause.

        Rule:
        - If AVALC has any non-numeric values for a PARAMCD/PARAM where-clause,
          create only AVALC VLM with character type.
        - Otherwise, if AVAL has values and all values are numeric, create only
          AVAL VLM with numeric type.
        - Do not create both AVAL and AVALC for the same where-clause.
        """
        rows = []
        for ds, df in self.datasets.items():
            cols = list(df.columns)
            if "PARAMCD" not in cols or "PARAM" not in cols:
                continue
            if "AVAL" not in cols and "AVALC" not in cols:
                continue

            group_cols = [c for c in ["PARCAT1", "PARCAT2", "PARCAT3"] if c in cols]
            order_cols = [c for c in ["PARCAT1N", "PARCAT2N", "PARCAT3N", "PARAMN"] if c in cols]
            available_result_cols = [c for c in ["AVAL", "AVALC"] if c in cols]
            base_cols = ["PARAMCD", "PARAM"] + group_cols + order_cols + available_result_cols
            tdf = df[[c for c in base_cols if c in cols]].copy()

            groupby_cols = group_cols + ["PARAMCD", "PARAM"]
            for key, gdf in tdf.groupby(groupby_cols, dropna=False):
                if not isinstance(key, tuple):
                    key = (key,)
                values = dict(zip(groupby_cols, key))
                paramcd = safe_text(values.get("PARAMCD")); param = safe_text(values.get("PARAM"))
                if not paramcd:
                    continue

                res_var = ""
                dtype = ""
                length = ""
                fmt = ""

                # Character/non-numeric PARAM result takes priority and should be represented by AVALC only.
                if "AVALC" in cols:
                    avalc_nonblank = gdf[gdf["AVALC"].apply(lambda x: safe_text(x) != "")]
                    if not avalc_nonblank.empty and self.adam_series_has_nonnumeric(avalc_nonblank["AVALC"]):
                        res_var = "AVALC"
                        dtype = "char"
                        length = max([len(safe_text(v)) for v in avalc_nonblank["AVALC"].tolist()] or [1])
                        fmt = ""

                # Numeric PARAM result should be represented by AVAL only.
                if not res_var and "AVAL" in cols:
                    aval_nonblank = gdf[gdf["AVAL"].apply(lambda x: safe_text(x) != "")]
                    if not aval_nonblank.empty and self.adam_series_all_numeric(aval_nonblank["AVAL"]):
                        res_var = "AVAL"
                        dtype, length, fmt = self.result_type_len_format(aval_nonblank["AVAL"])
                        if dtype in {"char", "date", "datetime"}:
                            dtype = "float" if any("." in safe_text(v) for v in aval_nonblank["AVAL"].tolist()) else "num"
                            length = 8
                            fmt = fmt or "8"

                # Fallback: when AVAL is unavailable/blank but AVALC has values, keep AVALC as character.
                if not res_var and "AVALC" in cols:
                    avalc_nonblank = gdf[gdf["AVALC"].apply(lambda x: safe_text(x) != "")]
                    if not avalc_nonblank.empty:
                        res_var = "AVALC"
                        dtype = "char"
                        length = max([len(safe_text(v)) for v in avalc_nonblank["AVALC"].tolist()] or [1])
                        fmt = ""

                if not res_var:
                    continue

                grp_pairs = [(g, values.get(g)) for g in group_cols if safe_text(values.get(g))]
                wc_parts = [f"PARAMCD EQ {paramcd}({param})"] + [f"{g} EQ {safe_text(v)}" for g, v in grp_pairs]
                rows.append(self.vlm_row(ds, "PARAMCD", paramcd, param, grp_pairs, " and ".join(wc_parts), res_var, dtype, length, fmt, "Derived", "Topic", ""))
        return rows

    def vlm_row(self, ds, gvar, gval, glabel, grp_pairs, where, res, dtype, length, fmt, origin, role, comment):
        row = {c: "" for c in VLM_COLUMNS}
        row.update({
            "Dataset": safe_upper(ds), "Grouping Variable": safe_upper(gvar), "Group Value": safe_text(gval),
            "Group Label": safe_text(glabel), "Where Clause": safe_text(where), "Result Variable": safe_upper(res),
            "Length": length, "Type": dtype, "Format": fmt, "Origin": origin, "Role": role, "Comment": comment,
        })
        for i, (gv, gvval) in enumerate(grp_pairs[:4], start=1):
            row[f"Grouping Variable {i}"] = safe_upper(gv)
            row[f"Group Value {i}"] = safe_text(gvval)
        return row

    def validate_define_inputs(self):
        meta = self.current_editor_df()
        self.formats_df = self.formats_model.df.copy()
        self.vlm_df = self.vlm_model.df.copy()
        issues = []
        for _, r in meta.iterrows():
            ds = safe_upper(r.get("Dataset")); var = safe_upper(r.get("Variable"))
            origin = safe_upper(r.get("Origin")); comments = safe_text(r.get("Comments"))
            if origin == "ASSIGNED" and not comments:
                issues.append({"Severity": "ERROR", "Check": "Assigned Comment", "Dataset": ds, "Variable": var, "Message": "Origin is Assigned but Comments is blank. Assigned variables usually need CommentDef in define output."})
            if origin == "DERIVED" and not comments:
                issues.append({"Severity": "ERROR", "Check": "Derived Method", "Dataset": ds, "Variable": var, "Message": "Origin is Derived but Comments is blank. This should go to MethodDef / computational method."})
            if ds not in self.datasets:
                issues.append({"Severity": "WARNING", "Check": "Dataset Missing", "Dataset": ds, "Variable": var, "Message": "Dataset was not found in selected data folder."})
            elif var not in self.datasets[ds].columns:
                issues.append({"Severity": "WARNING", "Check": "Variable Missing", "Dataset": ds, "Variable": var, "Message": "Variable was not found in dataset; define length/type may rely on spec fallback."})
            if safe_text(r.get("Format")) and self.formats_df is not None and not self.formats_df.empty:
                fmt = safe_text(r.get("Format"))
                if fmt not in set(self.formats_df["Format"].astype(str)) and is_define_codelist_format(fmt, var, r.get("Type")) and not is_external_dictionary_format(fmt):
                    issues.append({"Severity": "WARNING", "Check": "Format Missing", "Dataset": ds, "Variable": var, "Message": f"Format '{fmt}' is assigned but not present in generated Formats table."})
        # Blank Decode is now acceptable and should not be flagged.
        self.validation_df = pd.DataFrame(issues, columns=["Severity", "Check", "Dataset", "Variable", "Message"])

        # Missing Decode for format is no longer a validation issue.
        try:
            if isinstance(self.validation_df, pd.DataFrame) and not self.validation_df.empty:
                joined = self.validation_df.astype(str).agg(" ".join, axis=1).str.lower()
                self.validation_df = self.validation_df.loc[~joined.str.contains("missing decode|decode is missing", regex=True)].copy()
        except Exception:
            pass

        # Remove false "Format Missing" warnings after normalizing domain-qualified format names.
        # Example: metadata has CM.UNIT, generated table may have UNIT after CT normalization.
        try:
            ct_map = getattr(self, "ct_alias_map", {}) if hasattr(self, "ct_alias_map") else {}
            generated_keys = set()
            if isinstance(getattr(self, "formats_df", None), pd.DataFrame) and not self.formats_df.empty and "Format" in self.formats_df.columns:
                for f in self.formats_df["Format"].tolist():
                    generated_keys.update(format_presence_keys(f, ct_map))

            if generated_keys and isinstance(self.validation_df, pd.DataFrame) and not self.validation_df.empty:
                keep_rows = []
                for _, vr in self.validation_df.iterrows():
                    chk = safe_text(vr.get("Check"))
                    msg = safe_text(vr.get("Message"))
                    if chk == "Format Missing" and "assigned but not present" in msg:
                        # Extract quoted format from message.
                        m = re.search(r"Format '([^']+)'", msg)
                        fmt = m.group(1) if m else ""
                        keys = format_presence_keys(fmt, ct_map)
                        if keys & generated_keys:
                            keep_rows.append(False)
                        else:
                            keep_rows.append(True)
                    else:
                        keep_rows.append(True)
                self.validation_df = self.validation_df.loc[keep_rows].copy()
        except Exception:
            pass


        self.validation_model.set_df(self.validation_df)
        n_err = len(self.validation_df[self.validation_df["Severity"] == "ERROR"]) if not self.validation_df.empty else 0
        n_warn = len(self.validation_df[self.validation_df["Severity"] == "WARNING"]) if not self.validation_df.empty else 0
        if hasattr(self, "validation_summary_label"):
            self.validation_summary_label.setText(f"Errors - {n_err}, Warnings - {n_warn}")
            if n_err > 0:
                self.validation_summary_label.setStyleSheet("background-color: #ffe6e6; color: #8a0000; border: 1px solid #cc0000; border-radius: 8px; padding: 6px; font-weight: bold;")
            elif n_warn > 0:
                self.validation_summary_label.setStyleSheet("background-color: #fff4cc; color: #7a4a00; border: 1px solid #d6a400; border-radius: 8px; padding: 6px; font-weight: bold;")
            else:
                self.validation_summary_label.setStyleSheet("background-color: #e7f7e7; color: #155724; border: 1px solid #4f9e4f; border-radius: 8px; padding: 6px; font-weight: bold;")
        self.validation_view.resizeColumnsToContents()
        self.tabs.setCurrentWidget(self.tab_validation)
        self.review_export_available = True
        self.set_status(f"Validation complete: {n_err} errors, {n_warn} warnings", "done" if n_err == 0 and n_warn == 0 else "error" if n_err > 0 else "info")
        self.set_workflow_state("validated")
        return n_err == 0

    def generate_define_xml(self):
        self.set_buttons_busy(True)
        try:
            self.set_status("Preparing define.xml inputs", "running")
            self.editor_df = self.current_editor_df()
            self.formats_df = self.formats_model.df.copy()
            self.vlm_df = self.vlm_model.df.copy()
            if self.editor_df.empty:
                self.set_status("Generate Define stopped: metadata is missing", "error")
                QtWidgets.QMessageBox.warning(self, "Metadata missing", "Build Metadata Editor first.")
                return

            self.set_status("Validating define inputs", "running")
            if not self.validate_define_inputs():
                self.set_status("Validation failed - define.xml not created", "error")
                QtWidgets.QMessageBox.critical(self, "Validation failed", "Fix blocking validation errors before creating define.xml.")
                return
            self.set_status("Validating define inputs - done", "done")

            out_dir = self.dataset_path or self.dataset_edit.text().strip()
            if not out_dir or not os.path.isdir(out_dir):
                self.set_status("Generate Define stopped: XPT folder missing", "error")
                QtWidgets.QMessageBox.warning(self, "XPT folder missing", "Please load/select the XPT folder first. define.xml will be written there.")
                return

            self.set_status("Writing define.xml to XPT folder", "running")
            writer = DefineXmlWriter(
                standard=self.standard_combo.currentText(),
                define_version=self.define_combo.currentText(),
                ig_version=self.ig_edit.text(),
                ct_version=self.ct_edit.text(),
                study_oid=self.study_oid_edit.text(),
                study_name=self.study_name_edit.text(),
                protocol=self.protocol_edit.text(),
                metadata_df=self.editor_df,
                formats_df=self.formats_df,
                vlm_df=self.vlm_df,
                out_dir=out_dir,
                include_acrf=self.include_acrf.isChecked(),
                include_rg=self.include_rg.isChecked(),
                odm_version=self.odm_edit.text(),
                study_description=self.study_desc_edit.text(),
                meddra_version=self.meddra_edit.text(),
                whodrug_version=self.whodrug_edit.text(),
                acrf_file=self.acrf_file_edit.text(),
                csdrg_file=self.csdrg_file_edit.text(),
                adrg_file=self.adrg_file_edit.text(),
                domains_df=self.domains_df,
                documents_df=self.documents_df,
                document_links_df=self.document_links_df,
            )
            out_xml = writer.write()
            self.define_generated = True
            self.review_export_available = True
            self.set_status(f"define.xml created: {out_xml}", "done")
            self.set_workflow_state("generated")
            QtWidgets.QMessageBox.information(self, "Define created", f"define.xml created:\n\n{out_xml}")
        except Exception:
            QtWidgets.QMessageBox.critical(self, "Define generation failed", traceback.format_exc())
            self.set_status("Define generation failed", "error")
        finally:
            # Never leave buttons frozen after Generate Define, even when validation/message boxes return early.
            if getattr(self, "define_generated", False):
                self.set_workflow_state("generated")
            elif getattr(self, "review_export_available", False):
                self.set_workflow_state("validated")
            elif getattr(self, "editor_df", pd.DataFrame()).empty:
                self.set_workflow_state("initial")
            elif getattr(self, "formats_df", pd.DataFrame()).empty:
                self.set_workflow_state("metadata")
            else:
                self.set_workflow_state("formats")


    def vlm_review_path(self):
        standard = safe_upper(self.standard_combo.currentText()) or "SDTM"
        fname = "ADAM_VLM.xlsx" if standard == "ADAM" else "SDTM_VLM.xlsx"
        return Path(__file__).resolve().parent / fname

    def export_vlm_xlsx(self):
        self.vlm_df = self.vlm_model.df.copy()
        if self.vlm_df.empty:
            QtWidgets.QMessageBox.warning(self, "No VLM", "Generate VLM first before exporting.")
            return
        out_path = self.vlm_review_path()
        try:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                self.vlm_df.to_excel(writer, sheet_name="ValueMetadata", index=False)
            self.set_status(f"VLM exported: {out_path}")
            QtWidgets.QMessageBox.information(self, "VLM exported", f"VLM exported for review/editing:\n\n{out_path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "VLM export failed", str(e))
            self.set_status("VLM export failed")

    def import_vlm_xlsx(self):
        default_path = self.vlm_review_path()
        path = str(default_path)
        if not default_path.exists():
            selected, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Import VLM workbook", str(default_path.parent), "Excel Files (*.xlsx);;All Files (*)")
            if not selected:
                return
            path = selected
        try:
            df = pd.read_excel(path, sheet_name=0, dtype=str).fillna("")
            # Add any missing expected columns and preserve expected order.
            for c in VLM_COLUMNS:
                if c not in df.columns:
                    df[c] = ""
            df = df[VLM_COLUMNS].copy()
            self.vlm_df = df
            self.vlm_model.set_df(self.vlm_df)
            tune_table_widths(self.vlm_view, {
                "Dataset": 90, "Grouping Variable": 125, "Group Value": 150,
                "Group Label": 180, "Where Clause": 260, "Result Variable": 140,
                "Length": 80, "Type": 90, "Format": 140,
                "Origin": 120, "Role": 120, "Comment": 300,
            }, max_width=220)
            self.tabs.setCurrentWidget(self.tab_vlm)
            self.set_status(f"VLM imported: {path}")
            QtWidgets.QMessageBox.information(self, "VLM imported", f"VLM imported:\n\n{path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "VLM import failed", str(e))
            self.set_status("VLM import failed")

    def export_review_xlsx(self, auto_path=None):
        if auto_path:
            out_path = auto_path
        else:
            out_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export review workbook", str(Path(self.dataset_path or str(Path.home())) / "define_review_inputs.xlsx"), "Excel Files (*.xlsx)")
            if not out_path:
                return
        if not out_path.lower().endswith(".xlsx"):
            out_path += ".xlsx"
        self.editor_df = self.current_editor_df()
        self.formats_df = self.formats_model.df.copy()
        self.vlm_df = self.vlm_model.df.copy()
        self.validation_df = self.validation_model.df.copy()
        Path(out_path).parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            self.editor_df.to_excel(writer, sheet_name="Metadata", index=False)
            self.normalized_spec.to_excel(writer, sheet_name="Spec_Data", index=False)
            self.dataset_metadata.to_excel(writer, sheet_name="XPT_Metadata", index=False)
            self.dataset_inventory.to_excel(writer, sheet_name="XPT_Inventory", index=False)
            self.formats_df.to_excel(writer, sheet_name="Format_CT", index=False)
            self.vlm_df.to_excel(writer, sheet_name="ValueMetadata", index=False)
            self.validation_df.to_excel(writer, sheet_name="Validation", index=False)
            self.domains_df.to_excel(writer, sheet_name="Domains", index=False)
        if not auto_path:
            QtWidgets.QMessageBox.information(self, "Exported", f"Review workbook exported:\n\n{out_path}")
        self.set_status(f"Review workbook exported: {out_path}")



    def set_workflow_state(self, stage="initial"):
        """Enable buttons by workflow stage."""
        try:
            for b in [
                self.btn_load_spec,
                self.btn_refresh_formats,
                self.btn_validate,
                self.btn_define,
                self.btn_export,
            ]:
                b.setEnabled(False)

            self.btn_load_spec.setEnabled(True)

            if stage in {"metadata", "formats", "validated", "generated"}:
                self.btn_refresh_formats.setEnabled(True)

            if stage in {"formats", "validated", "generated"}:
                self.btn_validate.setEnabled(True)
                # Generate Define also runs validation internally, so it should not stay frozen
                # after Format/VLM generation.
                self.btn_define.setEnabled(True)

            if stage in {"validated", "generated"} or getattr(self, "review_export_available", False):
                self.btn_export.setEnabled(True)
        except Exception:
            pass

    def load_spec_data_and_build_metadata(self):
        """Combined button action: Load Spec, Load Data, Build Metadata Editor."""
        self.review_export_available = False
        self.define_generated = False
        self.set_buttons_busy(True)
        ok = False
        try:
            self.set_status("Loading SharePoint spec", "running")
            self.load_spec()
            self.set_status("Loading SharePoint spec - done", "done")

            self.set_status("Loading XPT data", "running")
            self.load_data()
            self.set_status("Loading XPT data - done", "done")

            self.set_status("Building metadata", "running")
            self.build_metadata_editor()
            self.set_status("Building metadata - done", "done")
            ok = True
        except Exception as e:
            self.set_status(f"Load Spec and Data failed: {e}", "error")
            QtWidgets.QMessageBox.critical(self, "Load Spec and Data failed", traceback.format_exc())
        finally:
            self.set_workflow_state("metadata" if ok else "initial")

    def generate_formats_and_vlm(self):
        """Combined button action: Generate/Refresh Formats and Generate VLM."""
        self.review_export_available = False
        self.define_generated = False
        self.set_buttons_busy(True)
        ok = False
        try:
            self.set_status("Generating CT / formats", "running")
            self.generate_formats()
            self.set_status("Generating CT / formats - done", "done")

            self.set_status("Generating VLM", "running")
            self.generate_vlm()
            self.set_status("Generating VLM - done", "done")
            ok = True
        except Exception as e:
            self.set_status(f"Generate Format and VLM failed: {e}", "error")
            QtWidgets.QMessageBox.critical(self, "Generate Format and VLM failed", traceback.format_exc())
        finally:
            self.set_workflow_state("formats" if ok else "metadata")


    def get_loaded_ct_dataframe(self):
        """Return the currently loaded CDISC CT/browser dataframe, if present."""
        candidates = [
            "ct_df", "cdisc_ct_df", "ct_browser_df", "ct_package_df",
            "ct_results_df", "library_ct_df"
        ]
        for name in candidates:
            obj = getattr(self, name, None)
            if isinstance(obj, pd.DataFrame) and not obj.empty:
                return obj

        # Try common table model names.
        for name in ["ct_model", "cdisc_ct_model", "ct_browser_model"]:
            model = getattr(self, name, None)
            df = getattr(model, "df", None)
            if isinstance(df, pd.DataFrame) and not df.empty:
                return df

        return pd.DataFrame()


    def get_work_area_dir(self):
        """Return the XPT/output folder used for define.xml and VLM sidecar files."""
        candidates = []
        try:
            candidates.append(safe_text(self.dataset_edit.text()))
        except Exception:
            pass
        try:
            candidates.append(safe_text(getattr(self, "dataset_path", "")))
        except Exception:
            pass

        for folder in candidates:
            if folder and Path(folder).exists():
                return Path(folder)

        if getattr(self, "spec_path", ""):
            return Path(self.spec_path).resolve().parent
        return Path(__file__).resolve().parent

    def get_vlm_excel_path(self):
        """Return SDTM_VLM.xlsx or ADAM_VLM.xlsx in the active XPT/output folder."""
        std = safe_upper(self.standard_combo.currentText()) if hasattr(self, "standard_combo") else "SDTM"
        prefix = "ADAM" if std == "ADAM" else "SDTM"
        return self.get_work_area_dir() / f"{prefix}_VLM.xlsx"

    def export_vlm_excel(self):
        """Export current VLM grid to SDTM_VLM.xlsx / ADAM_VLM.xlsx for user editing."""
        try:
            if hasattr(self, "vlm_model") and self.vlm_model is not None:
                df = self.vlm_model.df.copy()
            elif isinstance(getattr(self, "vlm_df", None), pd.DataFrame):
                df = self.vlm_df.copy()
            else:
                df = pd.DataFrame(columns=VLM_COLUMNS)

            for col in VLM_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            ordered = VLM_COLUMNS + [c for c in df.columns if c not in VLM_COLUMNS]
            df = df[ordered]

            out_path = self.get_vlm_excel_path()
            out_path.parent.mkdir(parents=True, exist_ok=True)

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="VLM", index=False)

            self.vlm_df = df.copy()
            if hasattr(self, "vlm_model") and self.vlm_model is not None:
                self.vlm_model.set_df(self.vlm_df)

            self.set_status(f"VLM exported: {out_path}") if hasattr(self, "set_status") else None
            QtWidgets.QMessageBox.information(self, "Export VLM", f"VLM exported successfully:\n\n{out_path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export VLM failed", str(e))

    def import_vlm_excel(self):
        """Import edited SDTM_VLM.xlsx / ADAM_VLM.xlsx back into the VLM grid."""
        try:
            default_path = self.get_vlm_excel_path()
            file_path = default_path

            if not file_path.exists():
                selected, _ = QtWidgets.QFileDialog.getOpenFileName(
                    self,
                    "Select VLM Excel file",
                    str(default_path.parent),
                    "Excel Files (*.xlsx *.xlsm *.xls)"
                )
                if not selected:
                    return
                file_path = Path(selected)

            df = pd.read_excel(file_path, sheet_name="VLM", dtype=str).fillna("")

            for col in VLM_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            ordered = VLM_COLUMNS + [c for c in df.columns if c not in VLM_COLUMNS]
            df = df[ordered]

            self.vlm_df = df.copy()

            if hasattr(self, "vlm_model") and self.vlm_model is not None:
                self.vlm_model.set_df(self.vlm_df)
            else:
                self.vlm_model = PandasTableModel(self.vlm_df, editable_columns=set(VLM_COLUMNS))
                self.vlm_table.setModel(self.vlm_model)

            try:
                tune_table_widths(
                    self.vlm_table,
                    preferred={"Where Clause": 220, "Comment": 260, "Group Label": 180},
                    max_width=220
                )
            except Exception:
                pass

            self.set_status(f"VLM imported: {file_path}") if hasattr(self, "set_status") else None
            QtWidgets.QMessageBox.information(self, "Import VLM", f"VLM imported successfully:\n\n{file_path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Import VLM failed", str(e))

    # Compatibility names used by some earlier button wiring
    def export_vlm_xlsx(self):
        return self.export_vlm_excel()

    def import_vlm_xlsx(self):
        return self.import_vlm_excel()


def main():
    app = QtWidgets.QApplication(sys.argv)
    win = DefineStudio()
    win.showMaximized()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
