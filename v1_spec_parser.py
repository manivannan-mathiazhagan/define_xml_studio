"""
Define.xml Studio v1 - Spec Parser and Dataset Comparison Backend

Purpose
-------
Reads SDTM and ADaM specification workbooks based on the current internal templates,
normalizes variable metadata, optionally scans SAS7BDAT/XPT files, compares actual
datasets against specifications, and writes a review workbook.

This is a backend foundation script. It does not generate Define-XML yet.
"""

from __future__ import annotations

import argparse
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

try:
    import pyreadstat
except Exception:
    pyreadstat = None


# -----------------------------
# Configuration
# -----------------------------

COMMON_VARIABLE_ALIASES = {
    "dataset": "Dataset",
    "variable": "Variable",
    "label": "Label",
    "id var": "ID Var",
    "keep": "Keep",
    "type": "Type",
    "len": "Len",
    "control or format": "Control or Format",
    "term": "Terms",
    "terms": "Terms",
    "core": "Core",
    "role": "Role",
    "origin": "Origin",
}

SDTM_NON_DOMAIN_SHEETS = {
    "ReadMe",
    "STATUS",
    "Domains",
    "Blank Spec",
    "Source Data",
    "SUPPQUAL",
    "SUPP_TEMP",
    "Formats",
    "ValueMetadata",
    "DEV_FORMATS",
    "QC_FORMATS",
    "Lookups",
}

ADAM_NON_DOMAIN_SHEETS = {
    "ReadMe",
    "STATUS",
    "Domains",
    "Valuemetadata",
    "Formats",
    "DEV_FORMATS",
    "QC_FORMATS",
    "Lookups",
    "Blank Spec",
}

SUPPORTED_EXTENSIONS = {".sas7bdat", ".xpt"}


@dataclass
class WorkbookParseResult:
    workbook_type: str
    domain_summary: pd.DataFrame
    variables: pd.DataFrame
    value_metadata: pd.DataFrame
    formats: pd.DataFrame
    issues: pd.DataFrame


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return COMMON_VARIABLE_ALIASES.get(text.lower(), text)


def row_is_empty(values: List[object]) -> bool:
    return all(v in (None, "") for v in values)


def worksheet_to_dataframe(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int = 1) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    if header_row > len(rows):
        return pd.DataFrame()

    header = [normalize_header(v) for v in rows[header_row - 1]]
    data = rows[header_row:]
    df = pd.DataFrame(data, columns=header)

    # Drop completely empty trailing columns/rows
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")
    return df


def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> Optional[int]:
    """
    Looks for the common variable spec header row.
    """
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = [normalize_header(v) for v in next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))]
        if "Dataset" in values and "Variable" in values and "Label" in values:
            return row_idx
    return None


def parse_domain_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, workbook_type: str) -> Tuple[pd.DataFrame, List[dict]]:
    issues: List[dict] = []
    header_row = find_header_row(ws)

    if header_row is None:
        issues.append(
            {
                "WorkbookType": workbook_type,
                "Sheet": ws.title,
                "IssueType": "Header",
                "Severity": "Error",
                "Message": "Could not detect variable-header row.",
            }
        )
        return pd.DataFrame(), issues

    df = worksheet_to_dataframe(ws, header_row=header_row)

    # Standardize expected columns
    expected_cols = [
        "Dataset",
        "Variable",
        "Label",
        "ID Var",
        "Keep",
        "Type",
        "Len",
        "Control or Format",
        "Terms",
        "Core",
        "Role",
        "Origin",
    ]
    for col in expected_cols:
        if col not in df.columns:
            df[col] = pd.NA

    df = df[expected_cols].copy()
    df["WorkbookType"] = workbook_type
    df["Sheet"] = ws.title
    df["VariableOrder"] = range(1, len(df) + 1)

    # Remove rows where Dataset and Variable are both blank
    df = df[~(df["Dataset"].isna() & df["Variable"].isna())].copy()

    # Basic issues
    missing_dataset = df["Dataset"].isna() | (df["Dataset"].astype(str).str.strip() == "")
    missing_variable = df["Variable"].isna() | (df["Variable"].astype(str).str.strip() == "")

    for idx in df.index[missing_dataset]:
        issues.append(
            {
                "WorkbookType": workbook_type,
                "Sheet": ws.title,
                "IssueType": "Missing Dataset",
                "Severity": "Warning",
                "Message": f"Missing dataset name on row index {idx}.",
            }
        )

    for idx in df.index[missing_variable]:
        issues.append(
            {
                "WorkbookType": workbook_type,
                "Sheet": ws.title,
                "IssueType": "Missing Variable",
                "Severity": "Warning",
                "Message": f"Missing variable name on row index {idx}.",
            }
        )

    return df, issues


def parse_domains_sheet(wb: openpyxl.Workbook, sheet_name: str, workbook_type: str) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    df = worksheet_to_dataframe(ws, header_row=1)
    if df.empty:
        return df
    df["WorkbookType"] = workbook_type
    return df


def parse_support_sheet(wb: openpyxl.Workbook, sheet_name: str, workbook_type: str) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    df = worksheet_to_dataframe(ws, header_row=1)
    if df.empty:
        return df
    df["WorkbookType"] = workbook_type
    return df


def parse_spec_workbook(path: str, workbook_type: str) -> WorkbookParseResult:
    wb = openpyxl.load_workbook(path, data_only=False)

    if workbook_type.upper() == "SDTM":
        non_domain_sheets = SDTM_NON_DOMAIN_SHEETS
        domain_sheet_name = "Domains"
        value_meta_sheet = "ValueMetadata"
        formats_sheet = "Formats"
    else:
        non_domain_sheets = ADAM_NON_DOMAIN_SHEETS
        domain_sheet_name = "Domains"
        value_meta_sheet = "Valuemetadata"
        formats_sheet = "Formats"

    all_variables: List[pd.DataFrame] = []
    issues: List[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in non_domain_sheets:
            continue

        # Skip hidden technical sheets that start with underscore for now
        if sheet_name.startswith("_"):
            continue

        ws = wb[sheet_name]
        df_vars, sheet_issues = parse_domain_sheet(ws, workbook_type)
        if not df_vars.empty:
            all_variables.append(df_vars)
        issues.extend(sheet_issues)

    variables_df = pd.concat(all_variables, ignore_index=True) if all_variables else pd.DataFrame()
    domain_summary_df = parse_domains_sheet(wb, domain_sheet_name, workbook_type)
    value_meta_df = parse_support_sheet(wb, value_meta_sheet, workbook_type)
    formats_df = parse_support_sheet(wb, formats_sheet, workbook_type)
    issues_df = pd.DataFrame(issues)

    if not variables_df.empty:
        dupes = variables_df.duplicated(subset=["Dataset", "Variable", "WorkbookType"], keep=False)
        for _, row in variables_df.loc[dupes, ["Dataset", "Variable", "Sheet"]].drop_duplicates().iterrows():
            issues.append(
                {
                    "WorkbookType": workbook_type,
                    "Sheet": row["Sheet"],
                    "IssueType": "Duplicate Variable",
                    "Severity": "Warning",
                    "Message": f"Duplicate variable mapping found for {row['Dataset']}.{row['Variable']}.",
                }
            )
        issues_df = pd.DataFrame(issues)

    return WorkbookParseResult(
        workbook_type=workbook_type,
        domain_summary=domain_summary_df,
        variables=variables_df,
        value_metadata=value_meta_df,
        formats=formats_df,
        issues=issues_df,
    )


def read_dataset_file(path: Path) -> Tuple[pd.DataFrame, List[dict]]:
    """
    Returns metadata rows for a single dataset file.
    """
    issues: List[dict] = []
    rows: List[dict] = []

    if pyreadstat is None:
        issues.append(
            {
                "IssueType": "Dependency",
                "Severity": "Error",
                "Message": "pyreadstat is not installed, so dataset scanning cannot run.",
                "Path": str(path),
            }
        )
        return pd.DataFrame(), issues

    try:
        suffix = path.suffix.lower()
        if suffix == ".sas7bdat":
            df, meta = pyreadstat.read_sas7bdat(str(path), metadataonly=True)
        elif suffix == ".xpt":
            df, meta = pyreadstat.read_xport(str(path), metadataonly=True)
        else:
            return pd.DataFrame(), issues

        dataset_name = path.stem.upper()
        col_names = meta.column_names or []
        col_labels = meta.column_labels or []
        readstat_types = meta.readstat_variable_types or {}
        file_label = getattr(meta, "file_label", None)

        for idx, var in enumerate(col_names, start=1):
            rows.append(
                {
                    "Dataset": dataset_name,
                    "Variable": var,
                    "Label_Actual": col_labels[idx - 1] if idx - 1 < len(col_labels) else None,
                    "Type_Actual": readstat_types.get(var),
                    "VariableOrder_Actual": idx,
                    "FilePath": str(path),
                    "FileType": suffix.replace(".", "").upper(),
                    "DatasetLabel_Actual": file_label,
                }
            )

    except Exception as exc:
        issues.append(
            {
                "IssueType": "Read Error",
                "Severity": "Error",
                "Message": f"Failed to read dataset: {exc}",
                "Path": str(path),
            }
        )

    return pd.DataFrame(rows), issues


def scan_dataset_directory(data_dir: Optional[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not data_dir:
        return pd.DataFrame(), pd.DataFrame()

    base = Path(data_dir)
    if not base.exists():
        issues = pd.DataFrame(
            [
                {
                    "IssueType": "Path",
                    "Severity": "Error",
                    "Message": f"Dataset directory does not exist: {data_dir}",
                }
            ]
        )
        return pd.DataFrame(), issues

    metadata_frames: List[pd.DataFrame] = []
    issues: List[dict] = []

    for path in sorted(base.rglob("*")):
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
            df_meta, file_issues = read_dataset_file(path)
            if not df_meta.empty:
                metadata_frames.append(df_meta)
            issues.extend(file_issues)

    metadata_df = pd.concat(metadata_frames, ignore_index=True) if metadata_frames else pd.DataFrame()
    issues_df = pd.DataFrame(issues)
    return metadata_df, issues_df


def compare_spec_vs_data(spec_variables: pd.DataFrame, actual_metadata: pd.DataFrame) -> pd.DataFrame:
    if spec_variables.empty and actual_metadata.empty:
        return pd.DataFrame()

    spec = spec_variables.copy()
    actual = actual_metadata.copy()

    if spec.empty:
        actual["ComparisonStatus"] = "Actual Only"
        return actual

    if actual.empty:
        spec["ComparisonStatus"] = "Spec Only"
        return spec

    spec["Dataset"] = spec["Dataset"].astype(str).str.upper()
    spec["Variable"] = spec["Variable"].astype(str).str.upper()
    actual["Dataset"] = actual["Dataset"].astype(str).str.upper()
    actual["Variable"] = actual["Variable"].astype(str).str.upper()

    merged = spec.merge(
        actual,
        on=["Dataset", "Variable"],
        how="outer",
        suffixes=("_Spec", "_Actual"),
        indicator=True,
    )

    status_map = {
        "both": "Matched",
        "left_only": "Spec Only",
        "right_only": "Actual Only",
    }
    merged["ComparisonStatus"] = merged["_merge"].map(status_map)

    if "Label" in merged.columns and "Label_Actual" in merged.columns:
        merged["LabelMatch"] = (
            merged["Label"].fillna("").astype(str).str.strip()
            == merged["Label_Actual"].fillna("").astype(str).str.strip()
        )
    else:
        merged["LabelMatch"] = pd.NA

    if "Type" in merged.columns and "Type_Actual" in merged.columns:
        merged["TypeMatch"] = (
            merged["Type"].fillna("").astype(str).str.upper().str.strip()
            == merged["Type_Actual"].fillna("").astype(str).str.upper().str.strip()
        )
    else:
        merged["TypeMatch"] = pd.NA

    return merged.drop(columns=["_merge"])


def build_summary(
    sdtm_result: WorkbookParseResult,
    adam_result: WorkbookParseResult,
    actual_metadata: pd.DataFrame,
    comparison_df: pd.DataFrame,
) -> pd.DataFrame:
    rows = [
        {
            "Metric": "SDTM Domain Rows",
            "Value": len(sdtm_result.domain_summary),
        },
        {
            "Metric": "SDTM Variable Rows",
            "Value": len(sdtm_result.variables),
        },
        {
            "Metric": "SDTM Value Metadata Rows",
            "Value": len(sdtm_result.value_metadata),
        },
        {
            "Metric": "ADaM Domain Rows",
            "Value": len(adam_result.domain_summary),
        },
        {
            "Metric": "ADaM Variable Rows",
            "Value": len(adam_result.variables),
        },
        {
            "Metric": "ADaM Value Metadata Rows",
            "Value": len(adam_result.value_metadata),
        },
        {
            "Metric": "Actual Dataset Variable Rows",
            "Value": len(actual_metadata),
        },
        {
            "Metric": "Matched Spec/Data Rows",
            "Value": int((comparison_df.get("ComparisonStatus", pd.Series(dtype="object")) == "Matched").sum())
            if not comparison_df.empty
            else 0,
        },
        {
            "Metric": "Spec Only Rows",
            "Value": int((comparison_df.get("ComparisonStatus", pd.Series(dtype="object")) == "Spec Only").sum())
            if not comparison_df.empty
            else 0,
        },
        {
            "Metric": "Actual Only Rows",
            "Value": int((comparison_df.get("ComparisonStatus", pd.Series(dtype="object")) == "Actual Only").sum())
            if not comparison_df.empty
            else 0,
        },
    ]
    return pd.DataFrame(rows)


def write_excel_report(
    output_path: str,
    summary_df: pd.DataFrame,
    sdtm_result: WorkbookParseResult,
    adam_result: WorkbookParseResult,
    actual_metadata: pd.DataFrame,
    comparison_df: pd.DataFrame,
    dataset_scan_issues: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        sdtm_result.domain_summary.to_excel(writer, sheet_name="SDTM_Domains", index=False)
        sdtm_result.variables.to_excel(writer, sheet_name="SDTM_Variables", index=False)
        sdtm_result.value_metadata.to_excel(writer, sheet_name="SDTM_ValueMetadata", index=False)
        sdtm_result.formats.to_excel(writer, sheet_name="SDTM_Formats", index=False)

        adam_result.domain_summary.to_excel(writer, sheet_name="ADAM_Domains", index=False)
        adam_result.variables.to_excel(writer, sheet_name="ADAM_Variables", index=False)
        adam_result.value_metadata.to_excel(writer, sheet_name="ADAM_ValueMetadata", index=False)
        adam_result.formats.to_excel(writer, sheet_name="ADAM_Formats", index=False)

        actual_metadata.to_excel(writer, sheet_name="Actual_Datasets", index=False)
        comparison_df.to_excel(writer, sheet_name="Spec_vs_Data", index=False)

        all_issues = pd.concat(
            [df for df in [sdtm_result.issues, adam_result.issues, dataset_scan_issues] if not df.empty],
            ignore_index=True,
        ) if any(not df.empty for df in [sdtm_result.issues, adam_result.issues, dataset_scan_issues]) else pd.DataFrame()

        all_issues.to_excel(writer, sheet_name="Issues", index=False)

        wb = writer.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for column_cells in ws.columns:
                max_length = 0
                col_letter = column_cells[0].column_letter
                for cell in column_cells[:200]:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                ws.column_dimensions[col_letter].width = min(max(12, max_length + 2), 40)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Define.xml Studio v1 spec parser")
    parser.add_argument("--sdtm-spec", required=True, help="Path to SDTM specification workbook")
    parser.add_argument("--adam-spec", required=True, help="Path to ADaM specification workbook")
    parser.add_argument("--data-dir", required=False, default=None, help="Optional directory containing SAS7BDAT/XPT files")
    parser.add_argument("--output", required=True, help="Path to output Excel report")
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    sdtm_result = parse_spec_workbook(args.sdtm_spec, "SDTM")
    adam_result = parse_spec_workbook(args.adam_spec, "ADAM")

    spec_variables = pd.concat(
        [df for df in [sdtm_result.variables, adam_result.variables] if not df.empty],
        ignore_index=True,
    ) if any(not df.empty for df in [sdtm_result.variables, adam_result.variables]) else pd.DataFrame()

    actual_metadata, dataset_scan_issues = scan_dataset_directory(args.data_dir)
    comparison_df = compare_spec_vs_data(spec_variables, actual_metadata)
    summary_df = build_summary(sdtm_result, adam_result, actual_metadata, comparison_df)

    output_path = str(Path(args.output).resolve())
    write_excel_report(
        output_path=output_path,
        summary_df=summary_df,
        sdtm_result=sdtm_result,
        adam_result=adam_result,
        actual_metadata=actual_metadata,
        comparison_df=comparison_df,
        dataset_scan_issues=dataset_scan_issues,
    )

    print(f"Report written to: {output_path}")


if __name__ == "__main__":
    main()
